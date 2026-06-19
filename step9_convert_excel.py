"""
convert_excel.py — Converts each Excel Data sheet to a CSV file in data/.
Run this after filling in data in the Excel files.
"""

import csv
from pathlib import Path
import openpyxl

DATA_DIR = Path(__file__).parent / "data"

EXCEL_TO_CSV = {
    "cell_site.xlsx":           "cell_site.csv",
    "active_components.xlsx":   "active_components.csv",
    "passive_components.xlsx":  "passive_components.csv",
    "infrastructure.xlsx":      "infrastructure.csv",
}


def is_empty_row(row):
    return all(cell.value is None or str(cell.value).strip() == "" for cell in row)


def convert(xlsx_name, csv_name):
    xlsx_path = DATA_DIR / xlsx_name
    csv_path  = DATA_DIR / csv_name

    if not xlsx_path.exists():
        print(f"SKIP: {xlsx_name} not found.")
        return

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if "Data" not in wb.sheetnames:
        print(f"SKIP: {xlsx_name} has no 'Data' sheet.")
        return

    ws = wb["Data"]
    rows = list(ws.iter_rows(values_only=False))

    if not rows:
        print(f"SKIP: {xlsx_name} Data sheet is empty.")
        return

    headers = [cell.value for cell in rows[0]]
    data_rows = [row for row in rows[1:] if not is_empty_row(row)]

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in data_rows:
            writer.writerow([("" if cell.value is None else cell.value) for cell in row])

    print(f"Written {len(data_rows)} rows to {csv_path.name}")


if __name__ == "__main__":
    for xlsx_name, csv_name in EXCEL_TO_CSV.items():
        convert(xlsx_name, csv_name)
    print("\nConversion complete.")
