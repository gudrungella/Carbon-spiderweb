"""
validate.py — Reads each Excel file, runs the full validation pipeline, and POSTs valid rows
to telecom_api.  Requires the server to be running:
    uvicorn main:app --host 0.0.0.0 --port 3000 --reload  (from inside telecom_api/)

Severity levels
  ERROR   — row rejected; must be corrected and resubmitted; no further processing
  WARNING — row proceeds; data quality score reduced; flagged field routed to enrichment
  INFO    — row proceeds unchanged; logged for audit
"""

from dataclasses import dataclass, field as dc_field
from functools import partial
from pathlib import Path
import sys

import openpyxl
import requests

BASE_URL = "http://localhost:3000"
DATA_DIR = Path(__file__).parent / "data"

# ── Enum definitions (mirrors swagger.yaml) ───────────────────────────────────

SITE_TYPES            = ["macro","micro","pico","femto","rooftop","indoor","greenfield","small","medium","large"]
NETWORK_TYPES         = ["Access","metro","aggregation","Backbone","core"]

# Variant spellings / capitalisations that map to a canonical network type
_NETWORK_TYPE_NORM: dict = {
    "metro":    "metro",
    "Metro":    "metro",
    "access":   "Access",
    "Access":   "Access",
    "backbone": "Backbone",
    "Backbone": "Backbone",
    "backcobe": "Backbone",   # typo fix
    "#N/A":     "unknown",
    "NA":       "unknown",
}

def normalise_network_type(nt) -> str:
    """Return the canonical network_type string.  Blank / #N/A / NA → 'unknown'."""
    s = str(nt or "").strip()
    if not s or s in ("#N/A", "NA"):
        return "unknown"
    return _NETWORK_TYPE_NORM.get(s, s)
OPERATIONAL_STATUSES  = ["active","planned","under_construction","decommissioned","maintenance"]
ELECTRICITY_SOURCES   = ["grid","renewable","solar","wind","none"]
FUEL_TYPES            = ["bensin","diesel","alternative","none","other"]
REFRIGERANT_TYPES     = ["none","R-11","R-12","R-22","R-123","R-134a","R-410A",
                         "R-245fa","R-32","R-1233zd(E)","R-1234yf","R-513A","other"]
ACTIVE_TYPES          = ["generator","cooling","fire_suppression",
                         "network_equipment","radio_equipment","power_equipment"]
ACTIVE_SUBTYPES       = [
    # generator
    "standby_power_generator","prime_power_generator","portable_industrial_generator",
    "inverter_generator","container_sized_generator",
    # cooling
    "water_cooled_systems","air_cooled_systems","industrial_chillers",
    "evaporative_cooling_systems","hybrid_system","specialized_cooling",
    # fire_suppression
    "gas_suppression","sprinkler_system","foam_system",
    # network_equipment
    "aggregation_router","chassis","edge_platform","firewall","gateway",
    "isam","memory_card","router","switch","sfp",
    # radio_equipment
    "air","cellular_modem","radio_unit","baseband_unit","base_station","antenna","wls",
    # power_equipment
    "sla_vrla","bci","din_en","dc_holder","power_distribution","light","camera","sensor",
]
POWER_SOURCES         = ["battery","electricity","fuel","refrigerant","other"]
PASSIVE_TYPES         = ["fiber_cable","electrical_cables","COAX","splitters","shelters",
                         "cabinets","plugs","fencing","steel","aluminum","plastic"]
INFRA_TYPES           = ["tower","mast","rooftop_mount","pole","building","underground","container",
                         "real estate","manhole","concrete","ducts & pipes"]
TECHNOLOGY_TYPES      = ["2G","3G","4G","5G","6G","ADSL","DSL","FTTH","FTTC","GPON","HFC",
                         "LoRa","MPLS","MEC","Network slicing","PON","POTS/PSTN","RFID",
                         "SD-WAN","VoIP","WLAN","XPON"]
INSTALL_METHODS       = ["traditional_excavation_green","microtrenching_green","milling_green",
                         "plowing_green","traditional_excavation_urban","microtrenching_urban",
                         "milling_urban","plowing_urban","truck","van","ship","airplane","crane","other"]
MAINT_METHODS         = ["physical_security_services","decommissioning","replacements",
                         "environmental_monitoring","failure_monitoring","software_updates","other"]
INSTALL_UNITS         = ["km","kW","kWh","L","m3","m2"]
POWER_UNITS           = ["kW","W","kWh","L","m3"]
INSTALL_EMIT_UNITS    = ["kgCO2eq/km","kgCO2eq/kW","kgCO2eq/kWh","kgCO2eq/L","kgCO2eq/m3","kgCO2eq/m2"]
POWER_EMIT_UNITS      = ["kgCO2eq/kW","kgCO2eq/W","kgCO2eq/kWh","kgCO2eq/L","kgCO2eq/m3"]
EMISSION_UNIT         = ["kgCO2eq/unit"]

MEASURED_ELECTRICITY_UNITS = ["kWh"]
MEASURED_FUEL_UNITS        = ["L", "m3", "kg"]
MEASURED_REFRIGERANT_UNITS = ["kg", "m3"]
ELECTRICITY_EF_UNITS       = ["kgCO2eq/kWh"]
FUEL_EF_UNITS              = ["kgCO2eq/L", "kgCO2eq/m3", "kgCO2eq/kg"]
REFRIGERANT_EF_UNITS       = ["kgCO2eq/m3", "kgCO2eq/kg"]

# Maps each measured consumption to its unit field, EF unit field, and compatible pairings.
# The denominator of the EF unit must match the consumption unit.
_MEASURED_EF_UNIT_COMPAT = [
    ("measured_electricity_unit", "electricity_emission_factor_unit",
     {"kWh": "kgCO2eq/kWh"}),
    ("measured_fuel_unit",        "fuel_emission_factor_unit",
     {"L": "kgCO2eq/L", "m3": "kgCO2eq/m3", "kg": "kgCO2eq/kg"}),
    ("measured_refrigerant_unit", "refrigerant_emission_factor_unit",
     {"kg": "kgCO2eq/kg", "m3": "kgCO2eq/m3"}),
]

# ── Data classes ──────────────────────────────────────────────────────────────

@dataclass
class Issue:
    level:   str   # ERROR | WARNING | INFO
    field:   str
    message: str

@dataclass
class RowResult:
    row_num:           int
    status:            str   # REJECTED | OK | WARNED | API_ERROR
    issues:            list  = dc_field(default_factory=list)
    quality:           int   = 100
    api_ids:           dict  = dc_field(default_factory=dict)
    enrichment_fields: list  = dc_field(default_factory=list)

# ── Helpers ───────────────────────────────────────────────────────────────────

def present(v):
    return v is not None and str(v).strip() != ""

def to_num(v):
    try:    return float(v)
    except: return None

def compute_quality(issues):
    return max(0, 100 - 10 * sum(1 for i in issues if i.level == "WARNING"))

# ── Generic check functions ───────────────────────────────────────────────────

def chk_mandatory(row, fields):
    return [Issue("ERROR", f, f"Required field '{f}' is missing")
            for f in fields if not present(row.get(f))]

def chk_enums(row, enum_map):
    issues = []
    for fname, allowed in enum_map.items():
        v = row.get(fname)
        if present(v) and str(v) not in allowed:
            issues.append(Issue("WARNING", fname, f"'{v}' is not a valid value for '{fname}'"))
    return issues

def chk_num_unit_pairs(row, pairs):
    issues = []
    for nf, uf in pairs:
        has_n = present(row.get(nf))
        has_u = present(row.get(uf))
        if has_n and not has_u:
            issues.append(Issue("ERROR", uf,
                                f"'{uf}' is required when '{nf}' is provided"))
        elif has_u and not has_n:
            issues.append(Issue("INFO", nf,
                                f"'{uf}' is provided but '{nf}' is empty"))
    return issues

def chk_positive(row, fields):
    issues = []
    for f in fields:
        v = row.get(f)
        if present(v):
            n = to_num(v)
            if n is None:
                issues.append(Issue("ERROR", f, f"'{f}' must be a number, got '{v}'"))
            elif n < 0:
                issues.append(Issue("ERROR", f, f"'{f}' must be >= 0, got {n}"))
    return issues

def chk_life_time(row):
    v = row.get("life_time")
    if not present(v):
        return [Issue("WARNING", "life_time", "life_time is missing — will be filled by enrichment")]
    n = to_num(v)
    if n is None:
        return [Issue("ERROR", "life_time", f"life_time must be a number, got '{v}'")]
    if not (0 <= n <= 200):
        return [Issue("ERROR", "life_time", f"life_time must be between 0 and 200, got {n}")]
    if n < 1:
        return [Issue("WARNING", "life_time", f"life_time is unusually short ({n} years)")]
    return []

def chk_power_fields(row):
    has_qty  = present(row.get("power_quantity"))
    has_idle = present(row.get("power_idle"))
    has_max  = present(row.get("power_max"))
    if not has_qty and not (has_idle and has_max):
        return [Issue("WARNING", "power_quantity",
                      "Active component should have either power_quantity or both power_idle and power_max")]
    return []

def chk_install_maint_complete(row):
    keys = ["installation_method","installation_quantity",
            "maintenance_method","maintenance_quantity"]
    if all(not present(row.get(k)) for k in keys):
        return [Issue("WARNING", "installation_method",
                      "All installation and maintenance fields are empty")]
    return []

def chk_embodied_endoflife(row):
    he = present(row.get("production_emissions"))
    hl = present(row.get("endoflife_emissions"))
    if he and not hl:
        return [Issue("INFO", "endoflife_emissions",
                      "production_emissions provided but endoflife_emissions is missing")]
    if hl and not he:
        return [Issue("INFO", "production_emissions",
                      "endoflife_emissions provided but production_emissions is missing")]
    return []

def chk_ef_without_qty(row):
    issues = []
    for ef, qty in [("installation_emission_factor", "installation_quantity"),
                    ("maintenance_emission_factor",  "maintenance_quantity")]:
        if present(row.get(ef)) and not present(row.get(qty)):
            issues.append(Issue("INFO", qty, f"'{ef}' provided but '{qty}' is missing"))
    return issues

# ── Shared enum/pair config ───────────────────────────────────────────────────

_COMMON_ENUM = {
    "installation_method":              INSTALL_METHODS,
    "installation_unit":                INSTALL_UNITS,
    "installation_emission_factor_unit":INSTALL_EMIT_UNITS,
    "maintenance_method":               MAINT_METHODS,
    "maintenance_unit":                 INSTALL_UNITS,
    "maintenance_emission_factor_unit": INSTALL_EMIT_UNITS,
    "production_emissions_unit":          EMISSION_UNIT,
    "endoflife_emissions_unit":         EMISSION_UNIT,
}

_COMMON_PAIRS = [
    ("installation_quantity",        "installation_unit"),
    ("installation_emission_factor", "installation_emission_factor_unit"),
    ("maintenance_quantity",         "maintenance_unit"),
    ("maintenance_emission_factor",  "maintenance_emission_factor_unit"),
    ("production_emissions",           "production_emissions_unit"),
    ("endoflife_emissions",          "endoflife_emissions_unit"),
]

# ── Schema validators ─────────────────────────────────────────────────────────

def validate_cell_site(row):
    row = dict(row)
    if "network_type" in row:
        row["network_type"] = normalise_network_type(row["network_type"])
    issues = (
        chk_mandatory(row, ["site_type", "network_type", "country"]) +
        chk_enums(row, {
            "site_type":                     SITE_TYPES,
            "network_type":                  NETWORK_TYPES,
            "operational_status":            OPERATIONAL_STATUSES,
            "electricity_source":            ELECTRICITY_SOURCES,
            "fuel_type":                     FUEL_TYPES,
            "refrigerant_type":              REFRIGERANT_TYPES,
            "measured_electricity_unit":     MEASURED_ELECTRICITY_UNITS,
            "measured_fuel_unit":            MEASURED_FUEL_UNITS,
            "measured_refrigerant_unit":     MEASURED_REFRIGERANT_UNITS,
            "electricity_emission_factor_unit": ELECTRICITY_EF_UNITS,
            "fuel_emission_factor_unit":        FUEL_EF_UNITS,
            "refrigerant_emission_factor_unit": REFRIGERANT_EF_UNITS,
        })
    )
    pr = to_num(row.get("per_rented"))
    if pr is not None and not (0 <= pr <= 100):
        issues.append(Issue("ERROR", "per_rented",
                            f"per_rented must be between 0 and 100, got {pr}"))
    for meas_unit_field, ef_unit_field, compat in _MEASURED_EF_UNIT_COMPAT:
        meas_unit = str(row.get(meas_unit_field) or "").strip()
        ef_unit   = str(row.get(ef_unit_field)   or "").strip()
        if meas_unit and ef_unit:
            expected = compat.get(meas_unit)
            if expected and ef_unit != expected:
                issues.append(Issue("ERROR", ef_unit_field,
                    f"{ef_unit_field} '{ef_unit}' is inconsistent with "
                    f"{meas_unit_field} '{meas_unit}' — expected '{expected}'"))
    return issues

def validate_active_component(row, cell_site_lookup):
    issues = (
        chk_mandatory(row, ["power_source"]) +
        chk_enums(row, {
            **_COMMON_ENUM,
            "active_type":                      ACTIVE_TYPES,
            "active_subtype":                   ACTIVE_SUBTYPES,
            "technology_type":                  TECHNOLOGY_TYPES,
            "power_source":                     POWER_SOURCES,
            "power_unit":                       POWER_UNITS,
            "power_idle_unit":                  POWER_UNITS,
            "power_max_unit":                   POWER_UNITS,
            "power_source_emission_factor_unit":POWER_EMIT_UNITS,
        }) +
        chk_life_time(row) +
        chk_positive(row, ["power_quantity", "power_idle", "power_max",
                            "production_emissions", "endoflife_emissions"]) +
        chk_num_unit_pairs(row, _COMMON_PAIRS + [
            ("power_quantity",              "power_unit"),
            ("power_idle",                  "power_idle_unit"),
            ("power_max",                   "power_max_unit"),
            ("power_source_emission_factor","power_source_emission_factor_unit"),
        ]) +
        chk_power_fields(row) +
        chk_install_maint_complete(row) +
        chk_embodied_endoflife(row) +
        chk_ef_without_qty(row)
    )

    # active_type and active_subtype both absent
    if not present(row.get("active_type")) and not present(row.get("active_subtype")):
        issues.append(Issue("WARNING", "active_type",
                            "Neither active_type nor active_subtype is provided"))

    # power_max >= power_idle
    idle = to_num(row.get("power_idle"))
    pmax = to_num(row.get("power_max"))
    if idle is not None and pmax is not None and pmax < idle:
        issues.append(Issue("ERROR", "power_max",
                            f"power_max ({pmax}) must be >= power_idle ({idle})"))

    # power_source=fuel but cell site has no fuel_type
    cs_id_raw = row.get("cell_site_id")
    if present(cs_id_raw) and str(row.get("power_source", "")).strip() == "fuel":
        site = cell_site_lookup.get(_cs_id_key(cs_id_raw))
        if site and (not present(site.get("fuel_type")) or site.get("fuel_type") == "none"):
            issues.append(Issue("WARNING", "power_source",
                                "power_source is 'fuel' but the linked cell site has no fuel_type"))

    return issues

def validate_passive_component(row):
    return (
        chk_mandatory(row, ["passive_type"]) +
        chk_enums(row, {
            **_COMMON_ENUM,
            "passive_type":    PASSIVE_TYPES,
            "technology_type": TECHNOLOGY_TYPES,
        }) +
        chk_life_time(row) +
        chk_positive(row, ["production_emissions", "endoflife_emissions"]) +
        chk_num_unit_pairs(row, _COMMON_PAIRS) +
        chk_install_maint_complete(row) +
        chk_embodied_endoflife(row) +
        chk_ef_without_qty(row)
    )

def validate_infrastructure(row):
    row = dict(row)
    if "network_type" in row:
        row["network_type"] = normalise_network_type(row["network_type"])
    return (
        chk_mandatory(row, ["infrastructure_type", "network_type"]) +
        chk_enums(row, {
            **_COMMON_ENUM,
            "infrastructure_type": INFRA_TYPES,
            "network_type":        NETWORK_TYPES,
        }) +
        chk_life_time(row) +
        chk_positive(row, ["production_emissions", "endoflife_emissions"]) +
        chk_num_unit_pairs(row, _COMMON_PAIRS) +
        chk_install_maint_complete(row) +
        chk_embodied_endoflife(row) +
        chk_ef_without_qty(row)
    )

# ── Excel helpers ─────────────────────────────────────────────────────────────

def load_sheet(xlsx_name):
    """Return (headers, data_rows) or (None, None) if unusable."""
    path = DATA_DIR / xlsx_name
    if not path.exists():
        return None, None
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Data" not in wb.sheetnames:
        return None, None
    ws = wb["Data"]
    all_rows = list(ws.iter_rows(values_only=False))
    if len(all_rows) < 2:
        return None, None
    headers   = [c.value for c in all_rows[0]]
    data_rows = [r for r in all_rows[1:]
                 if any(c.value is not None and str(c.value).strip() for c in r)]
    return headers, data_rows

def to_dict(headers, row):
    d = {}
    for h, cell in zip(headers, row):
        v = cell.value
        if h == "cell_site_id" and v is not None:
            v = str(int(v)) if isinstance(v, float) and v == int(v) else str(v)
        d[h] = v.strip() if isinstance(v, str) else v
    return d

def collect_cell_site_ids(xlsx_name):
    """Collect all cell_site_id values from a file without full processing."""
    headers, rows = load_sheet(xlsx_name)
    if headers is None:
        return set()
    ids = set()
    for row in rows:
        rd = to_dict(headers, row)
        v = rd.get("cell_site_id")
        if present(v):
            n = to_num(v)
            ids.add(int(n) if n is not None else str(v).strip())
    return ids


def _cs_id_key(raw):
    """Normalise a cell_site_id value to int (if numeric) or str."""
    n = to_num(raw)
    return int(n) if n is not None else str(raw).strip()

# ── API helpers ───────────────────────────────────────────────────────────────

def check_server():
    try:
        requests.get(f"{BASE_URL}/cell-sites", timeout=3)
    except Exception:
        print(f"ERROR: Cannot connect to API at {BASE_URL}")
        print("Start the server with:  uvicorn main:app --host 0.0.0.0 --port 3000 --reload  (from inside telecom_api/)")
        sys.exit(1)

def fetch_cell_site_lookup():
    """Return {cell_site_id: site_dict} for all sites currently in the API."""
    try:
        r = requests.get(f"{BASE_URL}/cell-sites", timeout=5)
        if r.status_code == 200:
            return {s["cell_site_id"]: s for s in r.json()}
    except Exception:
        pass
    return {}

def post_row(url, body):
    try:
        return requests.post(url, json=body, timeout=5)
    except Exception as exc:
        return exc   # caller checks isinstance(result, Exception)

# ── Output ────────────────────────────────────────────────────────────────────

def print_row(rr):
    errors   = [i for i in rr.issues if i.level == "ERROR"]
    warnings = [i for i in rr.issues if i.level == "WARNING"]
    infos    = [i for i in rr.issues if i.level == "INFO"]
    id_str   = "  " + "  ".join(f"{k}={v}" for k, v in rr.api_ids.items()) if rr.api_ids else ""

    if rr.status == "REJECTED":
        print(f"  Row {rr.row_num}: REJECTED")
        for i in errors:
            print(f"    [ERROR]   {i.field}: {i.message}")
    elif rr.status == "OK":
        print(f"  Row {rr.row_num}: OK  quality={rr.quality}%{id_str}")
        for i in infos:
            print(f"    [INFO]    {i.field}: {i.message}")
    elif rr.status == "WARNED":
        print(f"  Row {rr.row_num}: OK (warnings)  quality={rr.quality}%{id_str}")
        for i in warnings:
            print(f"    [WARNING] {i.field}: {i.message}")
        for i in infos:
            print(f"    [INFO]    {i.field}: {i.message}")
    elif rr.status == "API_ERROR":
        print(f"  Row {rr.row_num}: API ERROR")
        for i in rr.issues:
            print(f"    [{i.level:7}] {i.field}: {i.message}")

def print_summary(all_results):
    print("\n" + "=" * 62)
    print("VALIDATION SUMMARY")
    print("=" * 62)
    for name, results in all_results.items():
        if not results:
            print(f"\n  {name}  —  skipped (no data)")
            continue
        total    = len(results)
        rejected = sum(1 for r in results if r.status == "REJECTED")
        ok       = sum(1 for r in results if r.status == "OK")
        warned   = sum(1 for r in results if r.status == "WARNED")
        api_err  = sum(1 for r in results if r.status == "API_ERROR")
        qs       = [r.quality for r in results if r.status in ("OK", "WARNED")]
        avg_q    = f"{round(sum(qs)/len(qs))}%" if qs else "n/a"
        ef       = sorted({f for r in results if r.status == "WARNED"
                           for f in r.enrichment_fields})
        print(f"\n  {name}")
        print(f"    Rows: {total}  |  OK: {ok}  |  Warned: {warned}  "
              f"|  Rejected: {rejected}  |  API errors: {api_err}")
        print(f"    Avg quality score: {avg_q}")
        if ef:
            print(f"    Fields flagged for enrichment: {', '.join(ef)}")

# ── File processor ────────────────────────────────────────────────────────────

def process_file(xlsx_name, validator_fn, url_fn, skip_keys, known_cell_site_ids,
                 allow_missing_cs=False):
    """Process a component xlsx file: validate each row and POST to the API.

    allow_missing_cs: when True, rows with no cell_site_id are treated as
    fleet-level (WARNED, no API POST) instead of REJECTED.  Rows whose
    cell_site_id is not in known_cell_site_ids are treated as 3rd-party sites
    (WARNED, no API POST) regardless of this flag.  Both categories are
    summarised as a single count line rather than printed row-by-row.
    """
    headers, data_rows = load_sheet(xlsx_name)
    if headers is None:
        print(f"\nSKIP: {xlsx_name} — not found or empty.")
        return []

    print(f"\n--- {xlsx_name} ({len(data_rows)} rows) ---")
    results           = []
    third_party_count = 0
    fleet_level_count = 0

    for row in data_rows:
        row_num = row[0].row
        rd      = to_dict(headers, row)
        issues  = []

        cs_id_raw = rd.get("cell_site_id")
        if not present(cs_id_raw):
            if not allow_missing_cs:
                rr = RowResult(row_num, "REJECTED",
                               [Issue("ERROR", "cell_site_id", "cell_site_id is missing")])
                print_row(rr)
                results.append(rr)
                continue
            # Fleet-level row: no cell_site_id is expected
            issues.append(Issue("WARNING", "cell_site_id",
                                "No cell_site_id — treated as fleet-level"))
            skip_post = True
            is_fleet  = True
        else:
            is_fleet  = False
            cs_id_key = _cs_id_key(cs_id_raw)
            if cs_id_key not in known_cell_site_ids:
                issues.append(Issue("WARNING", "cell_site_id",
                                    f"cell_site_id {cs_id_raw} not found — treated as 3rd party"))
                skip_post = True
            else:
                skip_post = False

        # Schema validation
        issues += validator_fn(rd)

        errors = [i for i in issues if i.level == "ERROR"]
        if errors:
            rr = RowResult(row_num, "REJECTED", issues)
            print_row(rr)
            results.append(rr)
            continue

        if skip_post:
            warnings = [i for i in issues if i.level == "WARNING"]
            quality  = compute_quality(issues)
            ef       = [i.field for i in warnings]
            api_ids  = {"cell_site_id": cs_id_raw} if present(cs_id_raw) else {}
            rr = RowResult(row_num, "WARNED", issues, quality, api_ids, ef)
            results.append(rr)
            if is_fleet:
                fleet_level_count += 1
            else:
                third_party_count += 1
            continue

        # Build and POST
        body = {k: v for k, v in rd.items() if k not in skip_keys and present(v)}
        resp = post_row(url_fn(rd), body)

        if isinstance(resp, Exception):
            rr = RowResult(row_num, "API_ERROR",
                           issues + [Issue("ERROR", "api", str(resp))])
        elif resp.status_code in (200, 201):
            returned = resp.json()
            api_ids  = {k: returned[k]
                        for k in ("cell_site_id","active_id","passive_id","infrastructure_id")
                        if k in returned}
            warnings = [i for i in issues if i.level == "WARNING"]
            quality  = compute_quality(issues)
            status   = "WARNED" if warnings else "OK"
            ef       = [i.field for i in warnings]
            rr = RowResult(row_num, status, issues, quality, api_ids, ef)
        else:
            try:    msg = resp.json().get("message", resp.text)
            except: msg = resp.text
            rr = RowResult(row_num, "API_ERROR",
                           issues + [Issue("ERROR", "api", f"HTTP {resp.status_code}: {msg}")])

        print_row(rr)
        results.append(rr)

    if third_party_count:
        print(f"  {third_party_count} row(s) with unknown cell_site_id — treated as 3rd party (no API POST)")
    if fleet_level_count:
        print(f"  {fleet_level_count} row(s) with no cell_site_id — treated as fleet-level (no API POST)")

    return results

# ── API source helpers ────────────────────────────────────────────────────────

def fetch_all_from_api():
    """Fetch all entities from telecom_api.
    Returns {"cell_sites": [...], "active": [...], "passive": [...], "infrastructure": [...]}
    """
    try:
        r = requests.get(f"{BASE_URL}/cell-sites", timeout=5)
        cell_sites = r.json() if r.status_code == 200 else []
    except Exception:
        cell_sites = []

    active, passive, infra = [], [], []
    for site in cell_sites:
        cs_id = site["cell_site_id"]
        try:
            r = requests.get(f"{BASE_URL}/cell-sites/{cs_id}/active-components", timeout=5)
            if r.status_code == 200:
                active.extend(r.json())
        except Exception:
            pass
        try:
            r = requests.get(f"{BASE_URL}/cell-sites/{cs_id}/passive-components", timeout=5)
            if r.status_code == 200:
                passive.extend(r.json())
        except Exception:
            pass
        try:
            r = requests.get(f"{BASE_URL}/cell-sites/{cs_id}/infrastructure", timeout=5)
            if r.status_code == 200:
                infra.extend(r.json())
        except Exception:
            pass

    return {
        "cell_sites":     cell_sites,
        "active":         active,
        "passive":        passive,
        "infrastructure": infra,
    }


def validate_from_api():
    """Validate all data currently in telecom_api without POSTing anything.
    Data is already in the API — this runs quality checks only.
    Returns (all_results, rows_by_schema) where rows_by_schema feeds enrichment.
    """
    check_server()
    api_data = fetch_all_from_api()

    cell_sites          = api_data["cell_sites"]
    cell_site_lookup    = {s["cell_site_id"]: s for s in cell_sites}
    known_cell_site_ids = set(cell_site_lookup.keys())
    passive_cs_ids      = {r["cell_site_id"] for r in api_data["passive"]}
    infra_cs_ids        = {r["cell_site_id"] for r in api_data["infrastructure"]}

    all_results = {}

    # ── Cell sites ─────────────────────────────────────────────────────────────
    cs_results = []
    print(f"\n--- cell_sites ({len(cell_sites)} rows) ---")
    for i, site in enumerate(cell_sites, start=1):
        cs_id  = site["cell_site_id"]
        issues = validate_cell_site(site)
        if cs_id not in passive_cs_ids:
            issues.append(Issue("WARNING", "cell_site_id",
                                f"Cell site {cs_id} has no linked passive component"))
        if cs_id not in infra_cs_ids:
            issues.append(Issue("WARNING", "cell_site_id",
                                f"Cell site {cs_id} has no linked infrastructure"))
        errors   = [x for x in issues if x.level == "ERROR"]
        warnings = [x for x in issues if x.level == "WARNING"]
        quality  = compute_quality(issues)
        status   = "REJECTED" if errors else ("WARNED" if warnings else "OK")
        rr = RowResult(i, status, issues, quality,
                       {"cell_site_id": cs_id},
                       [x.field for x in warnings])
        print_row(rr)
        cs_results.append(rr)
    all_results["cell_site.xlsx"] = cs_results

    # ── Active components ──────────────────────────────────────────────────────
    ac_results = []
    print(f"\n--- active_components ({len(api_data['active'])} rows) ---")
    for i, comp in enumerate(api_data["active"], start=1):
        cs_id_raw = comp.get("cell_site_id")
        issues    = []
        if not present(cs_id_raw) or cs_id_raw not in known_cell_site_ids:
            issues.append(Issue("WARNING", "cell_site_id",
                                f"cell_site_id {cs_id_raw} not found in the database"))
        issues  += validate_active_component(comp, cell_site_lookup=cell_site_lookup)
        errors   = [x for x in issues if x.level == "ERROR"]
        warnings = [x for x in issues if x.level == "WARNING"]
        quality  = compute_quality(issues)
        status   = "REJECTED" if errors else ("WARNED" if warnings else "OK")
        api_ids  = {k: comp[k] for k in ("active_id", "cell_site_id") if k in comp}
        rr = RowResult(i, status, issues, quality, api_ids, [x.field for x in warnings])
        print_row(rr)
        ac_results.append(rr)
    all_results["active_components.xlsx"] = ac_results

    # ── Passive components ─────────────────────────────────────────────────────
    pa_results = []
    print(f"\n--- passive_components ({len(api_data['passive'])} rows) ---")
    for i, comp in enumerate(api_data["passive"], start=1):
        cs_id_raw = comp.get("cell_site_id")
        issues    = []
        if not present(cs_id_raw) or cs_id_raw not in known_cell_site_ids:
            issues.append(Issue("WARNING", "cell_site_id",
                                f"cell_site_id {cs_id_raw} not found in the database"))
        issues  += validate_passive_component(comp)
        errors   = [x for x in issues if x.level == "ERROR"]
        warnings = [x for x in issues if x.level == "WARNING"]
        quality  = compute_quality(issues)
        status   = "REJECTED" if errors else ("WARNED" if warnings else "OK")
        api_ids  = {k: comp[k] for k in ("passive_id", "cell_site_id") if k in comp}
        rr = RowResult(i, status, issues, quality, api_ids, [x.field for x in warnings])
        print_row(rr)
        pa_results.append(rr)
    all_results["passive_components.xlsx"] = pa_results

    # ── Infrastructure ─────────────────────────────────────────────────────────
    in_results = []
    print(f"\n--- infrastructure ({len(api_data['infrastructure'])} rows) ---")
    for i, item in enumerate(api_data["infrastructure"], start=1):
        cs_id_raw = item.get("cell_site_id")
        issues    = []
        if not present(cs_id_raw) or cs_id_raw not in known_cell_site_ids:
            issues.append(Issue("WARNING", "cell_site_id",
                                f"cell_site_id {cs_id_raw} not found in the database"))
        issues  += validate_infrastructure(item)
        errors   = [x for x in issues if x.level == "ERROR"]
        warnings = [x for x in issues if x.level == "WARNING"]
        quality  = compute_quality(issues)
        status   = "REJECTED" if errors else ("WARNED" if warnings else "OK")
        api_ids  = {k: item[k] for k in ("infrastructure_id", "cell_site_id") if k in item}
        rr = RowResult(i, status, issues, quality, api_ids, [x.field for x in warnings])
        print_row(rr)
        in_results.append(rr)
    all_results["infrastructure.xlsx"] = in_results

    rows_by_schema = {
        "cell_sites":     api_data["cell_sites"],
        "active":         api_data["active"],
        "passive":        api_data["passive"],
        "infrastructure": api_data["infrastructure"],
    }

    return all_results, rows_by_schema


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    check_server()

    # Pre-load cell_site_ids from passive and infrastructure files (for cross-file linkage check)
    passive_cell_site_ids = collect_cell_site_ids("passive_components.xlsx")
    infra_cell_site_ids   = collect_cell_site_ids("infrastructure.xlsx")

    # Fetch cell sites already in the API
    cell_site_lookup    = fetch_cell_site_lookup()
    known_cell_site_ids = set(cell_site_lookup.keys())

    all_results = {}

    # ── 1. Cell sites ──────────────────────────────────────────────────────────
    cs_headers, cs_rows = load_sheet("cell_site.xlsx")
    cs_results = []

    if cs_headers is None:
        print("\nSKIP: cell_site.xlsx — not found or empty.")
    else:
        print(f"\n--- cell_site.xlsx ({len(cs_rows)} rows) ---")
        for row in cs_rows:
            row_num = row[0].row
            rd      = to_dict(cs_headers, row)
            issues  = validate_cell_site(rd)

            errors = [i for i in issues if i.level == "ERROR"]
            if errors:
                rr = RowResult(row_num, "REJECTED", issues)
                print_row(rr)
                cs_results.append(rr)
                continue

            body = {k: v for k, v in rd.items() if present(v)}
            resp = post_row(f"{BASE_URL}/cell-sites", body)

            if isinstance(resp, Exception):
                rr = RowResult(row_num, "API_ERROR",
                               issues + [Issue("ERROR", "api", str(resp))])
                print_row(rr)
                cs_results.append(rr)
                continue

            if resp.status_code in (200, 201):
                returned = resp.json()
                cs_id    = returned.get("cell_site_id")

                # Register new site so nested files can reference it
                cell_site_lookup[cs_id]    = returned
                known_cell_site_ids.add(cs_id)

                # Cross-file linkage checks (warning on cell site row)
                if cs_id not in passive_cell_site_ids:
                    issues.append(Issue("WARNING", "cell_site_id",
                                        f"Cell site {cs_id} has no linked passive component"))
                if cs_id not in infra_cell_site_ids:
                    issues.append(Issue("WARNING", "cell_site_id",
                                        f"Cell site {cs_id} has no linked infrastructure"))

                warnings = [i for i in issues if i.level == "WARNING"]
                quality  = compute_quality(issues)
                status   = "WARNED" if warnings else "OK"
                ef       = [i.field for i in warnings]
                rr = RowResult(row_num, status, issues, quality,
                               {"cell_site_id": cs_id}, ef)
            else:
                try:    msg = resp.json().get("message", resp.text)
                except: msg = resp.text
                rr = RowResult(row_num, "API_ERROR",
                               issues + [Issue("ERROR", "api",
                                               f"HTTP {resp.status_code}: {msg}")])

            print_row(rr)
            cs_results.append(rr)

    all_results["cell_site.xlsx"] = cs_results

    # ── 2. Active components ───────────────────────────────────────────────────
    all_results["active_components.xlsx"] = process_file(
        "active_components.xlsx",
        validator_fn=partial(validate_active_component, cell_site_lookup=cell_site_lookup),
        url_fn=lambda rd: f"{BASE_URL}/cell-sites/{rd['cell_site_id']}/active-components",
        skip_keys={"cell_site_id"},
        known_cell_site_ids=known_cell_site_ids,
    )

    # ── 3. Passive components ──────────────────────────────────────────────────
    all_results["passive_components.xlsx"] = process_file(
        "passive_components.xlsx",
        validator_fn=validate_passive_component,
        url_fn=lambda rd: f"{BASE_URL}/cell-sites/{rd['cell_site_id']}/passive-components",
        skip_keys={"cell_site_id"},
        known_cell_site_ids=known_cell_site_ids,
    )

    # ── 4. Infrastructure ──────────────────────────────────────────────────────
    all_results["infrastructure.xlsx"] = process_file(
        "infrastructure.xlsx",
        validator_fn=validate_infrastructure,
        url_fn=lambda rd: f"{BASE_URL}/cell-sites/{rd['cell_site_id']}/infrastructure",
        skip_keys={"cell_site_id", "infrastructure_id"},
        known_cell_site_ids=known_cell_site_ids,
        allow_missing_cs=True,
    )

    print_summary(all_results)


if __name__ == "__main__":
    main()
