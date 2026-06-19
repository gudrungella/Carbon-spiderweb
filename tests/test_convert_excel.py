"""Tests for convert_excel.py"""

import csv
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))
from step9_convert_excel import convert, is_empty_row


@pytest.fixture
def tmp_xlsx(tmp_path):
    """Create a minimal Excel file with a Data sheet for testing."""
    def _make(rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        for row in rows:
            ws.append(row)
        path = tmp_path / "test.xlsx"
        wb.save(path)
        return path
    return _make


def test_is_empty_row_all_none():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None, None, None])
    row = list(ws.iter_rows())[0]
    assert is_empty_row(row) is True


def test_is_empty_row_has_value():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["hello", None])
    row = list(ws.iter_rows())[0]
    assert is_empty_row(row) is False


def test_convert_writes_csv(tmp_path, tmp_xlsx):
    xlsx_path = tmp_xlsx([
        ["site_code", "country", "site_type"],
        ["SITE-001", "NO", "macro"],
        ["SITE-002", "SE", "micro"],
    ])
    csv_path = tmp_path / "output.csv"

    # Patch DATA_DIR temporarily
    import step9_convert_excel as convert_excel
    original_dir = convert_excel.DATA_DIR
    convert_excel.DATA_DIR = tmp_path

    # Rename to expected name
    xlsx_path.rename(tmp_path / "cell_site.xlsx")
    convert_excel.convert("cell_site.xlsx", "output.csv")

    convert_excel.DATA_DIR = original_dir

    with open(csv_path, newline="") as f:
        rows = list(csv.reader(f))

    assert rows[0] == ["site_code", "country", "site_type"]
    assert rows[1] == ["SITE-001", "NO", "macro"]
    assert rows[2] == ["SITE-002", "SE", "micro"]
    assert len(rows) == 3


def test_convert_skips_empty_rows(tmp_path, tmp_xlsx):
    xlsx_path = tmp_xlsx([
        ["site_code", "country"],
        ["SITE-001", "NO"],
        [None, None],
        ["SITE-002", "SE"],
    ])
    csv_path = tmp_path / "output.csv"

    import step9_convert_excel as convert_excel
    original_dir = convert_excel.DATA_DIR
    convert_excel.DATA_DIR = tmp_path
    xlsx_path.rename(tmp_path / "cell_site.xlsx")
    convert_excel.convert("cell_site.xlsx", "output.csv")
    convert_excel.DATA_DIR = original_dir

    with open(csv_path, newline="") as f:
        rows = list(csv.reader(f))

    assert len(rows) == 3  # header + 2 data rows (empty row skipped)


def test_convert_missing_file(tmp_path, capsys):
    import step9_convert_excel as convert_excel
    original_dir = convert_excel.DATA_DIR
    convert_excel.DATA_DIR = tmp_path
    convert_excel.convert("nonexistent.xlsx", "output.csv")
    convert_excel.DATA_DIR = original_dir

    captured = capsys.readouterr()
    assert "SKIP" in captured.out
