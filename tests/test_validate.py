"""Tests for validate.py — uses unittest.mock to avoid needing a running server."""

import sys
from pathlib import Path
from unittest.mock import patch, MagicMock

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))
from step3_validate import build_body, is_empty_row


def test_build_body_excludes_skip_keys():
    headers = ["cell_site_id", "active_type", "life_time"]
    values = [1, "generator", 15]
    body = build_body(headers, values, skip_keys={"cell_site_id"})
    assert "cell_site_id" not in body
    assert body["active_type"] == "generator"
    assert body["life_time"] == 15


def test_build_body_omits_none_values():
    headers = ["active_type", "active_subtype", "life_time"]
    values = ["generator", None, 10]
    body = build_body(headers, values, skip_keys=set())
    assert "active_subtype" not in body
    assert body["active_type"] == "generator"


def test_build_body_omits_empty_strings():
    headers = ["active_type", "operator_name"]
    values = ["generator", "  "]
    body = build_body(headers, values, skip_keys=set())
    assert "operator_name" not in body


def test_is_empty_row_all_none():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None, None])
    row = list(ws.iter_rows())[0]
    assert is_empty_row(row) is True


def test_is_empty_row_with_value():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["generator", None])
    row = list(ws.iter_rows())[0]
    assert is_empty_row(row) is False


def test_check_server_exits_on_connection_error():
    import requests
    import step3_validate as validate
    with patch("validate.requests.get", side_effect=requests.exceptions.ConnectionError):
        with pytest.raises(SystemExit) as exc_info:
            validate.check_server()
        assert exc_info.value.code == 1
