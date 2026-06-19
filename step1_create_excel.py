"""
create_excel.py — One-time builder for all 4 telecom inventory Excel files.
Run this script to regenerate the Excel files in data/ from scratch.
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DATA_DIR = Path(__file__).parent / "data"

# ---------------------------------------------------------------------------
# Shared enum lists
# ---------------------------------------------------------------------------

INSTALLATION_METHODS = [
    "traditional_excavation_green", "microtrenching_green", "milling_green",
    "plowing_green", "traditional_excavation_urban", "microtrenching_urban",
    "milling_urban", "plowing_urban", "truck", "van", "ship", "airplane",
    "crane", "other",
]

MAINTENANCE_METHODS = [
    "physical_security_services", "decommissioning", "replacements",
    "environmental_monitoring", "failure_monitoring", "software_updates", "other",
]

INSTALL_UNITS        = ["km", "kW", "kWh", "L", "m3", "m2"]
POWER_UNITS          = ["kW", "W", "kWh", "L", "m3"]
NETWORK_TYPES        = ["access", "metro", "aggregation", "backbone", "core"]
TECHNOLOGY_TYPES     = ["2G", "3G", "4G", "5G", "6G", "ADSL", "DSL", "FTTH", "FTTC",
                        "GPON", "HFC", "LoRa", "MPLS", "MEC", "Network slicing", "PON",
                        "POTS/PSTN", "RFID", "SD-WAN", "VoIP", "WLAN", "XPON"]
EMISSION_UNIT        = ["kgCO2eq/unit"]
INSTALL_EMIT_UNITS   = ["kgCO2eq/km", "kgCO2eq/kW", "kgCO2eq/kWh", "kgCO2eq/L", "kgCO2eq/m3", "kgCO2eq/m2"]
POWER_EMIT_UNITS     = ["kgCO2eq/kW", "kgCO2eq/W", "kgCO2eq/kWh", "kgCO2eq/L", "kgCO2eq/m3"]

# ---------------------------------------------------------------------------
# Schema definitions
# Each field: name, type, required, enum (None = free text), description
# ---------------------------------------------------------------------------

CELL_SITE_SCHEMA = [
    {"name": "cell_site_id",        "type": "string",  "required": False, "enum": None,                                                                                      "description": "Original site identifier (e.g. 45872-WD1). Used as the API key if provided."},
    {"name": "site_type",           "type": "string",  "required": True,  "enum": ["small","medium","large","macro","micro","pico","femto","rooftop","indoor","greenfield"], "description": "Physical type of the cell site"},
    {"name": "network_type",        "type": "string",  "required": True,  "enum": NETWORK_TYPES,                                                                              "description": "Network layer type"},
    {"name": "country",             "type": "string",  "required": True,  "enum": None,                                                                                      "description": "Country where the site is located"},
    {"name": "operational_status",  "type": "string",  "required": False, "enum": ["active","planned","under_construction","decommissioned","maintenance"],                   "description": "Current operational status"},
    {"name": "operator_name",       "type": "string",  "required": False, "enum": None,                                                                                      "description": "Name of the network operator"},
    {"name": "per_rented",          "type": "number",  "required": False, "enum": None,                                                                                      "description": "Percentage of the site that is rented (0–100)"},
    {"name": "electricity_source",  "type": "string",  "required": False, "enum": ["grid","renewable","solar","wind","none"],                                                 "description": "Source of electricity supply"},
    {"name": "fuel_type",           "type": "string",  "required": False, "enum": ["bensin","diesel","alternative","none","other"],                                           "description": "Type of fuel used on site"},
    {"name": "measured_electricity",      "type": "number",  "required": False, "enum": None,             "description": "Measured electricity consumption — annual total"},
    {"name": "measured_electricity_unit", "type": "string",  "required": False, "enum": ["kWh"],          "description": "Unit for measured_electricity (kWh/year)"},
    {"name": "measured_electricity_jan",  "type": "number",  "required": False, "enum": None,             "description": "Measured electricity — January (kWh)"},
    {"name": "measured_electricity_feb",  "type": "number",  "required": False, "enum": None,             "description": "Measured electricity — February (kWh)"},
    {"name": "measured_electricity_mar",  "type": "number",  "required": False, "enum": None,             "description": "Measured electricity — March (kWh)"},
    {"name": "measured_electricity_apr",  "type": "number",  "required": False, "enum": None,             "description": "Measured electricity — April (kWh)"},
    {"name": "measured_electricity_may",  "type": "number",  "required": False, "enum": None,             "description": "Measured electricity — May (kWh)"},
    {"name": "measured_electricity_jun",  "type": "number",  "required": False, "enum": None,             "description": "Measured electricity — June (kWh)"},
    {"name": "measured_electricity_jul",  "type": "number",  "required": False, "enum": None,             "description": "Measured electricity — July (kWh)"},
    {"name": "measured_electricity_aug",  "type": "number",  "required": False, "enum": None,             "description": "Measured electricity — August (kWh)"},
    {"name": "measured_electricity_sep",  "type": "number",  "required": False, "enum": None,             "description": "Measured electricity — September (kWh)"},
    {"name": "measured_electricity_oct",  "type": "number",  "required": False, "enum": None,             "description": "Measured electricity — October (kWh)"},
    {"name": "measured_electricity_nov",  "type": "number",  "required": False, "enum": None,             "description": "Measured electricity — November (kWh)"},
    {"name": "measured_electricity_dec",  "type": "number",  "required": False, "enum": None,             "description": "Measured electricity — December (kWh)"},
    {"name": "measured_fuel",             "type": "number",  "required": False, "enum": None,             "description": "Measured fuel consumption — annual total"},
    {"name": "measured_fuel_unit",        "type": "string",  "required": False, "enum": ["L","m3","kg"],  "description": "Unit for measured_fuel"},
    {"name": "measured_refrigerant",      "type": "number",  "required": False, "enum": None,             "description": "Measured refrigerant consumption — annual total"},
    {"name": "measured_refrigerant_unit", "type": "string",  "required": False, "enum": ["kg","m3"],      "description": "Unit for measured_refrigerant"},
    {"name": "refrigerant_type",              "type": "string",  "required": False, "enum": ["none","R-11","R-12","R-22","R-123","R-134a","R-410A","R-245fa","R-32","R-1233zd(E)","R-1234yf","R-513A","other"], "description": "Type of refrigerant used"},
    {"name": "electricity_emission_factor",      "type": "number",  "required": False, "enum": None,                  "description": "Site-level emission factor for electricity consumption"},
    {"name": "electricity_emission_factor_unit", "type": "string",  "required": False, "enum": ["kgCO2eq/kWh"],       "description": "Unit of the electricity emission factor"},
    {"name": "fuel_emission_factor",             "type": "number",  "required": False, "enum": None,                  "description": "Site-level emission factor for fuel consumption"},
    {"name": "fuel_emission_factor_unit",        "type": "string",  "required": False, "enum": ["kgCO2eq/L","kgCO2eq/m3","kgCO2eq/kg"], "description": "Unit of the fuel emission factor — must match measured_fuel_unit denominator"},
    {"name": "refrigerant_emission_factor",      "type": "number",  "required": False, "enum": None,                  "description": "Site-level emission factor for refrigerant consumption"},
    {"name": "refrigerant_emission_factor_unit", "type": "string",  "required": False, "enum": ["kgCO2eq/m3","kgCO2eq/kg"], "description": "Unit of the refrigerant emission factor — must match measured_refrigerant_unit denominator"},
]

ACTIVE_COMPONENTS_SCHEMA = [
    {"name": "cell_site_id",                     "type": "string",  "required": True,  "enum": None,                                                                                                                                                       "description": "Used as URL path parameter, not sent in POST body"},
    {"name": "manufacture_part_number",          "type": "string",  "required": False, "enum": None,                                                                                                                                                       "description": "Manufacturer part number"},
    {"name": "brand",                            "type": "string",  "required": False, "enum": None,                                                                                                                                                       "description": "Brand or manufacturer name"},
    {"name": "active_type",    "type": "string",  "required": False, "enum": ["generator","cooling","fire_suppression","network_equipment","radio_equipment","power_equipment"],                                                                                                                                                                                                                    "description": "Category of active component"},
    {"name": "active_subtype", "type": "string",  "required": False, "enum": ["standby_power_generator","prime_power_generator","portable_industrial_generator","inverter_generator","container_sized_generator","water_cooled_systems","air_cooled_systems","industrial_chillers","evaporative_cooling_systems","hybrid_system","specialized_cooling","gas_suppression","sprinkler_system","foam_system","aggregation_router","chassis","edge_platform","firewall","gateway","isam","memory_card","router","switch","sfp","air","cellular_modem","radio_unit","baseband_unit","base_station","antenna","wls","sla_vrla","bci","din_en","dc_holder","power_distribution","light","camera","sensor"], "description": "Specific equipment subtype"},
    {"name": "technology_type",                  "type": "string",  "required": False, "enum": TECHNOLOGY_TYPES,                                                                                                                                            "description": "Network technology type"},
    {"name": "power_source",                     "type": "string",  "required": True,  "enum": ["battery","electricity","fuel","refrigerant","other"],                                                                                                      "description": "Source of power for this component"},
    {"name": "life_time",                        "type": "number",  "required": True,  "enum": None,                                                                                                                                                       "description": "Expected lifetime in years (0–200)"},
    {"name": "production_emissions",               "type": "number",  "required": False, "enum": None,                                                                                                                                                       "description": "Manufacturing/production emissions only (kgCO2eq/unit)"},
    {"name": "production_emissions_unit",          "type": "string",  "required": False, "enum": EMISSION_UNIT,                                                                                                                                               "description": "Unit of the production emissions"},
    {"name": "endoflife_emissions",              "type": "number",  "required": False, "enum": None,                                                                                                                                                       "description": "Emissions from end-of-life disposal or recycling"},
    {"name": "endoflife_emissions_unit",         "type": "string",  "required": False, "enum": EMISSION_UNIT,                                                                                                                                               "description": "Unit of the end-of-life emissions"},
    {"name": "power_quantity",                   "type": "number",  "required": False, "enum": None,                                                                                                                                                       "description": "Estimated or assumed power quantity"},
    {"name": "power_unit",                       "type": "string",  "required": False, "enum": POWER_UNITS,                                                                                                                                                 "description": "Unit of the power quantity"},
    {"name": "power_source_emission_factor",     "type": "number",  "required": False, "enum": None,                                                                                                                                                       "description": "Emission factor for the power source"},
    {"name": "power_source_emission_factor_unit","type": "string",  "required": False, "enum": POWER_EMIT_UNITS,                                                                                                                                            "description": "Unit of the power source emission factor"},
    {"name": "power_idle",                       "type": "number",  "required": False, "enum": None,                                                                                                                                                       "description": "Power consumption at idle"},
    {"name": "power_idle_unit",                  "type": "string",  "required": False, "enum": POWER_UNITS,                                                                                                                                                 "description": "Unit of the idle power"},
    {"name": "power_max",                        "type": "number",  "required": False, "enum": None,                                                                                                                                                       "description": "Maximum power consumption"},
    {"name": "power_max_unit",                   "type": "string",  "required": False, "enum": POWER_UNITS,                                                                                                                                                 "description": "Unit of the maximum power"},
    {"name": "installation_method",              "type": "string",  "required": False, "enum": INSTALLATION_METHODS,                                                                                                                                        "description": "Method used for installation"},
    {"name": "installation_quantity",            "type": "number",  "required": False, "enum": None,                                                                                                                                                       "description": "Quantity of units installed"},
    {"name": "installation_unit",                "type": "string",  "required": False, "enum": INSTALL_UNITS,                                                                                                                                               "description": "Unit of the installation quantity"},
    {"name": "installation_emission_factor",     "type": "number",  "required": False, "enum": None,                                                                                                                                                       "description": "Emission factor for the installation method"},
    {"name": "installation_emission_factor_unit","type": "string",  "required": False, "enum": INSTALL_EMIT_UNITS,                                                                                                                                          "description": "Unit of the installation emission factor"},
    {"name": "maintenance_method",               "type": "string",  "required": False, "enum": MAINTENANCE_METHODS,                                                                                                                                         "description": "Method used for maintenance"},
    {"name": "maintenance_quantity",             "type": "number",  "required": False, "enum": None,                                                                                                                                                       "description": "Quantity of maintenance actions"},
    {"name": "maintenance_unit",                 "type": "string",  "required": False, "enum": INSTALL_UNITS,                                                                                                                                               "description": "Unit of the maintenance quantity"},
    {"name": "maintenance_emission_factor",      "type": "number",  "required": False, "enum": None,                                                                                                                                                       "description": "Emission factor for the maintenance method"},
    {"name": "maintenance_emission_factor_unit", "type": "string",  "required": False, "enum": INSTALL_EMIT_UNITS,                                                                                                                                          "description": "Unit of the maintenance emission factor"},
]

PASSIVE_COMPONENTS_SCHEMA = [
    {"name": "cell_site_id",                     "type": "string",  "required": True,  "enum": None,                                                                                                           "description": "Used as URL path parameter, not sent in POST body"},
    {"name": "manufacture_part_number",          "type": "string",  "required": False, "enum": None,                                                                                                           "description": "Manufacturer part number"},
    {"name": "brand",                            "type": "string",  "required": False, "enum": None,                                                                                                           "description": "Brand or manufacturer name"},
    {"name": "passive_type",                     "type": "string",  "required": True,  "enum": ["fiber_cable","electrical_cables","COAX","splitters","shelters","cabinets","plugs","fencing","steel","aluminum","plastic"], "description": "Type of passive component"},
    {"name": "technology_type",                  "type": "string",  "required": False, "enum": TECHNOLOGY_TYPES,                                                                                               "description": "Network technology type"},
    {"name": "life_time",                        "type": "number",  "required": True,  "enum": None,                                                                                                           "description": "Expected lifetime in years (0–200)"},
    {"name": "production_emissions",               "type": "number",  "required": False, "enum": None,                                                                                                           "description": "Manufacturing/production emissions only (kgCO2eq/unit)"},
    {"name": "production_emissions_unit",          "type": "string",  "required": False, "enum": EMISSION_UNIT,                                                                                                   "description": "Unit of the production emissions"},
    {"name": "endoflife_emissions",              "type": "number",  "required": False, "enum": None,                                                                                                           "description": "Emissions from end-of-life disposal or recycling"},
    {"name": "endoflife_emissions_unit",         "type": "string",  "required": False, "enum": EMISSION_UNIT,                                                                                                   "description": "Unit of the end-of-life emissions"},
    {"name": "installation_method",              "type": "string",  "required": False, "enum": INSTALLATION_METHODS,                                                                                            "description": "Method used for installation"},
    {"name": "installation_quantity",            "type": "number",  "required": False, "enum": None,                                                                                                           "description": "Quantity of units installed"},
    {"name": "installation_unit",                "type": "string",  "required": False, "enum": INSTALL_UNITS,                                                                                                   "description": "Unit of the installation quantity"},
    {"name": "installation_emission_factor",     "type": "number",  "required": False, "enum": None,                                                                                                           "description": "Emission factor for the installation method"},
    {"name": "installation_emission_factor_unit","type": "string",  "required": False, "enum": INSTALL_EMIT_UNITS,                                                                                              "description": "Unit of the installation emission factor"},
    {"name": "maintenance_method",               "type": "string",  "required": False, "enum": MAINTENANCE_METHODS,                                                                                             "description": "Method used for maintenance"},
    {"name": "maintenance_quantity",             "type": "number",  "required": False, "enum": None,                                                                                                           "description": "Quantity of maintenance actions"},
    {"name": "maintenance_unit",                 "type": "string",  "required": False, "enum": INSTALL_UNITS,                                                                                                   "description": "Unit of the maintenance quantity"},
    {"name": "maintenance_emission_factor",      "type": "number",  "required": False, "enum": None,                                                                                                           "description": "Emission factor for the maintenance method"},
    {"name": "maintenance_emission_factor_unit", "type": "string",  "required": False, "enum": INSTALL_EMIT_UNITS,                                                                                              "description": "Unit of the maintenance emission factor"},
]

INFRASTRUCTURE_SCHEMA = [
    {"name": "cell_site_id",                     "type": "string",  "required": True,  "enum": None,                                                                                                           "description": "Used as URL path parameter, not sent in POST body"},
    {"name": "infrastructure_id",                "type": "integer", "required": False, "enum": None,                                                                                                           "description": "Auto-assigned by API — for reference only, not sent in POST body"},
    {"name": "contractor_name",                  "type": "string",  "required": False, "enum": None,                                                                                                           "description": "Name of the contractor responsible for installation"},
    {"name": "infrastructure_type",              "type": "string",  "required": True,  "enum": ["tower","mast","rooftop_mount","pole","building","underground","container","real estate","manhole","concrete","ducts & pipes"], "description": "Type of infrastructure"},
    {"name": "network_type",                     "type": "string",  "required": True,  "enum": NETWORK_TYPES,                                                                                                  "description": "Network layer type"},
    {"name": "life_time",                        "type": "number",  "required": True,  "enum": None,                                                                                                           "description": "Expected lifetime in years (0–200)"},
    {"name": "production_emissions",               "type": "number",  "required": False, "enum": None,                                                                                                           "description": "Manufacturing/production emissions only (kgCO2eq/unit)"},
    {"name": "production_emissions_unit",          "type": "string",  "required": False, "enum": EMISSION_UNIT,                                                                                                   "description": "Unit of the production emissions"},
    {"name": "endoflife_emissions",              "type": "number",  "required": False, "enum": None,                                                                                                           "description": "Emissions from end-of-life disposal or recycling"},
    {"name": "endoflife_emissions_unit",         "type": "string",  "required": False, "enum": EMISSION_UNIT,                                                                                                   "description": "Unit of the end-of-life emissions"},
    {"name": "installation_method",              "type": "string",  "required": False, "enum": INSTALLATION_METHODS,                                                                                            "description": "Method used for installation"},
    {"name": "installation_quantity",            "type": "number",  "required": False, "enum": None,                                                                                                           "description": "Quantity of units installed"},
    {"name": "installation_unit",                "type": "string",  "required": False, "enum": INSTALL_UNITS,                                                                                                   "description": "Unit of the installation quantity"},
    {"name": "installation_emission_factor",     "type": "number",  "required": False, "enum": None,                                                                                                           "description": "Emission factor for the installation method"},
    {"name": "installation_emission_factor_unit","type": "string",  "required": False, "enum": INSTALL_EMIT_UNITS,                                                                                              "description": "Unit of the installation emission factor"},
    {"name": "maintenance_method",               "type": "string",  "required": False, "enum": MAINTENANCE_METHODS,                                                                                             "description": "Method used for maintenance"},
    {"name": "maintenance_quantity",             "type": "number",  "required": False, "enum": None,                                                                                                           "description": "Quantity of maintenance actions"},
    {"name": "maintenance_unit",                 "type": "string",  "required": False, "enum": INSTALL_UNITS,                                                                                                   "description": "Unit of the maintenance quantity"},
    {"name": "maintenance_emission_factor",      "type": "number",  "required": False, "enum": None,                                                                                                           "description": "Emission factor for the maintenance method"},
    {"name": "maintenance_emission_factor_unit", "type": "string",  "required": False, "enum": INSTALL_EMIT_UNITS,                                                                                              "description": "Unit of the maintenance emission factor"},
]

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

HEADER_FONT   = Font(bold=True, color="FFFFFF")
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
REQUIRED_FILL = PatternFill("solid", fgColor="D6E4F0")
INSTRUCT_HEADER_FONT = Font(bold=True, color="FFFFFF")
INSTRUCT_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")

# ---------------------------------------------------------------------------
# Core builder
# ---------------------------------------------------------------------------

def build_excel(schema, output_path):
    wb = openpyxl.Workbook()

    # --- Data sheet ---
    ws_data = wb.active
    ws_data.title = "Data"

    # --- Hidden _Dropdowns sheet ---
    ws_drop = wb.create_sheet("_Dropdowns")
    ws_drop.sheet_state = "hidden"

    # Collect all unique enums and assign them a column in _Dropdowns
    enum_col_map = {}  # enum tuple → column letter in _Dropdowns
    drop_col_idx = 1
    for field in schema:
        if field["enum"] is not None:
            key = tuple(field["enum"])
            if key not in enum_col_map:
                col_letter = get_column_letter(drop_col_idx)
                enum_col_map[key] = (col_letter, len(field["enum"]))
                for row_idx, val in enumerate(field["enum"], start=1):
                    ws_drop[f"{col_letter}{row_idx}"] = val
                drop_col_idx += 1

    # Write Data sheet headers
    for col_idx, field in enumerate(schema, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws_data[f"{col_letter}1"]
        cell.value = field["name"]
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws_data.column_dimensions[col_letter].width = max(len(field["name"]) + 4, 18)

        # Apply dropdown validation if enum field
        if field["enum"] is not None:
            key = tuple(field["enum"])
            drop_col_letter, count = enum_col_map[key]
            formula = f"'_Dropdowns'!${drop_col_letter}$1:${drop_col_letter}${count}"
            dv = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid value",
                error="Please select a value from the dropdown list.",
            )
            dv.sqref = f"{col_letter}2:{col_letter}1000"
            ws_data.add_data_validation(dv)

    # Freeze header row
    ws_data.freeze_panes = "A2"

    # --- Instructions sheet ---
    ws_inst = wb.create_sheet("Instructions")
    inst_headers = ["Column", "Type", "Required", "Valid Values", "Notes"]
    for col_idx, h in enumerate(inst_headers, start=1):
        cell = ws_inst[f"{get_column_letter(col_idx)}1"]
        cell.value = h
        cell.font = INSTRUCT_HEADER_FONT
        cell.fill = INSTRUCT_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    ws_inst.column_dimensions["A"].width = 28
    ws_inst.column_dimensions["B"].width = 12
    ws_inst.column_dimensions["C"].width = 12
    ws_inst.column_dimensions["D"].width = 60
    ws_inst.column_dimensions["E"].width = 40

    for row_idx, field in enumerate(schema, start=2):
        valid_values = ", ".join(field["enum"]) if field["enum"] else ("0–200" if "life_time" in field["name"] else ("0–100" if "per_rented" in field["name"] else "Any"))
        ws_inst[f"A{row_idx}"] = field["name"]
        ws_inst[f"B{row_idx}"] = field["type"]
        ws_inst[f"C{row_idx}"] = "Yes" if field["required"] else "No"
        ws_inst[f"D{row_idx}"] = valid_values
        ws_inst[f"E{row_idx}"] = field["description"]
        ws_inst[f"D{row_idx}"].alignment = Alignment(wrap_text=True)
        ws_inst.row_dimensions[row_idx].height = 30 if field["enum"] and len(field["enum"]) > 8 else 18

    ws_inst.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Created: {output_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    DATA_DIR.mkdir(exist_ok=True)

    build_excel(CELL_SITE_SCHEMA,           DATA_DIR / "cell_site.xlsx")
    build_excel(ACTIVE_COMPONENTS_SCHEMA,   DATA_DIR / "active_components.xlsx")
    build_excel(PASSIVE_COMPONENTS_SCHEMA,  DATA_DIR / "passive_components.xlsx")
    build_excel(INFRASTRUCTURE_SCHEMA,      DATA_DIR / "infrastructure.xlsx")

    print("\nAll Excel files created successfully.")
    print("Open the files in Excel/LibreOffice to fill in data.")
