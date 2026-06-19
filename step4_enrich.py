"""
Emission factor and power data enrichment for telecom inventory.

Two provider hierarchies (defined in api_config.py):
  EMISSION_HIERARCHY: custom_file → rejoose → climatiq → epd → resilio → ecoinvent
  POWER_HIERARCHY:    tscircuit

Fill strategy — provider-first:
  For each provider, every row is attempted before moving to the next provider.
  This ensures the highest-priority source fills as much as possible before
  lower-priority sources are consulted.

Output: data/<name>_enriched.xlsx, with original columns plus
        <field>_source and <field>_confidence for each enriched field.
        Rows that could not be filled are appended to emission_factors.xlsx
        for manual input.

Backward-compatible functions enrich_from_api() and report_unmatched()
are preserved for any existing callers.
"""

import math
import os
import statistics as _stats
from pathlib import Path
from typing import Optional

try:
    import requests
    _REQUESTS_AVAILABLE = True
except ImportError:
    _REQUESTS_AVAILABLE = False

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

import step2_api_config as cfg

DATA_DIR = Path(__file__).parent / "data"


# ---------------------------------------------------------------------------
# .env loader
# ---------------------------------------------------------------------------

def _load_env():
    """Load .env from the project directory into os.environ (won't overwrite)."""
    env_path = Path(__file__).parent / ".env"
    if not env_path.exists():
        return
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            os.environ.setdefault(key, value)


_load_env()


def _env(key: str) -> Optional[str]:
    return os.environ.get(key)


# ---------------------------------------------------------------------------
# Schema definitions — which fields to enrich per schema
# ---------------------------------------------------------------------------
# Each entry: field, unit_field, default_unit, precondition, provider_group
# precondition: name of a field that must be present, or None
# provider_group: "emission" uses EMISSION_HIERARCHY; "power" uses POWER_HIERARCHY

ENRICHABLE_FIELDS = {
    "cell_site": [
        {"field": "electricity_emission_factor", "unit_field": "electricity_emission_factor_unit", "default_unit": "kgCO2eq/kWh", "precondition": "electricity_source", "provider_group": "emission"},
        {"field": "fuel_emission_factor",        "unit_field": "fuel_emission_factor_unit",        "default_unit": "kgCO2eq/L",   "precondition": "fuel_type",          "provider_group": "emission"},
        {"field": "refrigerant_emission_factor", "unit_field": "refrigerant_emission_factor_unit", "default_unit": "kgCO2eq/m3",  "precondition": "refrigerant_type",   "provider_group": "emission"},
    ],
    "active": [
        {"field": "life_time",                    "unit_field": None,                                "default_unit": None,           "precondition": None,                    "provider_group": "lifetime"},
        {"field": "production_emissions",           "unit_field": "production_emissions_unit",           "default_unit": "kgCO2eq/unit", "precondition": None,                    "provider_group": "emission"},
        {"field": "endoflife_emissions",          "unit_field": "endoflife_emissions_unit",          "default_unit": "kgCO2eq/unit", "precondition": None,                    "provider_group": "emission"},
        {"field": "power_source_emission_factor", "unit_field": "power_source_emission_factor_unit", "default_unit": "kgCO2eq/kWh",  "precondition": "power_source",          "provider_group": "emission"},
        {"field": "installation_emission_factor", "unit_field": "installation_emission_factor_unit", "default_unit": None,           "precondition": "installation_quantity", "provider_group": "emission"},
        {"field": "maintenance_emission_factor",  "unit_field": "maintenance_emission_factor_unit",  "default_unit": None,           "precondition": "maintenance_quantity",  "provider_group": "emission"},
        {"field": "power_idle",                   "unit_field": "power_idle_unit",                   "default_unit": "W",            "precondition": None,                    "provider_group": "power", "skip_if": "power_quantity"},
        {"field": "power_max",                    "unit_field": "power_max_unit",                    "default_unit": "W",            "precondition": None,                    "provider_group": "power", "skip_if": "power_quantity"},
    ],
    "passive": [
        {"field": "life_time",                    "unit_field": None,                                "default_unit": None,           "precondition": None,                    "provider_group": "lifetime"},
        {"field": "production_emissions",           "unit_field": "production_emissions_unit",           "default_unit": "kgCO2eq/unit", "precondition": None,                    "provider_group": "emission"},
        {"field": "endoflife_emissions",          "unit_field": "endoflife_emissions_unit",          "default_unit": "kgCO2eq/unit", "precondition": None,                    "provider_group": "emission"},
        {"field": "installation_emission_factor", "unit_field": "installation_emission_factor_unit", "default_unit": None,           "precondition": "installation_quantity", "provider_group": "emission"},
        {"field": "maintenance_emission_factor",  "unit_field": "maintenance_emission_factor_unit",  "default_unit": None,           "precondition": "maintenance_quantity",  "provider_group": "emission"},
    ],
    "infrastructure": [
        {"field": "life_time",                    "unit_field": None,                                "default_unit": None,           "precondition": None,                    "provider_group": "lifetime"},
        {"field": "production_emissions",           "unit_field": "production_emissions_unit",           "default_unit": "kgCO2eq/unit", "precondition": None,                    "provider_group": "emission"},
        {"field": "endoflife_emissions",          "unit_field": "endoflife_emissions_unit",          "default_unit": "kgCO2eq/unit", "precondition": None,                    "provider_group": "emission"},
        {"field": "installation_emission_factor", "unit_field": "installation_emission_factor_unit", "default_unit": None,           "precondition": "installation_quantity", "provider_group": "emission"},
        {"field": "maintenance_emission_factor",  "unit_field": "maintenance_emission_factor_unit",  "default_unit": None,           "precondition": "maintenance_quantity",  "provider_group": "emission"},
    ],
}

CONFIDENCE = {
    "custom_file":       1.0,
    "cell_site":         1.0,
    "rejoose":           0.95,
    "climatiq":          0.9,
    "epd":               0.85,
    "resilio":           0.8,
    "ecoinvent":         0.75,
    "tscircuit":         0.9,
    "lifetime_defaults": 0.5,
    "manual":            0.9,
}

# Maps component power_source → (cell_site EF field, cell_site EF unit field)
_POWER_SOURCE_TO_SITE_EF = {
    "electricity": ("electricity_emission_factor", "electricity_emission_factor_unit"),
    "battery":     ("electricity_emission_factor", "electricity_emission_factor_unit"),
    "fuel":        ("fuel_emission_factor",         "fuel_emission_factor_unit"),
    "refrigerant": ("refrigerant_emission_factor",  "refrigerant_emission_factor_unit"),
}

_KNOWN_ENRICHABLE_FIELDS = frozenset({
    "production_emissions", "endoflife_emissions",
    "power_idle", "power_max", "power_source_emission_factor",
    "installation_emission_factor", "maintenance_emission_factor", "life_time",
    "electricity_emission_factor", "fuel_emission_factor", "refrigerant_emission_factor",
})

# ---------------------------------------------------------------------------
# Group uncertainty — SD of key emission fields within each component type
# ---------------------------------------------------------------------------

_GROUP_TYPE_FIELD = {
    "active":         "active_subtype",
    "passive":        "passive_type",
    "infrastructure": "infrastructure_type",
}

_EMBODIED_SD_FIELDS = [
    "production_emissions",
    "endoflife_emissions",
    "installation_emission_factor",
    "maintenance_emission_factor",
]

_POWER_LOAD_LOW  = 0.3
_POWER_LOAD_HIGH = 1.0


def _group_stdev(values: list) -> float:
    """Sample SD (n≥2) of non-None numeric values, else 0.0."""
    nums = [v for v in values if v is not None]
    if len(nums) < 2:
        return 0.0
    try:
        return _stats.stdev(nums)
    except Exception:
        return 0.0


def compute_group_uncertainty(rows: list, schema: str) -> tuple:
    """Add group-level SD columns and power min/max estimates to each row.

    Groups are formed by the type field for each schema.  Power min/max use
    load factors 0.3 (min) and 1.0 (max) in place of the standard 0.8.

    New columns per row:
      production_emissions_sd, endoflife_emissions_sd,
      installation_emission_factor_sd, maintenance_emission_factor_sd
      power_quantity_sd           (active, quantity-path rows, grouped by type+unit)
      power_estimated_low_w       (active, idle/max-path rows)
      power_estimated_high_w      (active, idle/max-path rows)

    Returns (updated_rows, list_of_new_column_names).
    """
    type_field = _GROUP_TYPE_FIELD.get(schema, "")

    # ── Embodied + maintenance SD ────────────────────────────────────────────
    groups: dict = {}
    for row in rows:
        t = str(row.get(type_field, "") or "").strip() or "__unknown__"
        if t not in groups:
            groups[t] = {f: [] for f in _EMBODIED_SD_FIELDS}
        for f in _EMBODIED_SD_FIELDS:
            v = row.get(f)
            if v is not None:
                try:
                    groups[t][f].append(float(v))
                except (ValueError, TypeError):
                    pass

    group_sds = {t: {f: _group_stdev(vals) for f, vals in fv.items()}
                 for t, fv in groups.items()}

    # ── Power quantity SD (active only, grouped by type + unit) ─────────────
    pq_groups: dict = {}
    if schema == "active":
        for row in rows:
            if row.get("power_quantity") is None:
                continue
            t    = str(row.get(type_field, "") or "").strip() or "__unknown__"
            unit = str(row.get("power_unit", "") or "").strip().lower()
            v = row.get("power_quantity")
            if v is not None:
                try:
                    pq_groups.setdefault((t, unit), []).append(float(v))
                except (ValueError, TypeError):
                    pass
    pq_sds = {k: _group_stdev(vals) for k, vals in pq_groups.items()}

    # ── Attach columns to rows ───────────────────────────────────────────────
    for row in rows:
        t   = str(row.get(type_field, "") or "").strip() or "__unknown__"
        sds = group_sds.get(t, {})
        for f in _EMBODIED_SD_FIELDS:
            row[f"{f}_sd"] = sds.get(f, 0.0)

        if schema == "active":
            # Power quantity SD
            unit = str(row.get("power_unit", "") or "").strip().lower()
            row["power_quantity_sd"] = (
                pq_sds.get((t, unit), 0.0)
                if row.get("power_quantity") is not None else None
            )
            # Power estimated min/max
            p_idle_v = row.get("power_idle")
            p_max_v  = row.get("power_max")
            if p_idle_v is not None and p_max_v is not None:
                try:
                    p_idle = float(p_idle_v)
                    p_max  = float(p_max_v)
                    if str(row.get("power_idle_unit", "") or "").strip().lower() == "kw":
                        p_idle *= 1000.0
                    if str(row.get("power_max_unit",  "") or "").strip().lower() == "kw":
                        p_max  *= 1000.0
                    row["power_estimated_low_w"]  = round(
                        p_idle + _POWER_LOAD_LOW  * (p_max - p_idle), 4)
                    row["power_estimated_high_w"] = round(
                        p_idle + _POWER_LOAD_HIGH * (p_max - p_idle), 4)
                except (ValueError, TypeError):
                    row["power_estimated_low_w"]  = None
                    row["power_estimated_high_w"] = None
            else:
                row["power_estimated_low_w"]  = None
                row["power_estimated_high_w"] = None
        else:
            row["power_quantity_sd"]      = None
            row["power_estimated_low_w"]  = None
            row["power_estimated_high_w"] = None

    new_cols = [f"{f}_sd" for f in _EMBODIED_SD_FIELDS]
    if schema == "active":
        new_cols += ["power_quantity_sd", "power_estimated_low_w", "power_estimated_high_w"]

    return rows, new_cols


# ---------------------------------------------------------------------------
# Lifetime defaults lookup — average expected lifetimes in years by type/subtype
# ---------------------------------------------------------------------------

LIFETIME_DEFAULTS = {
    # Active subtypes
    "WLS":                         8,
    "SWITCH":                      8,
    "switch":                      8,
    "KRAFT":                       12,
    "ROUTER":                      8,
    "router":                      8,
    "DC-HÅLLARE":                  15,
    "aggregation_router":          8,
    "AIR":                         8,
    "cellular_modem":              8,
    "chassis":                     10,
    "edge_platform":               8,
    "firewall":                    8,
    "gateway":                     8,
    "radio_unit":                  8,
    "baseband_unit":               8,
    "base_station":                10,
    "ISAM":                        10,
    "memory_card":                 5,
    "antenna":                     15,
    "camera":                      7,
    "sensor":                      7,
    "SFP":                         7,
    "light":                       10,
    "BCI":                         10,
    "DIN/EN":                      15,
    "SLA/VRLA":                    5,
    "standby_power_generator":     15,
    "prime_power_generator":       15,
    "portable_industrial_generator": 10,
    "inverter_generator":          12,
    "container_sized_generator":   15,
    "water_cooled_systems":        15,
    "air_cooled_systems":          15,
    "industrial_chillers":         15,
    "evaporative_cooling_systems": 12,
    "hybrid_system":               12,
    "specialized_cooling":         15,
    # Active types (fallback when no subtype match)
    "generator":           15,
    "cooling":             15,
    "fire_suppression":    15,
    "electrical_equipment": 10,
    # Passive types
    "fiber_cable":         25,
    "electrical_cables":   25,
    "COAX":                20,
    "splitters":           20,
    "shelters":            25,
    "cabinets":            20,
    "plugs":               10,
    "fencing":             25,
    "steel":               30,
    "aluminum":            30,
    "plastic":             15,
    # Infrastructure types
    "tower":               40,
    "mast":                35,
    "rooftop_mount":       25,
    "pole":                25,
    "building":            40,
    "underground":         40,
    "container":           20,
    "real estate":         40,
    "manhole":             40,
    "concrete":            40,
    "ducts & pipes":       30,
}


# ---------------------------------------------------------------------------
# In-memory cache: (provider, field, search_key) -> result dict or {}
# ---------------------------------------------------------------------------

_cache: dict = {}


def _cache_get(provider: str, field: str, search_key: str) -> Optional[dict]:
    return _cache.get((provider, field, search_key))


def _cache_set(provider: str, field: str, search_key: str, result: dict) -> None:
    _cache[(provider, field, search_key)] = result


# ---------------------------------------------------------------------------
# Search key derivation
# ---------------------------------------------------------------------------

def get_search_keys(row: dict, schema: str, field: str) -> list:
    """Return ordered list of search keys to try for this row/field combination."""
    def _vals(*keys):
        return [str(row[k]).strip() for k in keys
                if row.get(k) and str(row.get(k, "")).strip()]

    if field == "life_time":
        if schema == "active":
            return _vals("active_subtype", "active_type")
        elif schema == "passive":
            return _vals("passive_type")
        elif schema == "infrastructure":
            return _vals("infrastructure_type")
    elif field in ("production_emissions", "endoflife_emissions"):
        if schema == "active":
            return _vals("manufacture_part_number", "active_subtype", "active_type", "brand")
        elif schema == "passive":
            return _vals("manufacture_part_number", "passive_type", "brand")
        elif schema == "infrastructure":
            return _vals("manufacture_part_number", "infrastructure_type", "brand")
    elif field in ("power_idle", "power_max"):
        return _vals("manufacture_part_number", "chip_id", "active_subtype", "active_type", "brand")
    elif field == "power_source_emission_factor":
        ps = str(row.get("power_source", "")).strip()
        # Try: direct power_source value, M-prefixed, and field name as id (user convention)
        keys = []
        if ps:
            keys += [ps, f"M-{ps}"]
        keys.append("power_source_emission_factor")
        return keys
    elif field == "electricity_emission_factor":
        src = str(row.get("electricity_source", "")).strip()
        return ([src, f"M-{src}"] if src else []) + ["electricity_emission_factor"]
    elif field == "fuel_emission_factor":
        ft = str(row.get("fuel_type", "")).strip()
        return ([ft, f"M-{ft}"] if ft else []) + ["fuel_emission_factor"]
    elif field == "refrigerant_emission_factor":
        rt = str(row.get("refrigerant_type", "")).strip()
        return ([rt, f"M-{rt}"] if rt else []) + ["refrigerant_emission_factor"]
    elif field == "installation_emission_factor":
        return _vals("installation_method")
    elif field == "maintenance_emission_factor":
        return _vals("maintenance_method")
    return []


# ---------------------------------------------------------------------------
# Precondition check
# ---------------------------------------------------------------------------

def precondition_met(row: dict, field_cfg: dict) -> bool:
    """Return True if the field's precondition is satisfied."""
    precond = field_cfg.get("precondition")
    if precond is not None:
        v = row.get(precond)
        if v is None or str(v).strip() == "":
            return False
    skip_if = field_cfg.get("skip_if")
    if skip_if is not None:
        v = row.get(skip_if)
        if v is not None and str(v).strip() != "":
            return False
    return True


# ---------------------------------------------------------------------------
# Unit validation and conversion
# ---------------------------------------------------------------------------

# Lowercase prefix → canonical prefix (avoids substring-replacement cascades)
_PREFIX_CANON = {
    "kgco2eq":  "kgCO2eq",
    "kgco2e":   "kgCO2eq",
    "kg co2e":  "kgCO2eq",
    "kg co2eq": "kgCO2eq",
    "gco2eq":   "gCO2eq",
    "gco2e":    "gCO2eq",
    "g co2e":   "gCO2eq",
    "g co2eq":  "gCO2eq",
    "tco2eq":   "tCO2eq",
    "tco2e":    "tCO2eq",
    "t co2e":   "tCO2eq",
    "t co2eq":  "tCO2eq",
}

_SCALE = {
    ("gCO2eq", "kgCO2eq"): 0.001,
    ("tCO2eq", "kgCO2eq"): 1000.0,
}


def _normalise_unit(unit_str: str) -> str:
    """Normalise CO2e unit to canonical form, e.g. 'kg CO2e/kWh' → 'kgCO2eq/kWh'."""
    s = str(unit_str).strip()
    if "/" in s:
        prefix, _, suffix = s.partition("/")
        normed = _PREFIX_CANON.get(prefix.strip().lower(), prefix.strip())
        return f"{normed}/{suffix}"
    else:
        return _PREFIX_CANON.get(s.lower(), s)


def validate_unit(value: float, returned_unit: str, expected_unit: str) -> Optional[float]:
    """
    Validate and convert value from returned_unit to expected_unit.
    Returns converted float, or None if units are incompatible.
    """
    nr = _normalise_unit(returned_unit)
    ne = _normalise_unit(expected_unit)
    if nr == ne:
        return float(value)
    # Try prefix scaling (e.g. gCO2eq/unit → kgCO2eq/unit)
    r_parts = nr.split("/", 1)
    e_parts = ne.split("/", 1)
    if len(r_parts) == 2 and len(e_parts) == 2 and r_parts[1] == e_parts[1]:
        scale = _SCALE.get((r_parts[0], e_parts[0]))
        if scale is not None:
            return float(value) * scale
    return None


# ---------------------------------------------------------------------------
# Custom file provider — reads emission_factors.xlsx
# ---------------------------------------------------------------------------

_ef_table: Optional[dict] = None


def _load_ef_table() -> dict:
    """Load emission_factors.xlsx once into memory. Returns lookup dict."""
    global _ef_table
    if _ef_table is not None:
        return _ef_table
    _ef_table = {}
    if not _OPENPYXL_AVAILABLE:
        return _ef_table
    path = Path(cfg.CUSTOM_FILE_PATH)
    if not path.exists():
        return _ef_table
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        # Find header row (first row where column 0 is "id")
        header_row_idx = None
        for i, row in enumerate(all_rows):
            if row and str(row[0]).strip().lower() == "id":
                header_row_idx = i
                break
        if header_row_idx is None:
            return _ef_table
        headers = [str(h).strip().lower() if h else "" for h in all_rows[header_row_idx]]
        _KNOWN_FIELDS = _KNOWN_ENRICHABLE_FIELDS
        for row in all_rows[header_row_idx + 1:]:
            if not row or not row[0]:
                continue
            d = dict(zip(headers, row))
            row_id    = str(d.get("id", "")).strip()
            row_field = str(d.get("field", "")).strip() if d.get("field") else ""
            # Fall back to notes column when field is empty and notes matches a known field name
            if not row_field:
                row_notes = str(d.get("notes", "")).strip() if d.get("notes") else ""
                if row_notes in _KNOWN_FIELDS:
                    row_field = row_notes
            row_country = str(d.get("country", "")).strip() if d.get("country") else ""
            ef        = d.get("emission_factor")
            unit      = str(d.get("unit", "")).strip() if d.get("unit") else ""
            if not row_id or ef is None:
                continue
            try:
                entry = {"value": float(ef), "unit": unit, "source": "custom_file"}
                # Key: (id, field, country) — country="" or field="" acts as wildcard
                # Country-specific rows are stored only under their country key;
                # rows without a country are stored under "" and act as the generic fallback.
                _ef_table[(row_id, row_field, row_country)] = entry
            except (ValueError, TypeError):
                pass  # non-numeric emission_factor = user hasn't filled it yet
    except Exception as exc:
        print(f"  [WARNING] Could not read custom file: {exc}")
    return _ef_table


def _fetch_custom_file(search_key: str, field: str, country: str = "") -> dict:
    """Look up search_key in emission_factors.xlsx.

    Priority order:
      1. (id, field, country)  — country + field specific
      2. (id, field, "")       — field specific, no country
      3. (id, "", country)     — country specific, any field
      4. (id, "", "")          — fully generic
    """
    table = _load_ef_table()
    c = country or ""
    return (
        table.get((search_key, field, c)) or
        table.get((search_key, field, "")) or
        table.get((search_key, "", c)) or
        table.get((search_key, "", "")) or
        {}
    )


# ---------------------------------------------------------------------------
# Rejoose provider
# ---------------------------------------------------------------------------

def _fetch_rejoose(search_key: str, field: str) -> dict:
    if not _REQUESTS_AVAILABLE:
        return {}
    key = _env("REJOOSE_API_KEY")
    if not key:
        return {}
    headers = {"Authorization": f"Bearer {key}"}
    try:
        resp = requests.get(
            cfg.REJOOSE_BASE_URL.rstrip("/") + "/api/v3/products",
            headers=headers,
            params={"mpn": search_key},
            timeout=10,
        )
        if resp.ok:
            data = resp.json()
            ef   = data.get("total_co2e") or data.get("co2e") or data.get("carbon_footprint")
            unit = data.get("unit", "kgCO2eq/unit")
            if ef is not None:
                return {"value": float(ef), "unit": unit, "source": "rejoose"}
    except Exception:
        pass
    return {}


# ---------------------------------------------------------------------------
# Climatiq provider
# ---------------------------------------------------------------------------

def _fetch_climatiq(search_key: str, field: str) -> dict:
    if not _REQUESTS_AVAILABLE:
        return {}
    key = _env("CLIMATIQ_API_KEY")
    if not key:
        return {}
    headers = {"Authorization": f"Bearer {key}"}
    try:
        resp = requests.get(
            cfg.CLIMATIQ_BASE_URL.rstrip("/") + "/data/v1/search",
            headers=headers,
            params={"query": search_key, "data_version": cfg.CLIMATIQ_DATA_VERSION},
            timeout=10,
        )
        if resp.ok:
            results = resp.json().get("results", [])
            if results:
                ef   = results[0].get("factor")
                unit = results[0].get("unit", "kgCO2eq/unit")
                if ef is not None:
                    return {"value": float(ef), "unit": unit, "source": "climatiq"}
    except Exception:
        pass
    return {}


# ---------------------------------------------------------------------------
# EPD provider — placeholder
# ---------------------------------------------------------------------------

def _fetch_epd(search_key: str, field: str) -> dict:
    # TODO: Implement when EPD API details are confirmed.
    # Base URL: cfg.EPD_BASE_URL  ("https://epd.apim.developer.azure-api.net")
    # Auth:     Ocp-Apim-Subscription-Key header — set EPD_API_KEY in .env
    return {}


# ---------------------------------------------------------------------------
# Resilio provider — placeholder
# ---------------------------------------------------------------------------

def _fetch_resilio(search_key: str, field: str) -> dict:
    # TODO: Implement when Resilio API details are confirmed.
    # Base URL: cfg.RESILIO_BASE_URL  ("https://db.resilio.tech")
    # Auth:     set RESILIO_API_KEY in .env
    return {}


# ---------------------------------------------------------------------------
# Ecoinvent provider
# ---------------------------------------------------------------------------

_ecoinvent_token: Optional[str] = None


def _get_ecoinvent_token() -> Optional[str]:
    global _ecoinvent_token
    if _ecoinvent_token:
        return _ecoinvent_token
    client_id     = _env("ECOINVENT_CLIENT_ID")
    client_secret = _env("ECOINVENT_CLIENT_SECRET")
    if not client_id or not client_secret:
        return None
    try:
        resp = requests.post(
            cfg.ECOINVENT_TOKEN_URL,
            data={
                "grant_type":    "client_credentials",
                "client_id":     client_id,
                "client_secret": client_secret,
            },
            timeout=10,
        )
        if resp.ok:
            _ecoinvent_token = resp.json().get("access_token")
            return _ecoinvent_token
    except Exception:
        pass
    return None


def _fetch_ecoinvent(search_key: str, field: str) -> dict:
    if not _REQUESTS_AVAILABLE:
        return {}
    token = _get_ecoinvent_token()
    if not token:
        return {}
    headers = {"Authorization": f"Bearer {token}"}
    try:
        search = requests.get(
            cfg.ECOINVENT_BASE_URL.rstrip("/") + "/datasets",
            headers=headers,
            params={"query": search_key},
            timeout=10,
        )
        if not search.ok:
            return {}
        datasets = search.json().get("datasets", [])
        if not datasets:
            return {}
        ds_id    = datasets[0].get("id")
        ind_resp = requests.get(
            cfg.ECOINVENT_BASE_URL.rstrip("/") + f"/datasets/{ds_id}/indicators",
            headers=headers,
            timeout=10,
        )
        if not ind_resp.ok:
            return {}
        for indicator in ind_resp.json().get("indicators", []):
            name = indicator.get("name", "")
            if "GWP" in name or "CO2" in name or "climate change" in name.lower():
                ef   = indicator.get("value")
                unit = indicator.get("unit", "kgCO2eq/unit")
                if ef is not None:
                    return {"value": float(ef), "unit": unit, "source": "ecoinvent"}
    except Exception:
        pass
    return {}


# ---------------------------------------------------------------------------
# tscircuit provider — placeholder (power fields only)
# ---------------------------------------------------------------------------

def _fetch_tscircuit(search_key: str, field: str) -> dict:
    # TODO: Implement when tscircuit API details are confirmed.
    # Base URL: cfg.TSCIRCUIT_BASE_URL  ("https://api.tscircuit.com")
    # Searches by chip_id / part name; returns power_idle and power_max in watts.
    # Set TSCIRCUIT_API_KEY in .env if auth is required.
    return {}


# ---------------------------------------------------------------------------
# Lifetime defaults provider — built-in lookup table
# ---------------------------------------------------------------------------

def _fetch_lifetime_defaults(search_key: str, field: str) -> dict:
    """Return average lifetime in years from the built-in LIFETIME_DEFAULTS table."""
    years = LIFETIME_DEFAULTS.get(search_key)
    if years is not None:
        return {"value": years, "unit": "years", "source": "lifetime_defaults"}
    return {}


# ---------------------------------------------------------------------------
# Provider dispatch tables
# ---------------------------------------------------------------------------

_EMISSION_PROVIDERS = {
    "custom_file": _fetch_custom_file,
    "rejoose":     _fetch_rejoose,
    "climatiq":    _fetch_climatiq,
    "epd":         _fetch_epd,
    "resilio":     _fetch_resilio,
    "ecoinvent":   _fetch_ecoinvent,
}

_POWER_PROVIDERS = {
    "tscircuit": _fetch_tscircuit,
}

_LIFETIME_PROVIDERS = {
    "lifetime_defaults": _fetch_lifetime_defaults,
}


def _get_provider_fn(provider: str, field_cfg: dict):
    group = field_cfg.get("provider_group", "emission")
    if group == "power":
        # custom_file serves all groups; group-specific providers checked first
        return _POWER_PROVIDERS.get(provider) or _EMISSION_PROVIDERS.get(provider)
    if group == "lifetime":
        return _LIFETIME_PROVIDERS.get(provider)
    return _EMISSION_PROVIDERS.get(provider)


# ---------------------------------------------------------------------------
# Cached provider fetch
# ---------------------------------------------------------------------------

def _fetch_from_provider(provider: str, search_key: str, field: str, field_cfg: dict,
                         country: str = "") -> dict:
    cached = _cache_get(provider, field, search_key)
    if cached is not None:
        return cached
    fn = _get_provider_fn(provider, field_cfg)
    if fn is None:
        return {}
    result = (_fetch_custom_file(search_key, field, country)
              if provider == "custom_file"
              else fn(search_key, field))
    if result:  # only cache hits — misses may succeed on a later run with more data
        _cache_set(provider, field, search_key, result)
    return result


# ---------------------------------------------------------------------------
# Row enrichment
# ---------------------------------------------------------------------------

def enrich_row(row: dict, schema: str, provider: str, field_cfgs: list,
               filled_log: list = None) -> int:
    """
    Try to fill empty enrichable fields in row using the given provider.
    Modifies row in-place. Returns count of fields filled.

    filled_log: if provided, appends (field, search_key, value) for each fill.
    """
    filled = 0
    for fc in field_cfgs:
        fname      = fc["field"]
        unit_field = fc.get("unit_field")

        # Skip already-filled fields
        v = row.get(fname)
        if v is not None and str(v).strip():
            continue
        # Check precondition
        if not precondition_met(row, fc):
            continue
        # Derive expected unit; None means this field carries no unit (e.g. life_time)
        if unit_field is not None:
            expected_unit = row.get(unit_field) or fc.get("default_unit")
            if not expected_unit:
                continue  # unit required but not derivable — skip
        else:
            expected_unit = None

        # Try each search key in priority order
        country = str(row.get("country", "")).strip()
        for sk in get_search_keys(row, schema, fname):
            result = _fetch_from_provider(provider, sk, fname, fc, country)
            if not result:
                continue
            if expected_unit is not None:
                converted = validate_unit(result["value"], result.get("unit", ""), expected_unit)
                if converted is None:
                    continue
                row[fname]      = converted
                row[unit_field] = expected_unit
            else:
                row[fname] = result["value"]

            row[f"{fname}_source"]     = result["source"]
            row[f"{fname}_confidence"] = CONFIDENCE.get(result["source"], 0.5)
            if filled_log is not None:
                filled_log.append((fname, sk, row[fname]))
            filled += 1
            break
    return filled


# ---------------------------------------------------------------------------
# Manual fallback — append unresolved entries to emission_factors.xlsx
# ---------------------------------------------------------------------------

def _append_manual_fallback(unresolved: list) -> None:
    """
    Append rows with blank emission_factor to emission_factors.xlsx.
    unresolved = list of (search_key, field, expected_unit), deduplicated.
    """
    if not _OPENPYXL_AVAILABLE or not unresolved:
        return
    path = Path(cfg.CUSTOM_FILE_PATH)
    if not path.exists():
        print(f"  [WARNING] Cannot write manual fallback — {path} not found.")
        return
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))

        # Locate header row
        header_row_num = None   # 1-based
        headers_lower  = []
        for i, row in enumerate(all_rows):
            if row and str(row[0]).strip().lower() == "id":
                header_row_num = i + 1
                headers_lower  = [str(h).strip().lower() if h else "" for h in row]
                break
        if header_row_num is None:
            return

        # Add 'field' column to header if missing
        if "field" not in headers_lower:
            field_col = len(headers_lower) + 1
            ws.cell(row=header_row_num, column=field_col).value = "field"
            headers_lower.append("field")
        else:
            field_col = headers_lower.index("field") + 1

        id_col   = headers_lower.index("id") + 1
        ef_col   = headers_lower.index("emission_factor") + 1
        unit_col = headers_lower.index("unit") + 1

        # notes column — field name goes here (user's preferred format)
        notes_col = headers_lower.index("notes") + 1 if "notes" in headers_lower else None

        # Build set of already-present (id, field) pairs — check both notes and field columns
        _KNOWN_FIELDS = _KNOWN_ENRICHABLE_FIELDS
        existing = set()
        for row in all_rows[header_row_num:]:
            if row and row[0]:
                fval = row[field_col - 1] if len(row) >= field_col else None
                fstr = str(fval).strip() if fval else ""
                if not fstr and notes_col and len(row) >= notes_col:
                    nval = str(row[notes_col - 1]).strip() if row[notes_col - 1] else ""
                    if nval in _KNOWN_FIELDS:
                        fstr = nval
                existing.add((str(row[0]).strip(), fstr))

        next_row = ws.max_row + 1
        added    = 0
        for sk, field, unit in unresolved:
            if (sk, field) in existing:
                continue
            ws.cell(row=next_row, column=id_col).value    = sk
            ws.cell(row=next_row, column=ef_col).value    = None   # blank for user
            ws.cell(row=next_row, column=unit_col).value  = unit
            if notes_col:
                ws.cell(row=next_row, column=notes_col).value = field  # field name in notes column
            ws.cell(row=next_row, column=field_col).value = field
            existing.add((sk, field))
            next_row += 1
            added    += 1

        wb.save(path)
        # Invalidate cache so next run re-reads the file with any filled values
        global _ef_table
        _ef_table = None

        if added:
            print(f"\n  {added} unresolved field(s) added to {path.name}.")
            print(f"  Fill in the 'emission_factor' column and re-run the pipeline.")
    except Exception as exc:
        print(f"  [WARNING] Could not write manual fallback: {exc}")


# ---------------------------------------------------------------------------
# Cell site EF inheritance
# ---------------------------------------------------------------------------

def _enrich_from_cell_site(data_rows: list, cell_site_lookup: dict) -> int:
    """Propagate electricity/fuel/refrigerant EFs from parent cell site to component rows.

    Fills power_source_emission_factor (and its unit) for rows that:
      - have a cell_site_id matching a site in the lookup
      - have a power_source mapped in _POWER_SOURCE_TO_SITE_EF
      - do not already have a power_source_emission_factor value

    Runs before the provider hierarchy — second only to explicit row values.
    Returns count of fields filled.
    """
    filled = 0
    for row in data_rows:
        if row.get("power_source_emission_factor") is not None and \
                str(row.get("power_source_emission_factor", "")).strip():
            continue
        cs_id = str(row.get("cell_site_id", "") or "").strip()
        if not cs_id:
            continue
        site = cell_site_lookup.get(cs_id)
        if not site:
            continue
        ps = str(row.get("power_source", "") or "").strip().lower()
        ef_info = _POWER_SOURCE_TO_SITE_EF.get(ps)
        if not ef_info:
            continue
        ef_field, unit_field = ef_info
        ef_val = None
        raw = site.get(ef_field)
        if raw is not None:
            try:
                ef_val = float(raw)
            except (ValueError, TypeError):
                pass
        if ef_val is None:
            continue
        row["power_source_emission_factor"]            = ef_val
        row["power_source_emission_factor_unit"]       = site.get(unit_field) or ""
        row["power_source_emission_factor_source"]     = "cell_site"
        row["power_source_emission_factor_confidence"] = CONFIDENCE["cell_site"]
        filled += 1
    return filled


def _load_cell_site_lookup(data_dir: Path) -> dict:
    """Load cell_site.xlsx from data_dir and return {cell_site_id: row_dict}."""
    if not _OPENPYXL_AVAILABLE:
        return {}
    path = data_dir / "cell_site.xlsx"
    if not path.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 2:
            return {}
        headers = list(all_rows[0])
        lookup  = {}
        for row in all_rows[1:]:
            d     = dict(zip(headers, row))
            cs_id = str(d.get("cell_site_id", "") or "").strip()
            if cs_id:
                lookup[cs_id] = d
        return lookup
    except Exception as exc:
        print(f"  [WARNING] Could not load cell_site.xlsx for EF inheritance: {exc}")
        return {}


# ---------------------------------------------------------------------------
# Main enrichment function
# ---------------------------------------------------------------------------

def _run_enrichment(data_rows: list, orig_headers: list, schema: str, out_path: Path,
                    cell_site_lookup: dict = None) -> list:
    """
    Core enrichment logic shared by enrich_file and enrich_rows.
    Modifies data_rows in-place, writes out_path, returns data_rows.
    """
    field_cfgs    = ENRICHABLE_FIELDS.get(schema, [])
    emission_cfgs = [fc for fc in field_cfgs if fc["provider_group"] == "emission"]
    power_cfgs    = [fc for fc in field_cfgs if fc["provider_group"] == "power"]
    lifetime_cfgs = [fc for fc in field_cfgs if fc["provider_group"] == "lifetime"]

    # Cell site EF inheritance — runs before all other providers
    if cell_site_lookup:
        cs_filled = _enrich_from_cell_site(data_rows, cell_site_lookup)
        if cs_filled:
            conf = CONFIDENCE["cell_site"]
            print(f"\n  [cell_site]  confidence {conf:.0%}  — {cs_filled} field fill(s):")
            print(f"    power_source_emission_factor  ← inherited from parent cell site")

    # Provider-first fill: each provider runs across all rows before the next.
    all_fills: dict = {}   # provider -> list of (field, search_key, value)

    for provider in cfg.LIFETIME_HIERARCHY:
        log: list = []
        for row in data_rows:
            enrich_row(row, schema, provider, lifetime_cfgs, filled_log=log)
        if log:
            all_fills[provider] = log

    for provider in cfg.EMISSION_HIERARCHY:
        log = []
        for row in data_rows:
            enrich_row(row, schema, provider, emission_cfgs, filled_log=log)
        if log:
            all_fills[provider] = log

    for provider in cfg.POWER_HIERARCHY:
        log = []
        for row in data_rows:
            enrich_row(row, schema, provider, power_cfgs, filled_log=log)
        if log:
            all_fills[provider] = log

    if all_fills:
        for provider, log in all_fills.items():
            conf = CONFIDENCE.get(provider, 0.5)
            print(f"\n  [{provider}]  confidence {conf:.0%}  — {len(log)} field fill(s):")
            from collections import Counter
            counts = Counter((f, sk, v) for f, sk, v in log)
            for (fname, sk, val), cnt in sorted(counts.items(), key=lambda x: (x[0][0], x[0][1])):
                unit_label = " yr" if fname == "life_time" else ""
                suffix     = f" (×{cnt})" if cnt > 1 else ""
                print(f"    {fname:<30}  {sk:<30} → {val}{unit_label}{suffix}")
    else:
        print("\n  No fields enriched.")

    # Collect unresolved fields for manual fallback (emission fields only —
    # lifetime fields use a built-in table and never need manual input)
    unresolved_seen = set()
    unresolved      = []
    for row in data_rows:
        for fc in field_cfgs:
            if fc.get("provider_group") == "lifetime":
                continue
            fname = fc["field"]
            v = row.get(fname)
            if v is not None and str(v).strip():
                continue
            if not precondition_met(row, fc):
                continue
            search_keys = get_search_keys(row, schema, fname)
            if not search_keys:
                continue
            sk            = search_keys[0]
            unit_field    = fc.get("unit_field")
            expected_unit = (row.get(unit_field) if unit_field else None) or fc.get("default_unit") or ""
            key           = (sk, fname)
            if key not in unresolved_seen:
                unresolved_seen.add(key)
                unresolved.append((sk, fname, expected_unit))

    if unresolved:
        _append_manual_fallback(unresolved)

    data_rows, unc_cols = compute_group_uncertainty(data_rows, schema)
    _write_enriched(orig_headers, data_rows, field_cfgs, out_path, uncertainty_cols=unc_cols)
    print(f"  Written: {out_path.name}")

    return data_rows


def enrich_file(filepath: str, schema: str) -> list:
    """
    Enrich all rows in filepath for the given schema.

    Args:
        filepath: Path to the source Excel file (must have a 'Data' sheet).
        schema:   One of "active", "passive", "infrastructure".

    Returns:
        List of enriched row dicts (one per non-empty data row).
        Writes data/<name>_enriched.xlsx alongside the source file.
    """
    if not _OPENPYXL_AVAILABLE:
        print("  [WARNING] openpyxl not available — skipping enrichment.")
        return []

    path = Path(filepath)
    if not path.exists():
        print(f"  [WARNING] File not found: {path}")
        return []

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Data" not in wb.sheetnames:
        print(f"  [WARNING] No 'Data' sheet in {path.name}")
        return []

    ws       = wb["Data"]
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 2:
        return []

    headers   = list(all_rows[0])
    data_rows = [
        dict(zip(headers, row))
        for row in all_rows[1:]
        if any(v is not None and str(v).strip() for v in row)
    ]

    cell_site_lookup = _load_cell_site_lookup(path.parent)
    out_path = path.parent / (path.stem + "_enriched" + path.suffix)
    return _run_enrichment(data_rows, headers, schema, out_path,
                           cell_site_lookup=cell_site_lookup)


def enrich_rows(rows: list, schema: str, out_filename: str,
                cell_sites: list = None) -> list:
    """
    Enrich a list of row dicts (e.g. fetched from telecom_api) for the given schema.

    Args:
        rows:         List of row dicts, as returned by the API GET endpoints.
        schema:       One of "active", "passive", "infrastructure".
        out_filename: Base name for the output file (without extension or _enriched suffix).
                      Output is written to data/<out_filename>_enriched.xlsx.
        cell_sites:   Optional list of cell site dicts for EF inheritance. When provided,
                      power_source_emission_factor is inherited from the parent cell site
                      before falling through to the provider hierarchy.

    Returns:
        List of enriched row dicts.
    """
    if not _OPENPYXL_AVAILABLE:
        print("  [WARNING] openpyxl not available — skipping enrichment.")
        return list(rows)
    if not rows:
        print("  No rows to enrich.")
        return []

    # Derive column order from union of all row keys (preserving first-seen order)
    seen: list = []
    seen_set: set = set()
    for row in rows:
        for k in row:
            if k not in seen_set:
                seen.append(k)
                seen_set.add(k)

    data_rows        = [dict(row) for row in rows]  # copy so we don't mutate caller's data
    cell_site_lookup = ({str(s.get("cell_site_id", "") or "").strip(): s
                         for s in cell_sites if s.get("cell_site_id")}
                        if cell_sites else {})
    out_path = DATA_DIR / (out_filename + "_enriched.xlsx")
    return _run_enrichment(data_rows, seen, schema, out_path,
                           cell_site_lookup=cell_site_lookup)


def _write_enriched(orig_headers: list, data_rows: list, field_cfgs: list, out_path: Path,
                    uncertainty_cols: list = None) -> None:
    """Write enriched rows to out_path; source/confidence and uncertainty columns appended."""
    if not data_rows:
        return
    extra_cols = []
    for fc in field_cfgs:
        extra_cols.append(f"{fc['field']}_source")
        extra_cols.append(f"{fc['field']}_confidence")
    if uncertainty_cols:
        extra_cols += list(uncertainty_cols)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(list(orig_headers) + extra_cols)
    for row in data_rows:
        out_row = [row.get(h) for h in orig_headers]
        out_row += [row.get(c) for c in extra_cols]
        ws.append(out_row)

    wb.save(out_path)


# ---------------------------------------------------------------------------
# Backward-compatible public interface
# ---------------------------------------------------------------------------

def _lookup_ids(product_id: str, method_ids: list) -> list:
    seen, result = set(), []
    for uid in [product_id] + list(method_ids):
        if uid and uid not in seen:
            seen.add(uid)
            result.append(uid)
    return result


def enrich_from_api(product_id: str, method_ids: list) -> dict:
    """
    Legacy interface — try each emission provider in hierarchy order.
    Returns the first successful result as {"emission_factor": float,
    "emission_data_source": str}, or {} if none match.
    """
    for provider_name in cfg.EMISSION_HIERARCHY:
        fn = _EMISSION_PROVIDERS.get(provider_name)
        if fn is None:
            continue
        for uid in _lookup_ids(product_id, list(method_ids)):
            result = fn(uid, "emission_factor")
            if result:
                return {
                    "emission_factor":      result.get("value"),
                    "emission_data_source": result.get("source"),
                }
    return {}


def report_unmatched(products: dict) -> None:
    """
    Print a warning for every product that has no emission factor after enrichment.
    """
    unmatched = [p for p in products.values() if p.get("emission_factor") is None]
    if not unmatched:
        return
    print("\n" + "!" * 60)
    print("  WARNING — UNMATCHED EMISSION FACTORS")
    print("!" * 60)
    print(f"\n  {len(unmatched)} product(s) could not be matched to an emission factor\n")
    print(f"  {'Product ID':<12}  {'Tried IDs'}")
    print(f"  {'-'*10}  {'-'*42}")
    for p in sorted(unmatched, key=lambda x: x["product_id"]):
        method_ids = sorted({
            m for m in [p.get("install_method"), p.get("maint_method"), p.get("elec_method")]
            if m and m != "—"
        })
        tried = ", ".join([p["product_id"]] + method_ids)
        print(f"  {p['product_id']:<12}  {tried}")
    print()
