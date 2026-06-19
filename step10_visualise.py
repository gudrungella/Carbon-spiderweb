"""
step10_visualise.py — HTML emissions report generator.

Reads data/emissions_report.xlsx (produced by step8_analyze.py) and
data/cell_site.xlsx, writes data/emissions_report.html.

Usage:
    python3 step10_visualise.py
"""

from collections import Counter
from pathlib import Path
import datetime
import json
import os

try:
    import openpyxl
except ImportError:
    raise SystemExit("ERROR: openpyxl is required.  pip install openpyxl")

DATA_DIR = Path(__file__).parent / "data"

# ── Colours ───────────────────────────────────────────────────────────────────

_C_EMB = "#3805E3"   # electric blue  — cradle-to-site
_C_EOL = "#7A9CDC"   # periwinkle     — end-of-life
_C_OP  = "#07011E"   # dark navy      — operational
_C_S1  = "#A80A02"   # red accent     — Scope 1 (direct)
_C_S2  = "#3805E3"   # electric blue  — Scope 2 (electricity)
_C_S3  = "#7A9CDC"   # periwinkle     — Scope 3 (indirect)

# "" maps to light blue-grey for sites with no recorded site_type
_SITE_TYPE_COLOURS = {
    "small": "#7A9CDC", "medium": "#3805E3", "large": "#07011E", "": "#D5E5E6",
}

_SCOPE_MAP = {"electricity": 2, "battery": 2, "fuel": 1, "refrigerant": 1}

_C_WARN = "#B45309"   # amber — data quality warnings

# Enrichment source → confidence score (mirrors step4_enrich.CONFIDENCE)
_SOURCE_CONFIDENCE = {
    "custom_file":       1.0,
    "cell_site":         1.0,
    "rejoose":           0.95,
    "climatiq":          0.9,
    "tscircuit":         0.9,
    "manual":            0.9,
    "epd":               0.85,
    "resilio":           0.8,
    "ecoinvent":         0.75,
    "lifetime_defaults": 0.5,
}

# Maps flag substrings → LCA data-quality category; unmatched → "Completeness"
# Unit mismatches, unrecognised values and logically inverted power values are
# internal consistency issues (within a row), not external validity concerns.
_FLAG_LCA_CATEGORY = [
    ("unit mismatch",            "Consistency"),
    ("not recognised",           "Consistency"),
    ("power_max < power_idle",   "Consistency"),
    ("not applicable",           "Info"),
    ("calculated at site level", "Info"),
    ("decommissioned",           "Info"),
    ("scaled by",                "Info"),
]

# Key enrichable fields to track per schema for coverage / consistency / validity
_KEY_FIELDS = {
    "active":         ["production_emissions", "endoflife_emissions",
                       "power_source_emission_factor", "power_idle", "power_max"],
    "passive":        ["production_emissions", "endoflife_emissions",
                       "installation_emission_factor"],
    "infrastructure": ["production_emissions", "endoflife_emissions",
                       "installation_emission_factor"],
}
_TYPE_FIELD = {
    "active":         "active_subtype",
    "passive":        "passive_type",
    "infrastructure": "infrastructure_type",
}

# Groups for the completeness display.
# "check" is a field name, except "operational" which triggers a multi-field check.
_COV_GROUPS = {
    "active": [
        {"label": "Embodied emissions",    "check": "production_emissions"},
        {"label": "Installation",          "check": "installation_emission_factor"},
        {"label": "Power data",            "check": "operational"},
        {"label": "Maintenance",           "check": "maintenance_emission_factor"},
        {"label": "End-of-life emissions", "check": "endoflife_emissions"},
    ],
    "passive": [
        {"label": "Embodied emissions",    "check": "production_emissions"},
        {"label": "Installation",          "check": "installation_emission_factor"},
        {"label": "Maintenance",           "check": "maintenance_emission_factor"},
        {"label": "End-of-life emissions", "check": "endoflife_emissions"},
    ],
    "infrastructure": [
        {"label": "Embodied emissions",    "check": "production_emissions"},
        {"label": "Installation",          "check": "installation_emission_factor"},
        {"label": "Maintenance",           "check": "maintenance_emission_factor"},
        {"label": "End-of-life emissions", "check": "endoflife_emissions"},
    ],
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def _f(v) -> float:
    if v is None or (isinstance(v, str) and ("→" in v or not v.strip())):
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _pct(part, total) -> float:
    return round(100 * part / total, 1) if total else 0.0


def _iter_sheet(filename, sheet_name=None):
    path = DATA_DIR / filename
    if not path.exists():
        raise FileNotFoundError(f"{path} — run step8_analyze.py first")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return
    headers = rows[0]
    for row in rows[1:]:
        if any(v is not None and str(v).strip() for v in row):
            yield dict(zip(headers, row))


def _parse_summary(filename):
    """Parse the three-section Summary sheet; returns {section_title: [row_dict]}."""
    path = DATA_DIR / filename
    wb  = openpyxl.load_workbook(path, data_only=True)
    ws  = wb["Summary"]
    _KNOWN = {"By Component Type", "By Network Type", "By Cell Site"}
    result = {}
    headers = []
    current = None
    for row in ws.iter_rows(values_only=True):
        v0 = str(row[0] or "").strip()
        if v0 in _KNOWN:
            current = v0
            headers = []
            result[current] = []
        elif current and not headers:
            headers = [str(c) for c in row if c is not None]
        elif current and headers:
            if all(c is None or str(c).strip() == "" for c in row):
                current = None
                headers = []
                if len(result) == len(_KNOWN):
                    break
            else:
                result[current].append(dict(zip(headers, row)))
    return result


# ── Data loading ──────────────────────────────────────────────────────────────

def _load_sensitivity() -> list:
    path = DATA_DIR / "sensitivity_results.json"
    if not path.exists():
        return []
    try:
        with open(path, encoding="utf-8") as fh:
            return json.load(fh)
    except Exception:
        return []


def _load_uncertainty() -> dict:
    path = DATA_DIR / "uncertainty_results.json"
    if not path.exists():
        return {}
    try:
        with open(path, encoding="utf-8") as fh:
            return json.load(fh)
    except Exception:
        return {}


def load_all():
    summary     = _parse_summary("emissions_report.xlsx")
    op_rows     = list(_iter_sheet("emissions_report.xlsx", "Operational"))
    emb_rows    = list(_iter_sheet("emissions_report.xlsx", "Embodied"))
    cell_sites  = list(_iter_sheet("cell_site.xlsx", "Data"))
    sensitivity  = _load_sensitivity()
    uncertainty  = _load_uncertainty()
    return summary, op_rows, emb_rows, cell_sites, sensitivity, uncertainty


# ── Derived data ──────────────────────────────────────────────────────────────

def _ghg_scopes(op_rows, emb_rows, cell_sites):
    s = {1: 0.0, 2: 0.0, 3: 0.0}

    for row in op_rows:
        if row.get("power_path") == "site_measured":
            continue
        ps = str(row.get("power_source", "") or "").strip().lower()
        s[_SCOPE_MAP.get(ps, 2)] += _f(row.get("op_energy"))
        # op_maintenance → Scope 1 (direct operational activity)
        s[1] += _f(row.get("op_maintenance"))

    for site in cell_sites:
        me = _f(site.get("measured_electricity"))
        ef = _f(site.get("electricity_emission_factor"))
        if me > 0 and ef > 0:
            s[2] += me * ef
        mf = _f(site.get("measured_fuel"))
        ef = _f(site.get("fuel_emission_factor"))
        if mf > 0 and ef > 0:
            s[1] += mf * ef
        mr = _f(site.get("measured_refrigerant"))
        ef = _f(site.get("refrigerant_emission_factor"))
        if mr > 0 and ef > 0:
            s[1] += mr * ef

    upstream   = sum(_f(r.get("cradle_to_site_annual")) for r in emb_rows)
    downstream = sum(_f(r.get("eol_emissions_annual"))  for r in emb_rows)
    s[3] = upstream + downstream

    return (
        {k: round(v, 1) for k, v in s.items()},
        {"upstream": round(upstream, 1), "downstream": round(downstream, 1)},
    )


def _flags_summary(op_rows, emb_rows):
    all_keys = []
    flagged  = 0
    total    = len(op_rows) + len(emb_rows)
    for row in list(op_rows) + list(emb_rows):
        f = str(row.get("flags", "") or "").strip()
        if f:
            flagged += 1
            for part in f.split(";"):
                p = part.strip()
                if p:
                    all_keys.append(p.split("—")[0].split("(")[0].strip())
    top3 = [flag for flag, _ in Counter(all_keys).most_common(3)]
    return flagged, total, top3


def _lca_category(flag: str) -> str:
    """Map a flag string to its LCA data-quality category."""
    fl = flag.lower()
    for substr, cat in _FLAG_LCA_CATEGORY:
        if substr.lower() in fl:
            return cat
    return "Completeness"


def _flag_severity(flag: str) -> str:
    """Classify a flag string as error / warning / info."""
    fl = flag.lower()
    if "could not be calculated" in fl or "no embodied input data provided" in fl:
        return "error"
    if any(p in fl for p in ("not applicable", "calculated at site level",
                              "decommissioned", "scaled by")):
        return "info"
    return "warning"


def _load_enriched_rows() -> dict:
    """Load enriched xlsx files; returns {schema: [row_dict]}."""
    _FILES = {
        "active":         "active_components_enriched.xlsx",
        "passive":        "passive_components_enriched.xlsx",
        "infrastructure": "infrastructure_enriched.xlsx",
    }
    result = {}
    for schema, fname in _FILES.items():
        path = DATA_DIR / fname
        if not path.exists():
            result[schema] = []
            continue
        try:
            wb   = openpyxl.load_workbook(path, data_only=True)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
            hdrs = list(rows[0]) if rows else []
            result[schema] = (
                [dict(zip(hdrs, r)) for r in rows[1:] if any(v for v in r)]
                if len(rows) > 1 else []
            )
        except Exception:
            result[schema] = []
    return result


def _has_val(row: dict, field: str) -> bool:
    v = row.get(field)
    return v is not None and str(v).strip() != ""


def _field_coverage(enriched: dict, measured_sites: set = None) -> dict:
    """Per-schema grouped coverage using _COV_GROUPS.

    Returns {schema: [{"label", "present", "total"}, ...]}.
    The "operational" check for active components counts rows where the emission
    factor is present AND at least one power data source is available
    (power_quantity, idle+max pair, or site-level measured consumption).
    """
    measured_sites = measured_sites or set()
    result = {}
    for schema, rows in enriched.items():
        if not rows:
            continue
        schema_cov = []
        for g in _COV_GROUPS.get(schema, []):
            if g["check"] == "operational":
                present = sum(
                    1 for r in rows
                    if _has_val(r, "power_source_emission_factor") and (
                        _has_val(r, "power_quantity") or
                        (_has_val(r, "power_idle") and _has_val(r, "power_max")) or
                        str(r.get("cell_site_id", "") or "").strip() in measured_sites
                    )
                )
            else:
                present = sum(1 for r in rows if _has_val(r, g["check"]))
            schema_cov.append({"label": g["label"], "present": present, "total": len(rows)})
        result[schema] = schema_cov
    return result


def _consistency_check(enriched: dict) -> list:
    """
    For each type group (≥2 rows), detect fields present for some items but
    not others (0 < coverage < 100%). Returns list of issue dicts.
    """
    issues = []
    for schema, rows in enriched.items():
        if not rows:
            continue
        tf     = _TYPE_FIELD.get(schema, "")
        groups: dict = {}
        for row in rows:
            t = str(row.get(tf, "") or "unknown").strip()
            groups.setdefault(t, []).append(row)
        for type_name, group_rows in groups.items():
            if len(group_rows) < 2:
                continue
            for field in _KEY_FIELDS.get(schema, []):
                present = sum(
                    1 for r in group_rows
                    if r.get(field) is not None and str(r.get(field)).strip()
                )
                total = len(group_rows)
                if 0 < present < total:
                    issues.append({"schema": schema, "type": type_name,
                                   "field": field, "present": present, "total": total})
    return issues


def _validity_summary(enriched: dict) -> dict:
    """Aggregate enrichment sources from *_source columns; return source names + scores."""
    source_counts: dict = {}
    for schema, rows in enriched.items():
        if not rows:
            continue
        for field in _KEY_FIELDS.get(schema, []):
            src_col = f"{field}_source"
            for row in rows:
                if row.get(field) is None or not str(row.get(field)).strip():
                    continue
                src = str(row.get(src_col) or "").strip() or "Direct input"
                source_counts[src] = source_counts.get(src, 0) + 1

    source_summary = sorted(
        [{"source": src, "count": cnt, "confidence": _SOURCE_CONFIDENCE.get(src)}
         for src, cnt in source_counts.items()],
        key=lambda x: (x["confidence"] is None, -(x["confidence"] or 0)),
    )
    return {"source_summary": source_summary}


def _flags_detail(op_rows, emb_rows) -> dict:
    """Parse flags; return severity counts, LCA category counts, top-flag list, and error rows."""
    severity_counts: dict = {"error": 0, "warning": 0, "info": 0}
    category_counts: dict = {"Completeness": 0, "Validity": 0, "Consistency": 0, "Info": 0}
    flag_agg: dict = {}
    error_rows: list = []
    for row in list(op_rows) + list(emb_rows):
        f = str(row.get("flags", "") or "").strip()
        if not f:
            continue
        for part in f.split(";"):
            p = part.strip()
            if not p:
                continue
            sev = _flag_severity(p)
            cat = _lca_category(p)
            severity_counts[sev] += 1
            category_counts[cat]  = category_counts.get(cat, 0) + 1
            key = p.split("—")[0].split("(")[0].strip()
            if key not in flag_agg:
                flag_agg[key] = {"count": 0, "sev": sev, "cat": cat}
            flag_agg[key]["count"] += 1
            if sev == "error":
                cs_id   = str(row.get("cell_site_id", "") or "").strip() or "—"
                subtype = (
                    str(row.get("active_subtype")      or "").strip() or
                    str(row.get("passive_type")         or "").strip() or
                    str(row.get("infrastructure_type")  or "").strip() or "—"
                )
                error_rows.append({
                    "cs_id":   cs_id,
                    "schema":  str(row.get("schema", "") or "").strip(),
                    "subtype": subtype,
                    "flag":    p,
                })
    top_flags = sorted(
        [{"text": k, "count": v["count"], "sev": v["sev"], "cat": v["cat"]}
         for k, v in flag_agg.items()],
        key=lambda x: x["count"], reverse=True,
    )[:12]
    return {"severity": severity_counts, "category": category_counts,
            "top_flags": top_flags, "error_rows": error_rows}


def _grand(by_network_rows):
    emb = sum(_f(r.get("Embodied (kgCO2e/yr)"))    for r in by_network_rows)
    op  = sum(_f(r.get("Operational (kgCO2e/yr)")) for r in by_network_rows)
    eol = sum(_f(r.get("End-of-life (kgCO2e/yr)")) for r in by_network_rows)
    t   = emb + op + eol
    return {"total": round(t,1), "emb": round(emb,1), "emb_pct": _pct(emb,t),
            "op": round(op,1), "op_pct": _pct(op,t), "eol": round(eol,1), "eol_pct": _pct(eol,t)}


# ── Assemble chart data ───────────────────────────────────────────────────────

def build_data(summary, op_rows, emb_rows, cell_sites, sensitivity=None, uncertainty=None):
    by_schema_rows  = summary.get("By Component Type", [])
    by_network_rows = summary.get("By Network Type",   [])
    by_site_rows    = summary.get("By Cell Site",      [])

    site_type_lkp = {
        str(s.get("cell_site_id", "")).strip(): str(s.get("site_type", "") or "").strip().lower()
        for s in cell_sites
    }

    schema_counts = {}
    for r in emb_rows:
        sc = str(r.get("schema", "unknown") or "unknown").strip()
        schema_counts[sc] = schema_counts.get(sc, 0) + 1

    site_count = len(cell_sites) if cell_sites else len(by_site_rows)
    gt = _grand(by_network_rows)
    scopes, scope3_detail = _ghg_scopes(op_rows, emb_rows, cell_sites)
    scope_total = round(sum(scopes.values()), 1)

    schema_chart = {
        "labels": [r.get("Type", "") for r in by_schema_rows],
        "emb": [round(_f(r.get("Embodied (kgCO2e/yr)")), 1)    for r in by_schema_rows],
        "op":  [round(_f(r.get("Operational (kgCO2e/yr)")), 1) for r in by_schema_rows],
        "eol": [round(_f(r.get("End-of-life (kgCO2e/yr)")), 1) for r in by_schema_rows],
    }

    network_chart = {
        "labels": [r.get("Network", "") for r in by_network_rows],
        "emb": [round(_f(r.get("Embodied (kgCO2e/yr)")), 1)    for r in by_network_rows],
        "op":  [round(_f(r.get("Operational (kgCO2e/yr)")), 1) for r in by_network_rows],
        "eol": [round(_f(r.get("End-of-life (kgCO2e/yr)")), 1) for r in by_network_rows],
    }

    site_data = []
    for r in by_site_rows:
        cs_id = str(r.get("Cell Site", "") or "").strip()
        if not cs_id or cs_id == "None":
            continue
        st    = site_type_lkp.get(cs_id, "")
        site_data.append({
            "label":     cs_id,
            "network":   str(r.get("Network", "") or ""),
            "site_type": st,
            "emb":   round(_f(r.get("Embodied (kgCO2e/yr)")), 1),
            "op":    round(_f(r.get("Operational (kgCO2e/yr)")), 1),
            "eol":   round(_f(r.get("End-of-life (kgCO2e/yr)")), 1),
            "total": round(_f(r.get("Total (kgCO2e/yr)")), 1),
            "colour": _SITE_TYPE_COLOURS.get(st, _SITE_TYPE_COLOURS[""]),
        })
    site_data.sort(key=lambda x: x["total"], reverse=True)

    _SITE_CHART_CAP = 10
    _chart_sites = site_data[:_SITE_CHART_CAP]
    _unc_by_site = (uncertainty or {}).get("by_site", {})
    site_chart = {
        "labels":      [s["label"]     for s in _chart_sites],
        "totals":      [s["total"]     for s in _chart_sites],
        "colours":     [s["colour"]    for s in _chart_sites],
        "networks":    [s["network"]   for s in _chart_sites],
        "site_types":  [s["site_type"] for s in _chart_sites],
        "total_count": len(site_data),
        "sds":         [_unc_by_site.get(s["label"], {}).get("t_sd", 0) for s in _chart_sites],
        "emb":         [s["emb"]  for s in _chart_sites],
        "op":          [s["op"]   for s in _chart_sites],
        "eol":         [s["eol"]  for s in _chart_sites],
    }

    # Tree: per-network summary — keeps JSON compact regardless of site count
    _site_net_lkp = {str(r.get("Cell Site", "") or "").strip(): str(r.get("Network", "") or "")
                     for r in by_site_rows}
    _TREE_NET_NORM = {
        "metro": "metro", "Metro": "metro",
        "access": "Access", "Access": "Access",
        "backbone": "Backbone", "Backbone": "Backbone", "backcobe": "Backbone",
        "#N/A": "unknown", "NA": "unknown",
    }

    def _norm_net(raw: str) -> str:
        s = str(raw or "").strip()
        if not s or s in ("#N/A", "NA"):
            return "unknown"
        return _TREE_NET_NORM.get(s, s)

    # "3rd party site" has no confirmed site count; "No site" infra has no site at all
    _SYNTHETIC_NETS = {"3rd party site", "No site"}
    _net_agg: dict = {}
    _infra_no_site = 0  # infrastructure rows with null cell_site_id — tracked separately
    for r in emb_rows:
        cs_id  = str(r.get("cell_site_id", "") or "").strip()
        schema = str(r.get("schema", "") or "").strip()
        if not cs_id and schema == "infrastructure":
            _infra_no_site += 1
            continue
        net = _norm_net(_site_net_lkp.get(cs_id, "") or "unknown")
        if net not in _net_agg:
            _net_agg[net] = {"network": net, "_sites": set(),
                              "active": 0, "passive": 0, "infrastructure": 0}
        _net_agg[net]["_sites"].add(cs_id)
        if schema in ("active", "passive", "infrastructure"):
            _net_agg[net][schema] += 1
    for cs_id, net in _site_net_lkp.items():
        if not cs_id:
            continue  # skip empty cs_id placeholder
        net = _norm_net(net or "")
        if net not in _net_agg:
            _net_agg[net] = {"network": net, "_sites": set(),
                              "active": 0, "passive": 0, "infrastructure": 0}
        _net_agg[net]["_sites"].add(cs_id)
    if _infra_no_site:
        _net_agg["No site"] = {"network": "No site", "_sites": set(),
                                "active": 0, "passive": 0, "infrastructure": _infra_no_site}
    _NET_ORDER = {"Access": 0, "aggregation": 1, "metro": 2, "Backbone": 3, "core": 4,
                  "unknown": 5, "3rd party site": 6, "No site": 7}
    tree_networks = [
        {"network": k, "sites": len(v["_sites"]),
         "active": v["active"], "passive": v["passive"], "infrastructure": v["infrastructure"],
         "synthetic": k in _SYNTHETIC_NETS}
        for k, v in sorted(_net_agg.items(),
                           key=lambda kv: (_NET_ORDER.get(kv[0], 99), kv[0]))
    ]

    _ELEC_PS = {"electricity", "battery"}
    _elec_kwh = 0.0
    for r in op_rows:
        if str(r.get("power_source", "") or "").strip().lower() in _ELEC_PS:
            if r.get("power_path") != "site_measured":
                _elec_kwh += _f(r.get("annual_consumption"))
    for site in cell_sites:
        _elec_kwh += _f(site.get("measured_electricity"))
    electricity_kwh = round(_elec_kwh, 1)

    _MONTHS = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]
    _monthly = [0.0] * 12
    _has_monthly = False
    for site in cell_sites:
        for i, m in enumerate(_MONTHS):
            v = _f(site.get(f"measured_electricity_{m}"))
            if v > 0:
                _has_monthly = True
            _monthly[i] += v
    if _has_monthly:
        monthly_electricity = [round(v, 1) for v in _monthly]
    else:
        _even = round(electricity_kwh / 12, 1)
        monthly_electricity = [_even] * 12

    largest_net = max(by_network_rows, key=lambda r: _f(r.get("Total (kgCO2e/yr)")), default={})
    ln_name     = largest_net.get("Network", "unknown")
    ln_pct      = _pct(_f(largest_net.get("Total (kgCO2e/yr)")), gt["total"])
    top_site    = site_data[0] if site_data else {}
    ts_pct      = _pct(top_site.get("total", 0), gt["total"])

    flagged, total_rows, top_flags = _flags_summary(op_rows, emb_rows)

    _measured_sites = {
        str(s.get("cell_site_id", "") or "").strip()
        for s in cell_sites
        if _f(s.get("measured_electricity")) or _f(s.get("measured_fuel"))
    }
    enriched = _load_enriched_rows()
    data_quality = {
        "coverage":    _field_coverage(enriched, _measured_sites),
        "consistency": _consistency_check(enriched),
        "validity":    _validity_summary(enriched),
        "flags":       _flags_detail(op_rows, emb_rows),
    }

    return {
        "generated":     datetime.date.today().isoformat(),
        "inventory":     {"sites": site_count,
                          "active": schema_counts.get("active", 0),
                          "passive": schema_counts.get("passive", 0),
                          "infrastructure": schema_counts.get("infrastructure", 0)},
        "grand":         gt,
        "scopes":        scopes,
        "scope_total":   scope_total,
        "scope3_detail": scope3_detail,
        "schema_chart":  schema_chart,
        "network_chart": network_chart,
        "site_chart":    site_chart,
        "electricity_kwh":    electricity_kwh,
        "monthly_electricity": monthly_electricity,
        "tree":          {"networks": tree_networks},
        "top5":          site_data[:5],
        "flags":         {"flagged": flagged, "total": total_rows, "top": top_flags},
        "findings": [
            f"Embodied emissions account for {gt['emb_pct']}% of total annual emissions ({gt['emb']:,.1f} kgCO2e/yr).",
            f"The {ln_name} network is the largest contributor at {ln_pct}% of the network total.",
            f"Top site {top_site.get('label', '—')} accounts for {ts_pct}% of network total.",
        ],
        "data_quality": data_quality,
        "colours": {"emb": _C_EMB, "eol": _C_EOL, "op": _C_OP,
                    "s1": _C_S1, "s2": _C_S2, "s3": _C_S3,
                    "warn": _C_WARN,
                    "site_type": _SITE_TYPE_COLOURS},
        "sensitivity":  sensitivity or [],
        "uncertainty":  uncertainty or {},
    }


# ── HTML template ─────────────────────────────────────────────────────────────

_HTML = """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Telecom Emissions Report</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4/dist/chart.umd.min.js"></script>
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:#f0f4f5;color:#07011E}
.page{max-width:960px;margin:0 auto;padding:24px 20px 60px}
h1{font-size:1.5rem;font-weight:700;margin-bottom:4px;color:#07011E}
.subtitle{color:#5a6472;font-size:.9rem;margin-bottom:32px}
h2{font-size:1.15rem;font-weight:600;margin:40px 0 16px;padding-bottom:6px;border-bottom:2px solid #3805E3;color:#07011E}
h3{font-size:.95rem;font-weight:600;margin:0 0 12px;color:#07011E}
.card{background:#fff;border:1px solid #D5E5E6;border-radius:8px;padding:20px;margin-bottom:20px}
.kpi-row{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:20px}
.kpi{background:#fff;border:1px solid #D5E5E6;border-radius:8px;padding:14px 20px;flex:1 1 160px}
.kpi-label{font-size:.78rem;color:#5a6472;text-transform:uppercase;letter-spacing:.04em;margin-bottom:6px}
.kpi-value{font-size:1.5rem;font-weight:700;color:#07011E}
.kpi-sub{font-size:.78rem;color:#5a6472;margin-top:2px}
.kpi-sd{font-size:.75rem;color:#A80A02;margin-top:3px;font-style:italic}
.cov-row{display:flex;flex-wrap:wrap;gap:10px}
.cov-chip{background:#D5E5E6;border-radius:6px;padding:10px 16px;flex:1 1 120px;text-align:center}
.cov-chip .n{font-size:1.3rem;font-weight:700;color:#07011E}
.cov-chip .lbl{font-size:.78rem;color:#5a6472;margin-top:2px}
.scope-grid{display:flex;gap:32px;align-items:flex-start;flex-wrap:wrap}
.scope-donut{width:200px;flex-shrink:0}
.scope-table-wrap{flex:1;min-width:260px}
table{width:100%;border-collapse:collapse;font-size:.85rem}
th{background:#D5E5E6;text-align:left;padding:8px 10px;font-weight:600;border-bottom:2px solid #7A9CDC;color:#07011E}
td{padding:7px 10px;border-bottom:1px solid #e8eef0}
tr:hover td{background:#f5f8f9}
.num{text-align:right}
.badge{display:inline-block;border-radius:4px;padding:2px 7px;font-size:.75rem;font-weight:600;color:#fff;margin-right:4px}
.chart-wrap{position:relative}
.tree-wrap{overflow-x:auto}
.findings{list-style:none;padding:0}
.findings li{margin:8px 0;font-size:.9rem;padding-left:18px;position:relative}
.findings li::before{content:"\\25B8";position:absolute;left:0;color:#3805E3}
.flag-box{background:#fff5f5;border-left:4px solid #A80A02;border-radius:4px;padding:12px 16px;margin-top:14px;font-size:.85rem;color:#07011E}
.flag-box ul{margin-top:6px;padding-left:18px}
.site-legend{margin-top:10px;font-size:.8rem;color:#5a6472;display:flex;flex-wrap:wrap;gap:12px}
.site-legend span{display:inline-flex;align-items:center;gap:5px}
.swatch{width:12px;height:12px;border-radius:2px;display:inline-block}
.cov-pill{display:inline-block;border-radius:3px;padding:1px 8px;font-size:.82rem;font-weight:600}
.cov-full{background:#D5E5E6;color:#3805E3}
.cov-partial{background:#FEF3C7;color:#B45309}
.cov-none{background:#fff5f5;color:#A80A02}
.sev-row{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:16px}
.sev-card{flex:1 1 130px;border-radius:6px;padding:12px 16px;text-align:center;border:1px solid #D5E5E6}
.sev-card .n{font-size:1.4rem;font-weight:700}
.sev-card .lbl{font-size:.78rem;color:#5a6472;margin-top:2px}
.sev-card .desc{font-size:.72rem;color:#888;margin-top:3px}
.limits{padding-left:18px;font-size:.85rem;line-height:1.6}
.limits li{margin:4px 0}
.cat-badge{display:inline-block;border-radius:3px;padding:1px 6px;font-size:.72rem;font-weight:600;color:#fff;margin-left:4px}
</style>
</head>
<body>
<div class="page">

<h1>Telecom Network Emission Report</h1>
<p class="subtitle">Generated: GENERATED_DATE</p>

<h2>1. Scope</h2>

<div class="card">
  <h3>Inventory coverage</h3>
  <div class="cov-row" id="covChips"></div>
</div>

<div class="card">
  <h3>Network topology</h3>
  <p id="treeDesc" style="font-size:.82rem;color:#666;margin-bottom:14px"></p>
  <div class="tree-wrap" id="treeViz"></div>
</div>

<h2>2. Results</h2>

<div class="kpi-row" id="resultKpis"></div>

<div class="card">
  <h3>Electricity consumption</h3>
  <div id="elecHighlight" style="margin-bottom:16px"></div>
  <canvas id="elecLineChart" height="120"></canvas>
</div>

<div class="card">
  <h3>GHG scope mapping (kgCO2e/yr)</h3>
  <div class="scope-grid">
    <div class="scope-donut"><canvas id="scopeDonut"></canvas></div>
    <div class="scope-table-wrap"><table id="scopeTable"></table></div>
  </div>
</div>

<div class="card">
  <h3>By component category</h3>
  <canvas id="schemaChart" height="110"></canvas>
</div>

<div class="card">
  <h3>By network segment</h3>
  <canvas id="networkChart" height="110"></canvas>
</div>

<div class="card">
  <h3>By cell site</h3>
  <div class="chart-wrap" id="siteWrap"><canvas id="siteChart"></canvas></div>
</div>

<h2>3. Analysis</h2>

<div class="card">
  <h3>Top 5 emitters</h3>
  <table id="top5Table"></table>
</div>

<div class="card" id="sensitivityCard" style="display:none">
  <h3>Sensitivity analysis &mdash; top parameters (&plusmn;20% perturbation)</h3>
  <p style="font-size:.85rem;color:#5a6472;margin-bottom:16px">One-at-a-time analysis: each parameter is varied &plusmn;20% fleet-wide while all others are held constant. Bars show the resulting change in grand total emissions as a percentage-point deviation from the baseline (100%). Showing the 8 most sensitive of all parameters tested &mdash; sorted by total swing, most sensitive at top.</p>
  <div class="chart-wrap" id="sensitivityWrap"><canvas id="sensitivityChart"></canvas></div>
</div>

<div class="card" id="highVarCard" style="display:none">
  <h3>High-variability groups &mdash; largest relative standard deviation (&sigma;/&mu;)</h3>
  <p style="font-size:.85rem;color:#5a6472;margin-bottom:12px">Component types with the highest ratio of standard deviation to mean emission factor within the type group. A high relative &sigma;/&mu; may indicate data quality gaps or genuine design variation that warrants separate sub-typing.</p>
  <table id="highVarTable"></table>
</div>

<div class="card">
  <h3>Key findings</h3>
  <ul class="findings" id="findings"></ul>
  <div id="flagsBox"></div>
</div>

<h2>4. Data Quality</h2>

<div class="card">
  <h3>Completeness</h3>
  <p style="font-size:.85rem;color:#5a6472;margin-bottom:12px">Proportion of component rows where each key emission field carries a value. An empty field means the value was not provided and no enrichment source could fill it.</p>
  <table id="covTable"></table>
  <div style="border-top:1px solid #D5E5E6;margin:24px 0 16px"></div>
  <h3>Flags</h3>
  <div id="errBanner" style="display:none;background:#fff5f5;border:2px solid #A80A02;border-radius:6px;padding:14px 18px;margin-bottom:16px;font-size:.88rem;color:#A80A02"></div>
  <div class="sev-row" id="sevRow"></div>
  <table id="flagTable"></table>
  <div id="consistencyFlagsNote"></div>
</div>

<div class="card">
  <h3>Consistency &#8212; cross-item field coverage</h3>
  <div id="consistencyContent"></div>
</div>

<div class="card">
  <h3>Validity &#8212; enrichment source confidence</h3>
  <p style="font-size:.85rem;color:#5a6472;margin-bottom:12px">Confidence scores reflect how verifiable each enrichment source is. <em>Direct input</em> means the value was entered directly in the inventory spreadsheet &#8212; it is user-verified but not cross-checked against an external database.</p>
  <table id="validityTable"></table>
</div>

<div class="card">
  <h3>Limitations &amp; modelling assumptions</h3>
  <ul class="limits">
    <li>Operational electricity for active components without direct consumption data is estimated at 80% of rated load: <em>p_idle + 0.8 &times; (p_max &minus; p_idle)</em>.</li>
    <li>Installation and end-of-life emission factors must be user-supplied. Rows where these are absent contribute only partial values to the cradle-to-site total.</li>
    <li>Scope&nbsp;1 maintenance emissions are one-time figures (per maintenance event), not annualised over equipment lifetime.</li>
    <li>Only modelled components contribute to totals. Unmodelled infrastructure &#8212; hosting and accommodation sites, modem pools, maintenance travel &#8212; is excluded from all totals.</li>
    <li><em>Timeliness:</em> emission factor publication years are not tracked; temporal variation within the reporting period is not captured.</li>
  </ul>
</div>

</div>
<script>
const D = DATA_JSON_PLACEHOLDER;

const fmt = v => Number(v).toLocaleString('en-GB',{minimumFractionDigits:1,maximumFractionDigits:1});
const pct = (v,t) => t ? (100*v/t).toFixed(1)+'%' : '\\u2014';
function fmtRounded(valueKg, sigmaKg) {
  if (!sigmaKg || sigmaKg <= 0) return fmt(valueKg) + ' kgCO₂e/yr';
  const logSig = Math.floor(Math.log10(Math.abs(sigmaKg)));
  const prec = Math.pow(10, logSig);
  const rKg = Math.round(valueKg / prec) * prec;
  if (Math.abs(rKg) >= 1000) {
    const decimals = Math.max(0, -(logSig - 3));
    return (rKg / 1000).toLocaleString('en-GB', {minimumFractionDigits: decimals, maximumFractionDigits: decimals}) + ' tCO₂e/yr';
  }
  return Math.round(rKg).toLocaleString('en-GB') + ' kgCO₂e/yr';
}
function fmtSigmaVal(sigmaKg) {
  if (!sigmaKg || sigmaKg <= 0) return '';
  if (sigmaKg >= 1000) return '± ' + Math.round(sigmaKg / 1000).toLocaleString('en-GB') + ' tCO₂e/yr';
  return '± ' + Math.round(sigmaKg).toLocaleString('en-GB') + ' kgCO₂e/yr';
}

// 1a. Coverage chips
const inv = D.inventory;
[{n:inv.sites,lbl:'Cell sites'},{n:inv.active,lbl:'Active components'},
 {n:inv.passive,lbl:'Passive components'},{n:inv.infrastructure,lbl:'Infrastructure'}
].forEach(d => {
  const el = document.createElement('div'); el.className='cov-chip';
  el.innerHTML = `<div class="n">${d.n}</div><div class="lbl">${d.lbl}</div>`;
  document.getElementById('covChips').appendChild(el);
});

// 1b. Network topology tree — one SVG row per network segment
(function renderTree(treeData) {
  const SCHEMA_C = {network:'#3805E3', active:'#07011E', passive:'#7A9CDC', infrastructure:'#A80A02'};
  const nets = treeData.networks;
  const container = document.getElementById('treeViz');
  const desc = document.getElementById('treeDesc');

  const totalSites = nets.reduce((s, n) => s + n.sites, 0);
  if (desc) desc.textContent = 'Network segments with component totals. Coloured badges: site count (blue), active components (dark), passive (light blue), infrastructure (red).';

  const SVG_W = 680, ROW_H = 40, PAD_T = 20, PAD_B = 36;
  const LABEL_X = 210, BADGE_X = 218, BADGE_W = 70, BADGE_GAP = 6;
  const svgH = PAD_T + nets.length * ROW_H + PAD_B;

  let svg = `<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 ${SVG_W} ${svgH}" style="width:100%;max-width:${SVG_W}px;display:block">`;

  nets.forEach((n, i) => {
    const y = PAD_T + i * ROW_H + ROW_H / 2;
    const lbl = n.network.length > 22 ? n.network.slice(0, 20) + '\\u2026' : n.network;
    svg += `<text x="${LABEL_X}" y="${y + 5}" font-size="12" fill="#222" font-family="sans-serif" text-anchor="end" font-weight="700">${lbl}</text>`;
    let bx = BADGE_X;
    const siteVal = n.synthetic ? 'n/a' : n.sites.toLocaleString('en-GB');
    const badges = [
      {c: SCHEMA_C.network, val: siteVal,                                     lbl2: 'sites'},
      {c: SCHEMA_C.active,  val: n.active.toLocaleString('en-GB'),            lbl2: 'active'},
      {c: SCHEMA_C.passive, val: n.passive.toLocaleString('en-GB'),           lbl2: 'passive'},
      {c: SCHEMA_C.infrastructure, val: n.infrastructure.toLocaleString('en-GB'), lbl2: 'infra'},
    ];
    badges.forEach(b => {
      svg += `<rect x="${bx}" y="${y - 14}" width="${BADGE_W}" height="28" rx="5" fill="${b.c}" opacity="0.88"/>`;
      svg += `<text x="${bx + BADGE_W / 2}" y="${y - 1}" font-size="11" fill="#fff" font-family="sans-serif" text-anchor="middle" font-weight="700">${b.val}</text>`;
      svg += `<text x="${bx + BADGE_W / 2}" y="${y + 12}" font-size="9" fill="#fff" font-family="sans-serif" text-anchor="middle" opacity="0.85">${b.lbl2}</text>`;
      bx += BADGE_W + BADGE_GAP;
    });
  });

  // Legend row
  const legItems = [{k:'active',lbl:'Active comps'},{k:'passive',lbl:'Passive comps'},{k:'infrastructure',lbl:'Infrastructure'}];
  let lx = 8, ly = svgH - 14;
  legItems.forEach(({k, lbl}) => {
    svg += `<circle cx="${lx + 5}" cy="${ly}" r="5" fill="${SCHEMA_C[k]}"/>`;
    svg += `<text x="${lx + 13}" y="${ly + 4}" font-size="10" fill="#666" font-family="sans-serif">${lbl}</text>`;
    lx += 140;
  });

  svg += '</svg>';
  container.innerHTML = svg;
})(D.tree);

// 2a. KPIs (with ±1σ uncertainty if available)
const g = D.grand;
const unc = D.uncertainty && D.uncertainty.grand ? D.uncertainty.grand : {};
[{lbl:'Grand Total',
  val: unc.t_sd   ? fmtRounded(g.total, unc.t_sd)   : fmt(g.total),
  sub: unc.t_sd   ? ''                               : 'kgCO₂e / year',
  sd:  unc.t_sd},
 {lbl:'Embodied',
  val: g.emb_pct+'%',
  sub: unc.emb_sd ? fmtRounded(g.emb, unc.emb_sd)   : fmt(g.emb)+' kgCO₂e/yr',
  sd:  unc.emb_sd},
 {lbl:'Operational',
  val: g.op_pct+'%',
  sub: unc.op_sd  ? fmtRounded(g.op,  unc.op_sd)    : fmt(g.op)+' kgCO₂e/yr',
  sd:  unc.op_sd},
 {lbl:'End-of-life',
  val: g.eol_pct+'%',
  sub: unc.eol_sd ? fmtRounded(g.eol, unc.eol_sd)   : fmt(g.eol)+' kgCO₂e/yr',
  sd:  unc.eol_sd}
].forEach(d => {
  const el = document.createElement('div'); el.className='kpi';
  const sdLine = d.sd ? `<div class="kpi-sd">${fmtSigmaVal(d.sd)}</div>` : '';
  el.innerHTML=`<div class="kpi-label">${d.lbl}</div><div class="kpi-value">${d.val}</div>`
    + (d.sub ? `<div class="kpi-sub">${d.sub}</div>` : '') + sdLine;
  document.getElementById('resultKpis').appendChild(el);
});

// 2b. Electricity highlight + monthly distribution
(function(){
  const elec = D.electricity_kwh;
  const monthlyData = D.monthly_electricity;
  const isFlat = monthlyData.every(v => v === monthlyData[0]);
  const note = isFlat
    ? 'estimated annual electricity consumption &mdash; evenly distributed across 12 months (no monthly data provided)'
    : 'estimated annual electricity consumption &mdash; monthly breakdown from measured data';
  const elecSd = D.uncertainty && D.uncertainty.electricity_kwh_sd;
  let elecValStr, elecSdStr = '';
  if (elecSd && elecSd > 0) {
    const logS = Math.floor(Math.log10(elecSd));
    const pr = Math.pow(10, logS);
    elecValStr = (Math.round(elec / pr) * pr).toLocaleString('en-GB');
    elecSdStr = ' <span style="font-size:1rem;font-weight:400;color:#A80A02">&plusmn;&thinsp;'
              + (Math.round(elecSd / pr) * pr).toLocaleString('en-GB') + ' kWh</span>';
  } else {
    elecValStr = Math.round(elec).toLocaleString('en-GB');
  }
  document.getElementById('elecHighlight').innerHTML =
    '<div style="font-size:2rem;font-weight:700;color:' + D.colours.s2 + '">' +
    elecValStr + ' kWh' + elecSdStr + '</div>' +
    '<div style="font-size:.82rem;color:#5a6472;margin-top:4px">' + note + '</div>';
  const monthLabels = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
  try {
    new Chart(document.getElementById('elecLineChart'), {
      type: 'line',
      data: {
        labels: monthLabels,
        datasets: [{
          label: 'Electricity (kWh/month)',
          data: monthlyData,
          borderColor: D.colours.s2,
          backgroundColor: 'rgba(56,5,227,0.07)',
          tension: 0,
          pointRadius: 4,
          pointBackgroundColor: D.colours.s2,
          fill: true,
        }]
      },
      options: {
        plugins: {
          legend: { display: false },
          tooltip: { callbacks: { label: ctx => ' ' + ctx.raw.toLocaleString('en-GB') + ' kWh' } }
        },
        scales: {
          y: {
            beginAtZero: true,
            ticks: { callback: v => v.toLocaleString('en-GB') }
          }
        }
      }
    });
  } catch(e) {
    document.getElementById('elecLineChart').parentNode.innerHTML =
      '<p style="color:#A80A02;padding:12px">Chart error: '+e.message+'</p>';
  }
})();

// 2c. Scope donut (now in Section 2)
const sc = D.scopes;
new Chart(document.getElementById('scopeDonut'),{
  type:'doughnut',
  data:{
    labels:['Scope 1 (direct)','Scope 2 (electricity)','Scope 3 (indirect)'],
    datasets:[{data:[sc['1'],sc['2'],sc['3']],
      backgroundColor:[D.colours.s1,D.colours.s2,D.colours.s3],borderWidth:1}]
  },
  options:{plugins:{legend:{display:false},tooltip:{callbacks:{label:ctx=>` ${fmt(ctx.raw)} kgCO2e/yr`}}}}
});

const st = D.scope_total; const s3 = D.scope3_detail;
const fmtT = v => Math.round(v/1000).toLocaleString('en-GB');
document.getElementById('scopeTable').innerHTML =
  `<thead><tr><th>Scope</th><th class="num">tCO₂e/yr</th><th class="num">%</th></tr></thead><tbody>
   <tr><td><span class="badge" style="background:${D.colours.s1}">Scope 1</span>Direct — fuel, refrigerant &amp; maintenance</td>
       <td class="num">${fmtT(sc['1'])}</td><td class="num">${pct(sc['1'],st)}</td></tr>
   <tr><td><span class="badge" style="background:${D.colours.s2}">Scope 2</span>Purchased electricity</td>
       <td class="num">${fmtT(sc['2'])}</td><td class="num">${pct(sc['2'],st)}</td></tr>
   <tr><td><span class="badge" style="background:${D.colours.s3}">Scope 3</span>All indirect</td>
       <td class="num">${fmtT(sc['3'])}</td><td class="num">${pct(sc['3'],st)}</td></tr>
   <tr><td style="padding-left:28px;color:#666">&#8627; Upstream (manufacturing + installation)</td>
       <td class="num">${fmtT(s3.upstream)}</td><td class="num">${pct(s3.upstream,st)}</td></tr>
   <tr><td style="padding-left:28px;color:#666">&#8627; Downstream (end-of-life)</td>
       <td class="num">${fmtT(s3.downstream)}</td><td class="num">${pct(s3.downstream,st)}</td></tr>
   <tr style="font-weight:600;border-top:2px solid #d0d8e8">
       <td>Total</td><td class="num">${fmtT(st)}</td><td class="num">100%</td></tr>
   </tbody>`;

// Error-bar plugin — draws ±1σ whiskers on the right edge of horizontal stacked bars
const errorBarPlugin = {
  id: 'errorBar',
  afterDatasetsDraw(chart) {
    const ebOpts = chart.options.plugins && chart.options.plugins.errorBar;
    if (!ebOpts || !ebOpts.bars) return;
    const bars   = ebOpts.bars;
    const totals = ebOpts.totals;
    const ctx = chart.ctx;
    const xScale = chart.scales.x;
    const yScale = chart.scales.y;
    const CAP = 6;
    ctx.save();
    ctx.strokeStyle = '#555';
    ctx.lineWidth = 1.5;
    bars.forEach((sd, i) => {
      if (!sd) return;
      const xPx  = xScale.getPixelForValue(totals[i]);
      const sdPx = xScale.getPixelForValue(totals[i] + sd)
                 - xScale.getPixelForValue(totals[i]);
      const yPx = yScale.getPixelForValue(chart.data.labels[i]);
      ctx.beginPath(); ctx.moveTo(xPx - sdPx, yPx); ctx.lineTo(xPx + sdPx, yPx); ctx.stroke();
      ctx.beginPath(); ctx.moveTo(xPx - sdPx, yPx - CAP); ctx.lineTo(xPx - sdPx, yPx + CAP); ctx.stroke();
      ctx.beginPath(); ctx.moveTo(xPx + sdPx, yPx - CAP); ctx.lineTo(xPx + sdPx, yPx + CAP); ctx.stroke();
    });
    ctx.restore();
  }
};
Chart.register(errorBarPlugin);

// 2c & 2d. Stacked bar helper with optional error bars
function stackedBar(id, labels, emb, op, eol, sdByLabel) {
  const sds    = sdByLabel ? labels.map(l => (sdByLabel[l] || {}).t_sd || 0) : null;
  const totals = labels.map((_, i) => (emb[i]||0) + (op[i]||0) + (eol[i]||0));
  try {
    new Chart(document.getElementById(id), {
      type:'bar',
      data:{labels,datasets:[
        {label:'Cradle-to-site',data:emb,backgroundColor:D.colours.emb,stack:'s'},
        {label:'End-of-life',   data:eol,backgroundColor:D.colours.eol,stack:'s'},
        {label:'Operational',   data:op, backgroundColor:D.colours.op, stack:'s'},
      ]},
      options:{
        indexAxis:'y',
        plugins:{
          tooltip:{callbacks:{
            label: ctx => ` ${ctx.dataset.label}: ${fmt(ctx.raw)} kgCO₂e/yr`,
            footer: ctx => {
              if (!sds) return '';
              const sd = sds[ctx[0].dataIndex];
              return sd ? `1σ uncertainty: ± ${fmt(sd)} kgCO₂e/yr` : '';
            }
          }},
          errorBar: sds && sds.some(v=>v>0) ? {bars: sds, totals: totals} : false,
        },
        scales:{x:{stacked:true,ticks:{callback:v=>fmt(v)}},y:{stacked:true}}
      },
    });
  } catch(e) {
    document.getElementById(id).parentNode.innerHTML='<p style="color:#A80A02;padding:12px">Chart error: '+e.message+'</p>';
  }
}

const uncSch = (D.uncertainty && D.uncertainty.by_schema)  || {};
const uncNet = (D.uncertainty && D.uncertainty.by_network) || {};
const sch = D.schema_chart;  stackedBar('schemaChart', sch.labels, sch.emb, sch.op, sch.eol, uncSch);
const net = D.network_chart; stackedBar('networkChart', net.labels, net.emb, net.op, net.eol, uncNet);

// 2e. By cell site (stacked by component category) — capped to avoid canvas size limits
const stc = D.site_chart;
if(stc.total_count > stc.labels.length){
  const note = document.createElement('p');
  note.style.cssText='font-size:.82rem;color:#666;margin-bottom:8px';
  note.textContent = `Top ${stc.labels.length.toLocaleString('en-GB')} of ${stc.total_count.toLocaleString('en-GB')} sites by total emissions`;
  const wrap = document.getElementById('siteWrap');
  wrap.parentNode.insertBefore(note, wrap);
}
const siteH = Math.max(250, stc.labels.length * 36 + 70);
document.getElementById('siteWrap').style.height = siteH + 'px';
(function(){
  const siteSds    = stc.sds && stc.sds.some(v => v > 0) ? stc.sds : null;
  const siteTotals = stc.labels.map((_, i) => (stc.emb[i]||0) + (stc.op[i]||0) + (stc.eol[i]||0));
  try {
    new Chart(document.getElementById('siteChart'),{
      type:'bar',
      data:{labels:stc.labels,datasets:[
        {label:'Cradle-to-site',data:stc.emb,backgroundColor:D.colours.emb,stack:'s'},
        {label:'End-of-life',   data:stc.eol,backgroundColor:D.colours.eol,stack:'s'},
        {label:'Operational',   data:stc.op, backgroundColor:D.colours.op, stack:'s'},
      ]},
      options:{
        maintainAspectRatio:false,
        indexAxis:'y',
        plugins:{
          legend:{display:true,position:'bottom'},
          tooltip:{callbacks:{
            label:ctx=>` ${ctx.dataset.label}: ${fmt(ctx.raw)} kgCO₂e/yr`,
            footer:ctx=>{
              const i=ctx[0].dataIndex;
              const parts=['site_type: '+(stc.site_types[i]||'unknown')+'  network: '+(stc.networks[i]||'\u2014')];
              if(siteSds&&siteSds[i])parts.push('1σ: ± '+fmt(siteSds[i])+' kgCO₂e/yr');
              return parts;
            }
          }},
          errorBar: siteSds ? {bars:siteSds, totals:siteTotals} : false,
        },
        scales:{x:{stacked:true,ticks:{callback:v=>fmt(v)}},y:{stacked:true}}
      },
    });
  } catch(e){ document.getElementById('siteWrap').innerHTML='<p style="color:#A80A02;padding:12px">Chart error: '+e.message+'</p>'; }
})()
// 3a. Top 5 table
const grandT = D.grand.total;
document.getElementById('top5Table').innerHTML =
  `<thead><tr><th>Cell Site</th><th>Network</th><th class="num">Embodied</th>
   <th class="num">Operational</th><th class="num">End-of-life</th>
   <th class="num">Total</th><th class="num">% of total</th></tr></thead><tbody>` +
  D.top5.map(r=>`<tr><td>${r.label}</td><td>${r.network}</td>
   <td class="num">${fmt(r.emb)}</td><td class="num">${fmt(r.op)}</td>
   <td class="num">${fmt(r.eol)}</td><td class="num">${fmt(r.total)}</td>
   <td class="num">${pct(r.total,grandT)}</td></tr>`).join('') + '</tbody>';

// 3b. Sensitivity tornado chart
(function(){
  const sensAll = D.sensitivity;
  if (!sensAll || !sensAll.length) return;
  document.getElementById('sensitivityCard').style.display = '';
  const sens = sensAll.slice(0, 8);
  const labels   = sens.map(r => r.label);
  const lowDevs  = sens.map(r => parseFloat((r.low  - 100).toFixed(2)));
  const highDevs = sens.map(r => parseFloat((r.high - 100).toFixed(2)));
  const sensH = Math.max(260, sens.length * 52 + 90);
  document.getElementById('sensitivityWrap').style.height = sensH + 'px';
  try {
    new Chart(document.getElementById('sensitivityChart'), {
      type: 'bar',
      data: {
        labels: labels,
        datasets: [
          {
            label: '\\u221220% perturbation',
            data: lowDevs,
            backgroundColor: D.colours.emb,
            barThickness: 16,
          },
          {
            label: '+20% perturbation',
            data: highDevs,
            backgroundColor: D.colours.s1,
            barThickness: 16,
          },
        ],
      },
      options: {
        maintainAspectRatio: false,
        indexAxis: 'y',
        plugins: {
          legend: { display: true, position: 'top' },
          tooltip: {
            callbacks: {
              label: function(ctx) {
                const r = sens[ctx.dataIndex];
                const val = ctx.datasetIndex === 0 ? r.low : r.high;
                const dev = val - 100;
                return ' ' + ctx.dataset.label + ': ' + val.toFixed(1) + '% of baseline (' + (dev >= 0 ? '+' : '') + dev.toFixed(1) + ' pp)';
              }
            }
          },
        },
        scales: {
          x: {
            title: { display: true, text: '% of baseline' },
            ticks: { callback: function(v) { return (100 + v).toFixed(0) + '%'; } },
            grid: { color: function(ctx) { return ctx.tick.value === 0 ? '#07011E' : '#e8eef0'; }, lineWidth: function(ctx) { return ctx.tick.value === 0 ? 2 : 1; } },
          },
          y: { stacked: false },
        },
      },
    });
  } catch(e) {
    document.getElementById('sensitivityWrap').innerHTML = '<p style="color:#A80A02;padding:12px">Chart error: ' + e.message + '</p>';
  }
})();

// 3c. High-variability groups
(function(){
  const hv = D.uncertainty && D.uncertainty.high_variability;
  if (!hv || !hv.length) return;
  document.getElementById('highVarCard').style.display = '';
  let html = '<thead><tr><th>Schema</th><th>Type</th><th>Field</th>'
           + '<th class="num">n</th><th class="num">Mean</th><th class="num">SD</th>'
           + '<th class="num">σ/μ</th></tr></thead><tbody>';
  hv.forEach(function(r) {
    const colour = r.relative_sd_pct > 50 ? D.colours.s1 : r.relative_sd_pct > 20 ? D.colours.warn : '#5a6472';
    const bold   = r.relative_sd_pct > 20 ? 'font-weight:600;' : '';
    html += '<tr><td>' + r.schema + '</td><td>' + r.type + '</td><td>' + r.field + '</td>'
          + '<td class="num">' + r.n + '</td>'
          + '<td class="num">' + r.mean.toLocaleString('en-GB',{maximumFractionDigits:1}) + '</td>'
          + '<td class="num">' + r.sd.toLocaleString('en-GB',{maximumFractionDigits:1}) + '</td>'
          + '<td class="num" style="color:' + colour + ';' + bold + '">' + r.relative_sd_pct.toFixed(1) + '%</td>'
          + '</tr>';
  });
  html += '</tbody>';
  document.getElementById('highVarTable').innerHTML = html;
})();

// 3d. Key findings
D.findings.forEach(f=>{
  const li=document.createElement('li'); li.textContent=f;
  document.getElementById('findings').appendChild(li);
});

const fl=D.flags;
if(fl.flagged>0){
  document.getElementById('flagsBox').innerHTML=
    `<div class="flag-box"><strong>Data flags:</strong> ${fl.flagged} of ${fl.total} component rows carry flags. See Section 4 for full breakdown.`+
    `</div>`;
}

// === Section 4: Data Quality ===
const DQ = D.data_quality;
const FIELD_LABELS_DQ = {
  production_emissions:         'Production emissions',
  endoflife_emissions:          'End-of-life emissions',
  power_source_emission_factor: 'Power source EF',
  installation_emission_factor: 'Installation EF',
  maintenance_emission_factor:  'Maintenance EF',
};

// 4a. Completeness — grouped coverage table
(function(){
  let html = '<thead><tr><th>Schema</th><th>Data category</th><th class="num">Present</th><th class="num">Total</th><th class="num">Coverage</th></tr></thead><tbody>';
  for (const [schema, groups] of Object.entries(DQ.coverage)) {
    let first = true;
    for (const g of groups) {
      const pct = g.total > 0 ? Math.round(100 * g.present / g.total) : null;
      const cls = pct === 100 ? 'cov-full' : pct === 0 ? 'cov-none' : 'cov-partial';
      const pctCell = pct === null
        ? '<em style="color:#888;font-size:.82rem">not required</em>'
        : '<span class="cov-pill ' + cls + '">' + pct + '%</span>';
      html += '<tr>'
            + '<td>' + (first ? '<strong>' + schema + '</strong>' : '') + '</td>'
            + '<td>' + g.label + '</td>'
            + '<td class="num">' + (g.total > 0 ? g.present : '&mdash;') + '</td>'
            + '<td class="num">' + (g.total > 0 ? g.total   : '&mdash;') + '</td>'
            + '<td class="num">' + pctCell + '</td>'
            + '</tr>';
      first = false;
    }
    html += '<tr><td colspan="5" style="padding:4px 0;border:none"></td></tr>';
  }
  html += '</tbody>';
  document.getElementById('covTable').innerHTML = html;
})();

// 4b. Consistency
(function(){
  const issues = DQ.consistency;
  const el = document.getElementById('consistencyContent');
  if (!issues || issues.length === 0) {
    el.innerHTML = '<p style="font-size:.88rem;color:#5a6472;margin-top:4px">No inconsistencies detected. All type groups with two or more rows show consistent field coverage across key emission fields.</p>';
    return;
  }
  let html = '<p style="font-size:.85rem;color:#5a6472;margin-bottom:12px">Fields present for some items of a type but absent for others &#8212; indicating inconsistent data collection within the type group.</p>'
           + '<table><thead><tr><th>Schema</th><th>Type</th><th>Field</th><th class="num">Present</th><th class="num">Total</th></tr></thead><tbody>';
  for (const iss of issues) {
    html += '<tr><td>' + iss.schema + '</td><td>' + iss.type + '</td>'
          + '<td>' + (FIELD_LABELS_DQ[iss.field] || iss.field) + '</td>'
          + '<td class="num">' + iss.present + '</td><td class="num">' + iss.total + '</td></tr>';
  }
  html += '</tbody></table>';
  el.innerHTML = html;
})();

// 4c. Validity — enrichment source confidence
(function(){
  const val = DQ.validity;
  let html = '<thead><tr><th>Source</th><th class="num">Confidence score</th><th class="num">Values filled</th></tr></thead><tbody>';
  for (const s of val.source_summary) {
    const conf = (s.confidence !== null && s.confidence !== undefined)
                 ? s.confidence.toFixed(2) : '\\u2014';
    const colour = s.source === 'Direct input' ? '#07011E' : D.colours.s2;
    html += '<tr>'
          + '<td><span class="badge" style="background:' + colour + '">' + s.source + '</span></td>'
          + '<td class="num">' + conf + '</td>'
          + '<td class="num">' + s.count + '</td></tr>';
  }
  html += '</tbody>';
  document.getElementById('validityTable').innerHTML = html;
})();

// 4d. Flags — severity only; consistency note shown when relevant
(function(){
  const flDQ = DQ.flags;
  const SEV_C = {error: D.colours.s1, warning: D.colours.warn, info: D.colours.s3};
  const SEV_LBL = {error: 'Error', warning: 'Warning', info: 'Info'};

  const sevEl = document.getElementById('sevRow');
  const errCount = flDQ.severity['error'];
  if (errCount > 0) {
    const banner = document.getElementById('errBanner');
    banner.style.display = 'block';
    const errRows = flDQ.error_rows || [];
    let bHtml = '<strong>&#9888; ' + errCount + ' computation error' + (errCount > 1 ? 's' : '') + ' detected</strong>'
      + ' &mdash; emission totals for the affected rows could not be calculated and are zero.'
      + ' Resolve these before relying on the results.';
    if (errRows.length > 0) {
      bHtml += '<table style="margin-top:10px;font-size:.82rem;border-collapse:collapse;width:100%;color:#07011E">'
             + '<thead><tr>'
             + '<th style="text-align:left;padding:5px 10px;border-bottom:1px solid #f0b0b0;background:#fef2f2">Cell site ID</th>'
             + '<th style="text-align:left;padding:5px 10px;border-bottom:1px solid #f0b0b0;background:#fef2f2">Schema</th>'
             + '<th style="text-align:left;padding:5px 10px;border-bottom:1px solid #f0b0b0;background:#fef2f2">Subtype</th>'
             + '<th style="text-align:left;padding:5px 10px;border-bottom:1px solid #f0b0b0;background:#fef2f2">Issue</th>'
             + '</tr></thead><tbody>';
      errRows.forEach(function(r, i) {
        const bg = i % 2 === 0 ? '#fff' : '#fef9f9';
        bHtml += '<tr style="background:' + bg + '">'
               + '<td style="padding:4px 10px;border-bottom:1px solid #fde8e8">' + r.cs_id   + '</td>'
               + '<td style="padding:4px 10px;border-bottom:1px solid #fde8e8">' + r.schema  + '</td>'
               + '<td style="padding:4px 10px;border-bottom:1px solid #fde8e8">' + r.subtype + '</td>'
               + '<td style="padding:4px 10px;border-bottom:1px solid #fde8e8">' + r.flag    + '</td>'
               + '</tr>';
      });
      bHtml += '</tbody></table>';
    }
    banner.innerHTML = bHtml;
  }
  [{k:'error', lbl:'Errors', desc:'Computation failed entirely'},
   {k:'warning', lbl:'Warnings', desc:'Partial or estimated data used'},
   {k:'info', lbl:'Informational', desc:'Normal processing notes'},
  ].forEach(function(d) {
    const el = document.createElement('div'); el.className = 'sev-card';
    const n = flDQ.severity[d.k];
    if (d.k === 'error' && n > 0) {
      el.style.cssText = 'background:#fff5f5;border:2px solid #A80A02;border-radius:6px;padding:12px 16px;text-align:center;flex:0 0 auto;min-width:160px';
    }
    el.innerHTML = '<div class="n" style="color:' + SEV_C[d.k] + (d.k === 'error' && n > 0 ? ';font-size:2rem' : '') + '">' + n + '</div>'
                 + '<div class="lbl"' + (d.k === 'error' && n > 0 ? ' style="font-weight:700;color:#A80A02"' : '') + '>' + d.lbl + '</div>'
                 + '<div class="desc">' + d.desc + '</div>';
    sevEl.appendChild(el);
  });

  document.getElementById('flagTable').innerHTML =
    '<thead><tr><th>Flag message</th><th class="num">Count</th><th>Severity</th></tr></thead><tbody>'
    + flDQ.top_flags.map(function(f) {
        return '<tr><td>' + f.text + '</td>'
             + '<td class="num">' + f.count + '</td>'
             + '<td><span class="badge" style="background:' + SEV_C[f.sev] + '">' + SEV_LBL[f.sev] + '</span></td></tr>';
      }).join('')
    + '</tbody>';

  const nConsistency = flDQ.category['Consistency'] || 0;
  if (nConsistency > 0) {
    document.getElementById('consistencyFlagsNote').innerHTML =
      '<div class="flag-box" style="margin-top:14px"><strong>Consistency flags (' + nConsistency + '):</strong> '
      + 'One or more rows have internal consistency issues: unit mismatches between quantity and emission factor, '
      + 'unrecognised unit strings, or power_max &lt; power_idle. Review the flagged rows directly.</div>';
  }
})();
</script>
</body>
</html>
"""


def generate_html(data: dict) -> str:
    html = _HTML.replace("GENERATED_DATE", data["generated"])
    html = html.replace("DATA_JSON_PLACEHOLDER", json.dumps(data, ensure_ascii=False))
    return html


def main():
    print("Loading data...")
    summary, op_rows, emb_rows, cell_sites, sensitivity, uncertainty = load_all()
    print(f"  Cell sites:       {len(cell_sites)}")
    print(f"  Embodied rows:    {len(emb_rows)}")
    print(f"  Operational rows: {len(op_rows)}")
    print(f"  Sensitivity params: {len(sensitivity)}")
    print(f"  Uncertainty loaded: {'yes' if uncertainty else 'no'}")

    data = build_data(summary, op_rows, emb_rows, cell_sites, sensitivity, uncertainty)
    html     = generate_html(data)
    out_path = DATA_DIR / "emissions_report.html"
    out_path.write_text(html, encoding="utf-8")
    print(f"  Written: {out_path}")

    g = data["grand"]
    print(f"\n  Grand total : {g['total']:>14,.1f} kgCO2e/yr")
    print(f"  Embodied    : {g['emb']:>14,.1f}  ({g['emb_pct']}%)")
    print(f"  Operational : {g['op']:>14,.1f}  ({g['op_pct']}%)")
    print(f"  End-of-life : {g['eol']:>14,.1f}  ({g['eol_pct']}%)")


if __name__ == "__main__":
    main()
