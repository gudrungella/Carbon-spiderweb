"""
analyze.py — Emission analysis orchestrator.

Loads enriched Excel files produced by the pipeline and runs two models:

  model_embodied    — annualised embodied + end-of-life emissions
  model_operational — operational emissions (electricity, fuel, refrigerant, maintenance)

Output:
  - Terminal report  (by component type, network type, cell site)
  - data/emissions_report.xlsx  (Summary + Embodied + Operational sheets)

Usage:
    python3 analyze.py
"""

import json
import math
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

import step4_enrich as _enrich
import step6_model_embodied as model_embodied
import step7_model_operational as model_operational
from step3_validate import normalise_network_type

DATA_DIR = Path(__file__).parent / "data"


def _f(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _build_site_op(cs_lookup: dict, ef_table: dict) -> dict:
    """For each site that has measured consumption, compute site-level op_energy per source type.

    Returns {cs_id: {"measured_electricity": {"qty", "ef", "op_energy"}, ...}}
    Only entries for source types where the measured field is present are included.
    """
    site_op = {}
    _sources = [
        ("measured_electricity", "electricity_emission_factor", "electricity_source"),
        ("measured_fuel",        "fuel_emission_factor",        "fuel_type"),
        ("measured_refrigerant", "refrigerant_emission_factor", "refrigerant_type"),
    ]
    for cs_id, site in cs_lookup.items():
        for measured_field, ef_field, source_type_field in _sources:
            qty = _f(site.get(measured_field))
            if qty is None:
                continue
            ef = _f(site.get(ef_field))
            if ef is None:
                source = str(site.get(source_type_field, "") or "").strip()
                for key in ([source, f"M-{source}"] if source else []):
                    result = (ef_table.get((key, "power_source_emission_factor"))
                              or ef_table.get((key, "")))
                    if result:
                        ef = _f(result.get("value"))
                        if ef is not None:
                            break
            op_energy = round(qty * ef, 4) if ef is not None else None
            site_op.setdefault(cs_id, {})[measured_field] = {
                "qty": qty, "ef": ef, "op_energy": op_energy,
            }

    # Apply site modifiers (decommissioned / per_rented) to site_op entries
    for cs_id, sources in site_op.items():
        site = cs_lookup.get(cs_id, {})
        _, factor = _site_scale_factor(site)
        for source_data in sources.values():
            if source_data["op_energy"] is not None:
                source_data["op_energy"] = round(source_data["op_energy"] * factor, 4)

    return site_op


# ---------------------------------------------------------------------------
# Site-level modifiers
# ---------------------------------------------------------------------------

_EMB_EMISSION_FIELDS = frozenset({
    "production_emissions_annual", "install_total", "install_total_annual",
    "cradle_to_site_annual", "eol_emissions_annual", "embodied_emissions_annual",
})
_OP_EMISSION_FIELDS = frozenset({
    "op_energy", "op_maintenance", "op_total",
})


def _site_scale_factor(site: dict):
    """Return (is_decommissioned, scale_factor) for a site dict.

    Decommissioned sites → factor 0.0.
    per_rented in (0, 100] → factor (1 − per_rented / 100).
    Otherwise → factor 1.0 (no change).
    """
    status = str(site.get("operational_status", "") or "").strip().lower()
    if status == "decommissioned":
        return True, 0.0

    raw = site.get("per_rented")
    if raw is not None and str(raw).strip() != "":
        try:
            per_rented = float(raw)
            if 0 < per_rented <= 100:
                return False, round(1.0 - per_rented / 100.0, 6)
        except (ValueError, TypeError):
            pass
    return False, 1.0


def _apply_site_modifiers(emb: dict, op: dict, site: dict) -> tuple:
    """Scale emission fields by ownership fraction; zero all if decommissioned.

    Returns new (emb, op) dicts — originals are not mutated.
    A note is appended to the flags field of each dict.
    """
    is_decom, factor = _site_scale_factor(site)
    if factor == 1.0:
        return emb, op

    def _scale(d, fields):
        result = dict(d)
        for f in fields:
            v = result.get(f)
            if v is not None:
                result[f] = round(v * factor, 4)
        return result

    adj_emb = _scale(emb, _EMB_EMISSION_FIELDS)
    adj_op  = _scale(op,  _OP_EMISSION_FIELDS)

    if is_decom:
        note = "site decommissioned — emissions zeroed"
    else:
        per_rented = round((1.0 - factor) * 100, 2)
        note = f"per_rented={per_rented}% — emissions scaled by {factor}"

    for d in (adj_emb, adj_op):
        existing = d.get("flags", "") or ""
        d["flags"] = f"{existing}; {note}".lstrip("; ")

    return adj_emb, adj_op


# ---------------------------------------------------------------------------
# Shared aggregation
# ---------------------------------------------------------------------------

def _aggregate(results: list, cs_lookup: dict, site_op: dict) -> tuple:
    """Aggregate emissions into by_schema, by_network, by_site dicts.

    Returns (by_schema, by_network, by_site).
    site_op contributions are added once per site to avoid double-counting
    when power_path = 'site_measured'.
    """
    by_schema  = defaultdict(lambda: {"emb": 0.0, "op": 0.0, "eol": 0.0, "t": 0.0, "n": 0})
    by_network = defaultdict(lambda: {"emb": 0.0, "op": 0.0, "eol": 0.0, "t": 0.0})
    by_site    = defaultdict(lambda: {"emb": 0.0, "op": 0.0, "eol": 0.0, "t": 0.0, "nt": ""})

    for schema, row, emb, op in results:
        cs_id   = str(row.get("cell_site_id", "")).strip()
        site    = cs_lookup.get(cs_id, {})
        _default_nt = "3rd party site" if cs_id not in cs_lookup else "unknown"
        _raw_nt = str(site.get("network_type") or row.get("network_type") or _default_nt).strip()
        nt      = normalise_network_type(_raw_nt) if _raw_nt not in ("3rd party site", "unknown") else _raw_nt
        emb_val = emb["cradle_to_site_annual"] or 0.0
        eol_val = emb["eol_emissions_annual"] or 0.0
        op_val  = (op["op_total"] or 0.0) if op["power_path"] != "site_measured" else 0.0
        total   = emb_val + op_val + eol_val
        for agg in (by_schema[schema], by_network[nt], by_site[cs_id]):
            agg["emb"] += emb_val
            agg["op"]  += op_val
            agg["eol"] += eol_val
            agg["t"]   += total
        by_schema[schema]["n"] += 1
        by_site[cs_id]["nt"] = nt

    for cs_id, sources in site_op.items():
        site    = cs_lookup.get(cs_id, {})
        _raw_nt = str(site.get("network_type") or "unknown").strip()
        nt      = normalise_network_type(_raw_nt) if _raw_nt != "unknown" else "unknown"
        for source_data in sources.values():
            op_val = source_data["op_energy"] or 0.0
            for agg in (by_network[nt], by_site[cs_id]):
                agg["op"] += op_val
                agg["t"]  += op_val

    return by_schema, by_network, by_site


# ---------------------------------------------------------------------------
# Sensitivity analysis
# ---------------------------------------------------------------------------

def _perturb_row(row: dict, field: str, factor: float) -> dict:
    v = row.get(field)
    if v is None:
        return row
    try:
        fv = float(v)
    except (ValueError, TypeError):
        return row
    result = dict(row)
    result[field] = round(fv * factor, 6)
    return result


def _compute_grand_total(cell_sites, active, passive, infra, ef_table) -> float:
    cs_lookup = {str(s.get("cell_site_id", "")).strip(): s for s in cell_sites}
    site_op   = _build_site_op(cs_lookup, ef_table)
    _na_op    = {f: None for f in model_operational.FIELDS}
    _na_op["power_path"] = "not_applicable"
    _na_op["flags"]      = "operational emissions not applicable for this component type"
    results = []
    for row in active:
        cs_id = str(row.get("cell_site_id", "")).strip()
        site  = cs_lookup.get(cs_id, {})
        emb   = model_embodied.compute(row)
        op    = model_operational.compute(row, "active", ef_table, site=site)
        emb, op = _apply_site_modifiers(emb, op, site)
        results.append(("active", row, emb, op))
    for row in passive:
        cs_id = str(row.get("cell_site_id", "")).strip()
        site  = cs_lookup.get(cs_id, {})
        emb   = model_embodied.compute(row)
        emb, _ = _apply_site_modifiers(emb, dict(_na_op), site)
        results.append(("passive", row, emb, dict(_na_op)))
    for row in infra:
        cs_id = str(row.get("cell_site_id", "")).strip()
        site  = cs_lookup.get(cs_id, {})
        emb   = model_embodied.compute(row)
        emb, _ = _apply_site_modifiers(emb, dict(_na_op), site)
        results.append(("infrastructure", row, emb, dict(_na_op)))
    by_schema, _, _ = _aggregate(results, cs_lookup, site_op)
    grand_t = sum(v["t"] for v in by_schema.values())
    for sources in site_op.values():
        for source_data in sources.values():
            grand_t += source_data["op_energy"] or 0.0
    return grand_t


def run_sensitivity(cell_sites, active, passive, infra, ef_table) -> list:
    """One-at-a-time (OAT) sensitivity: vary each parameter ±20% fleet-wide.

    Tests 15 parameters and returns all results sorted by swing (largest first).
    Each entry: {"label", "low" (% of baseline), "high" (% of baseline), "swing"}.
    """
    # Field sets used in composite perturbations
    _A_ALL = ["production_emissions", "endoflife_emissions",
              "power_quantity", "power_idle", "power_max", "maintenance_quantity"]
    _P_ALL = ["production_emissions", "endoflife_emissions", "installation_quantity"]
    _I_ALL = ["production_emissions", "endoflife_emissions", "installation_quantity"]
    _PWR   = ["power_quantity", "power_idle", "power_max"]

    # Each entry: (label, active_fields, passive_fields, infra_fields)
    _PARAMS = [
        ("Equipment lifetime",         ["life_time"],                      ["life_time"],                      ["life_time"]),
        ("Active component lifetime",  ["life_time"],                      [],                                 []),
        ("Infrastructure lifetime",    [],                                 [],                                 ["life_time"]),
        ("Production emissions EF",    ["production_emissions"],           ["production_emissions"],           ["production_emissions"]),
        ("Installation EF",            ["installation_emission_factor"],   ["installation_emission_factor"],   ["installation_emission_factor"]),
        ("Installation quantity",      ["installation_quantity"],          ["installation_quantity"],          ["installation_quantity"]),
        ("Grid emission factor",       ["power_source_emission_factor"],   [],                                 []),
        ("End-of-life emissions EF",   ["endoflife_emissions"],            ["endoflife_emissions"],            ["endoflife_emissions"]),
        ("Maintenance EF",             ["maintenance_emission_factor"],    [],                                 []),
        ("Maintenance frequency",      ["maintenance_quantity"],           [],                                 []),
        ("Network power draw",         _PWR,                              [],                                 []),
        ("Active component count",     _A_ALL,                            [],                                 []),
        ("Passive component count",    [],                                _P_ALL,                             []),
        ("Infrastructure count",       [],                                [],                                 _I_ALL),
        ("Fleet size",                 _A_ALL,                            _P_ALL,                             _I_ALL),
    ]

    baseline = _compute_grand_total(cell_sites, active, passive, infra, ef_table)
    if not baseline:
        return []

    def _perturb(rows, fields, factor):
        result = []
        for row in rows:
            r = row
            for f in fields:
                r = _perturb_row(r, f, factor)
            result.append(r)
        return result

    results = []
    for label, a_fields, p_fields, i_fields in _PARAMS:
        row_results = {}
        for factor in (0.8, 1.2):
            pa = _perturb(active,  a_fields, factor) if a_fields else active
            pp = _perturb(passive, p_fields, factor) if p_fields else passive
            pi = _perturb(infra,   i_fields, factor) if i_fields else infra
            total = _compute_grand_total(cell_sites, pa, pp, pi, ef_table)
            row_results[factor] = round(100.0 * total / baseline, 2)
        low   = row_results[0.8]
        high  = row_results[1.2]
        swing = round(abs(high - low), 2)
        results.append({"label": label, "low": low, "high": high, "swing": swing})

    results.sort(key=lambda x: x["swing"], reverse=True)
    return results


# ---------------------------------------------------------------------------
# Uncertainty analysis (1σ, sum-in-quadrature propagation)
# ---------------------------------------------------------------------------

def _fsd(v) -> float:
    """Safely read a standard-deviation field — returns 0.0 for None/missing/invalid."""
    if v is None:
        return 0.0
    try:
        return max(0.0, float(v))
    except (ValueError, TypeError):
        return 0.0


def _compute_row_uncertainty(row: dict, emb: dict, op: dict) -> dict:
    """Return 1σ uncertainty estimates for a single component row.

    Returns {"sigma_emb", "sigma_eol", "sigma_op", "sigma_total"} — all annualised kgCO2e/yr.
    Propagation rules:
      - Division by lifetime:      σ(x/n) = σ(x)/n
      - Product quantity × EF:     σ = quantity × σ(EF)      (quantity treated as exact)
      - Sum a + b:                 σ = √(σ_a² + σ_b²)
      - Power quantity path:       σ(op) = op_energy × (pq_sd / pq)   [relative]
      - Power idle/max path:       σ(op) = half-range between low/high energy estimates
    """
    lifetime = max(emb.get("lifetime_used") or 1.0, 1e-9)

    # ── Embodied (cradle-to-site) ─────────────────────────────────────────────
    sd_prod   = _fsd(row.get("production_emissions_sd"))
    sd_ef_ins = _fsd(row.get("installation_emission_factor_sd"))
    inst_qty  = _f(row.get("installation_quantity")) or 0.0

    sigma_prod_annual = sd_prod / lifetime
    sigma_inst_annual = sd_ef_ins * inst_qty / lifetime
    sigma_emb = math.sqrt(sigma_prod_annual ** 2 + sigma_inst_annual ** 2)

    # ── End-of-life ───────────────────────────────────────────────────────────
    sigma_eol = _fsd(row.get("endoflife_emissions_sd")) / lifetime

    # ── Operational ───────────────────────────────────────────────────────────
    sigma_op_e = 0.0
    power_path = op.get("power_path") or ""

    if power_path == "quantity":
        pq    = _f(row.get("power_quantity"))
        pq_sd = _fsd(row.get("power_quantity_sd"))
        op_e  = _f(op.get("op_energy")) or 0.0
        if pq and pq_sd:
            sigma_op_e = op_e * (pq_sd / pq)

    elif power_path == "estimated":
        low_w  = _f(row.get("power_estimated_low_w"))
        high_w = _f(row.get("power_estimated_high_w"))
        op_e   = _f(op.get("op_energy")) or 0.0
        if low_w is not None and high_w is not None:
            # Scale op_energy by low/high power ratios to get energy range
            p_idle = _f(row.get("power_idle")) or 0.0
            p_max  = _f(row.get("power_max"))  or 0.0
            if str(row.get("power_idle_unit", "") or "").strip().lower() == "kw":
                p_idle *= 1000.0
            if str(row.get("power_max_unit",  "") or "").strip().lower() == "kw":
                p_max  *= 1000.0
            p_central = p_idle + 0.8 * (p_max - p_idle)
            if p_central > 0:
                op_low_e  = op_e * (low_w  / p_central)
                op_high_e = op_e * (high_w / p_central)
                sigma_op_e = (op_high_e - op_low_e) / 2.0

    sd_maint = _fsd(row.get("maintenance_emission_factor_sd"))
    maint_qty = _f(row.get("maintenance_quantity")) or 0.0
    sigma_maint = sd_maint * maint_qty

    sigma_op = math.sqrt(sigma_op_e ** 2 + sigma_maint ** 2)

    sigma_total = math.sqrt(sigma_emb ** 2 + sigma_eol ** 2 + sigma_op ** 2)

    return {
        "sigma_emb":   round(sigma_emb,   4),
        "sigma_eol":   round(sigma_eol,   4),
        "sigma_op":    round(sigma_op,    4),
        "sigma_total": round(sigma_total, 4),
    }


def compute_uncertainty(results: list, cs_lookup: dict) -> dict:
    """Propagate row-level σ to grand total and aggregates via sum-in-quadrature.

    Returns a dict suitable for JSON export and step10 visualisation:
      {"grand": {t_sd, emb_sd, eol_sd, op_sd},
       "by_schema":  {schema:  {t_sd, ...}},
       "by_network": {network: {t_sd, ...}},
       "by_site":    {cs_id:   {t_sd, ...}},
       "electricity_kwh_sd": float}
    """
    def _zero():
        return {"var_emb": 0.0, "var_eol": 0.0, "var_op": 0.0, "var_t": 0.0}

    grand      = _zero()
    by_schema  = defaultdict(_zero)
    by_network = defaultdict(_zero)
    by_site    = defaultdict(_zero)
    sigma2_elec = 0.0

    for schema, row, emb, op in results:
        if op.get("power_path") == "site_measured":
            continue

        cs_id = str(row.get("cell_site_id", "")).strip()
        site  = cs_lookup.get(cs_id, {})
        _raw_nt = str(site.get("network_type") or row.get("network_type") or "unknown").strip()
        nt = normalise_network_type(_raw_nt) if _raw_nt not in ("3rd party site", "unknown") else _raw_nt

        u = _compute_row_uncertainty(row, emb, op)
        for agg in (grand, by_schema[schema], by_network[nt], by_site[cs_id]):
            agg["var_emb"] += u["sigma_emb"]   ** 2
            agg["var_eol"] += u["sigma_eol"]   ** 2
            agg["var_op"]  += u["sigma_op"]    ** 2
            agg["var_t"]   += u["sigma_total"] ** 2

        # Electricity-specific σ² for active electricity components
        if schema == "active":
            ps = str(row.get("power_source", "") or "").strip().lower()
            if ps in ("electricity", "battery"):
                power_path = op.get("power_path") or ""
                if power_path == "quantity":
                    pq    = _f(row.get("power_quantity"))
                    pq_sd = _fsd(row.get("power_quantity_sd"))
                    ac    = _f(op.get("annual_consumption")) or 0.0
                    if pq and pq_sd:
                        sigma2_elec += (ac * (pq_sd / pq)) ** 2
                elif power_path == "estimated":
                    low_w  = _f(row.get("power_estimated_low_w"))
                    high_w = _f(row.get("power_estimated_high_w"))
                    if low_w is not None and high_w is not None:
                        sigma2_elec += ((high_w - low_w) / 2 * 8760 / 1000) ** 2

    def _to_sd(d):
        return {k.replace("var_", "") + "_sd": round(math.sqrt(v), 1)
                for k, v in d.items()}

    return {
        "grand":             _to_sd(grand),
        "by_schema":         {k: _to_sd(v) for k, v in by_schema.items()},
        "by_network":        {k: _to_sd(v) for k, v in by_network.items()},
        "by_site":           {k: _to_sd(v) for k, v in by_site.items()},
        "electricity_kwh_sd": round(math.sqrt(sigma2_elec), 1),
    }


def compute_high_variability_groups(active: list, passive: list, infra: list,
                                     top_n: int = 15) -> list:
    """Identify (schema, type, field) groups with the highest relative SD (σ/μ).

    Returns top_n entries sorted by relative_sd descending.
    Each entry: {schema, type, field, n, mean, sd, relative_sd_pct}.
    Only groups with n≥2, sd>0, and mean>0 are included.
    """
    _FIELDS = {
        "active":         ["production_emissions", "endoflife_emissions",
                           "installation_emission_factor", "maintenance_emission_factor"],
        "passive":        ["production_emissions", "endoflife_emissions",
                           "installation_emission_factor"],
        "infrastructure": ["production_emissions", "endoflife_emissions",
                           "installation_emission_factor"],
    }
    _TYPE_FIELD = {
        "active": "active_subtype", "passive": "passive_type",
        "infrastructure": "infrastructure_type",
    }
    _FIELD_LABELS = {
        "production_emissions":         "Production emissions",
        "endoflife_emissions":          "End-of-life emissions",
        "installation_emission_factor": "Installation EF",
        "maintenance_emission_factor":  "Maintenance EF",
    }

    entries = []
    for schema, rows in [("active", active), ("passive", passive), ("infrastructure", infra)]:
        tf = _TYPE_FIELD[schema]
        groups: dict = {}
        for row in rows:
            t = str(row.get(tf, "") or "").strip() or "__unknown__"
            groups.setdefault(t, []).append(row)

        for type_name, grp in groups.items():
            n = len(grp)
            if n < 2:
                continue
            for field in _FIELDS[schema]:
                vals = [_f(r.get(field)) for r in grp]
                vals = [v for v in vals if v is not None]
                if not vals:
                    continue
                mean = sum(vals) / len(vals)
                if abs(mean) < 1e-9:
                    continue
                sd = _fsd(grp[0].get(f"{field}_sd"))
                if sd == 0:
                    continue
                entries.append({
                    "schema":          schema,
                    "type":            type_name,
                    "field":           _FIELD_LABELS.get(field, field),
                    "n":               n,
                    "mean":            round(mean, 2),
                    "sd":              round(sd, 2),
                    "relative_sd_pct": round(100.0 * sd / abs(mean), 1),
                })

    entries.sort(key=lambda x: x["relative_sd_pct"], reverse=True)
    return entries[:top_n]


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _load_xlsx(filename, data_sheet=False):
    path = DATA_DIR / filename
    if not path.exists():
        print(f"  [WARNING] {filename} not found — skipping.")
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Data"] if data_sheet and "Data" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = list(rows[0])
    return [
        dict(zip(headers, row))
        for row in rows[1:]
        if any(v is not None and str(v).strip() for v in row)
    ]


def load_data():
    cell_sites = _load_xlsx("cell_site.xlsx",   data_sheet=True)
    active     = _load_xlsx("active_components_enriched.xlsx")
    passive    = _load_xlsx("passive_components_enriched.xlsx")
    infra      = _load_xlsx("infrastructure_enriched.xlsx")
    return cell_sites, active, passive, infra


# ---------------------------------------------------------------------------
# Terminal report helpers
# ---------------------------------------------------------------------------

def _table(headers, rows, widths, indent="  "):
    sep = indent + "+" + "+".join("-" * (w + 2) for w in widths) + "+"
    fmt = indent + "|" + "|".join(f" {{:<{w}}} " for w in widths) + "|"
    print(sep)
    print(fmt.format(*[str(h) for h in headers]))
    print(sep)
    for row in rows:
        print(fmt.format(*[str(v) for v in row]))
    print(sep)


def _pct(part, total):
    return f"{100 * part / total:.0f}%" if total else "—"


# ---------------------------------------------------------------------------
# Terminal report
# ---------------------------------------------------------------------------

def print_report(results, cell_sites, site_op=None):
    """
    results: list of (schema, row, emb_dict, op_dict) for each component
    site_op: {cs_id: {measured_field: {op_energy, ...}}} for site-level measured emissions
    """
    site_op   = site_op or {}
    cs_lookup = {str(s.get("cell_site_id", "")).strip(): s for s in cell_sites}
    by_schema, by_network, by_site = _aggregate(results, cs_lookup, site_op)

    grand = {k: sum(v[k] for v in by_schema.values()) for k in ("emb", "op", "eol", "t")}
    for sources in site_op.values():
        for source_data in sources.values():
            grand["op"] += source_data["op_energy"] or 0.0
            grand["t"]  += source_data["op_energy"] or 0.0

    print("\n" + "=" * 76)
    print("  TELECOM NETWORK EMISSION ANALYSIS  (kgCO2e / year)")
    print("=" * 76)
    print(f"\n  Grand total: {grand['t']:>12,.1f} kgCO2e/year")
    print(f"    Embodied:    {grand['emb']:>12,.1f}  ({_pct(grand['emb'], grand['t'])})")
    print(f"    Operational: {grand['op']:>12,.1f}  ({_pct(grand['op'],  grand['t'])})")
    print(f"    End-of-life: {grand['eol']:>12,.1f}  ({_pct(grand['eol'], grand['t'])})")

    print(f"\n{'─' * 76}")
    print("  By component type")
    _table(
        ["Type", "Count", "Embodied", "Operational", "End-of-life", "Total"],
        [(s, v["n"], f"{v['emb']:,.1f}", f"{v['op']:,.1f}", f"{v['eol']:,.1f}", f"{v['t']:,.1f}")
         for s, v in sorted(by_schema.items())],
        [14, 6, 14, 14, 14, 14],
    )

    print(f"\n{'─' * 76}")
    print("  By network type")
    _table(
        ["Network", "Embodied", "Operational", "End-of-life", "Total"],
        [(nt, f"{v['emb']:,.1f}", f"{v['op']:,.1f}", f"{v['eol']:,.1f}", f"{v['t']:,.1f}")
         for nt, v in sorted(by_network.items())],
        [12, 14, 14, 14, 14],
    )

    print(f"\n{'─' * 76}")
    print("  By cell site")
    site_rows = []
    for cs_id, v in sorted(by_site.items()):
        site = cs_lookup.get(cs_id, {})
        nt   = str(site.get("network_type") or v["nt"]).strip()
        site_rows.append((cs_id, nt, f"{v['emb']:,.1f}", f"{v['op']:,.1f}", f"{v['eol']:,.1f}", f"{v['t']:,.1f}"))
    _table(
        ["Cell Site", "Network", "Embodied", "Operational", "End-of-life", "Total"],
        site_rows,
        [22, 10, 14, 14, 14, 14],
    )
    print()


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def write_excel(results, cell_sites, out_path, site_op=None):
    if not _OPENPYXL_AVAILABLE:
        print("  [WARNING] openpyxl not available — skipping Excel output.")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_font  = Font(bold=True)
    hdr_fill  = PatternFill("solid", fgColor="D9E1F2")
    emb_fill  = PatternFill("solid", fgColor="E2EFDA")   # green tint — embodied
    op_fill   = PatternFill("solid", fgColor="FCE4D6")   # orange tint — operational

    def _write_sheet(title, col_headers, data_rows, fill=None):
        ws = wb.create_sheet(title)
        for c, h in enumerate(col_headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = hdr_font
            cell.fill = fill or hdr_fill
        for r, row_vals in enumerate(data_rows, 2):
            for c, v in enumerate(row_vals, 1):
                ws.cell(r, c).value = v
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w + 2, 35)

    # ── Summary sheet ─────────────────────────────────────────────────────────
    site_op    = site_op or {}
    cs_lookup  = {str(s.get("cell_site_id", "")).strip(): s for s in cell_sites}
    by_schema, by_network, by_site = _aggregate(results, cs_lookup, site_op)

    ws_sum = wb.create_sheet("Summary", 0)
    ws_sum.cell(1, 1, "TELECOM EMISSIONS SUMMARY  (kgCO2e / year)").font = Font(bold=True, size=13)

    def _section(ws, start, title, hdrs, rows):
        ws.cell(start, 1, title).font = Font(bold=True, size=11)
        start += 1
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(start, c, h)
            cell.font = hdr_font
            cell.fill = hdr_fill
        start += 1
        for dr in rows:
            for c, v in enumerate(dr, 1):
                ws.cell(start, c).value = v
            start += 1
        return start + 1

    r = 3
    r = _section(ws_sum, r, "By Component Type",
        ["Type", "Count", "Embodied (kgCO2e/yr)", "Operational (kgCO2e/yr)", "End-of-life (kgCO2e/yr)", "Total (kgCO2e/yr)"],
        [(s, v["n"], round(v["emb"],2), round(v["op"],2), round(v["eol"],2), round(v["t"],2))
         for s, v in sorted(by_schema.items())])
    r = _section(ws_sum, r, "By Network Type",
        ["Network", "Embodied (kgCO2e/yr)", "Operational (kgCO2e/yr)", "End-of-life (kgCO2e/yr)", "Total (kgCO2e/yr)"],
        [(nt, round(v["emb"],2), round(v["op"],2), round(v["eol"],2), round(v["t"],2))
         for nt, v in sorted(by_network.items())])
    r = _section(ws_sum, r, "By Cell Site",
        ["Cell Site", "Network", "Embodied (kgCO2e/yr)", "Operational (kgCO2e/yr)", "End-of-life (kgCO2e/yr)", "Total (kgCO2e/yr)"],
        [(cs_id, v["nt"], round(v["emb"],2), round(v["op"],2), round(v["eol"],2), round(v["t"],2))
         for cs_id, v in sorted(by_site.items())])
    for col in ws_sum.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws_sum.column_dimensions[col[0].column_letter].width = min(w + 2, 40)

    # ── Embodied sheet ────────────────────────────────────────────────────────
    emb_hdrs = ["cell_site_id", "schema", "active_subtype", "passive_type",
                "infrastructure_type", "life_time",
                "production_emissions", "installation_quantity", "installation_emission_factor",
                "endoflife_emissions"] + model_embodied.FIELDS
    emb_rows = []
    for schema, row, emb, _ in results:
        emb_rows.append([
            row.get("cell_site_id"), schema,
            row.get("active_subtype"), row.get("passive_type"), row.get("infrastructure_type"),
            row.get("life_time"),
            row.get("production_emissions"), row.get("installation_quantity"),
            row.get("installation_emission_factor"), row.get("endoflife_emissions"),
        ] + [emb.get(f) for f in model_embodied.FIELDS])
    _write_sheet("Embodied", emb_hdrs, emb_rows, fill=emb_fill)

    # ── Operational sheet ─────────────────────────────────────────────────────
    op_hdrs = ["cell_site_id", "schema", "active_subtype", "power_source",
               "power_quantity", "power_unit",
               "power_idle", "power_max",
               "power_source_emission_factor", "power_source_emission_factor_unit",
               "maintenance_quantity", "maintenance_emission_factor"] + model_operational.FIELDS
    op_rows = []
    for schema, row, _, op in results:
        op_field_vals = []
        for f in model_operational.FIELDS:
            v = op.get(f)
            # Replace None op_energy with a note when emissions are captured at site level
            if f == "op_energy" and v is None and op.get("power_path") == "site_measured":
                v = "→ see site total"
            op_field_vals.append(v)
        op_rows.append([
            row.get("cell_site_id"), schema,
            row.get("active_subtype"), row.get("power_source"),
            row.get("power_quantity"), row.get("power_unit"),
            row.get("power_idle"), row.get("power_max"),
            row.get("power_source_emission_factor"), row.get("power_source_emission_factor_unit"),
            row.get("maintenance_quantity"), row.get("maintenance_emission_factor"),
        ] + op_field_vals)
    _write_sheet("Operational", op_hdrs, op_rows, fill=op_fill)

    wb.save(out_path)
    print(f"  Excel report written: {out_path.name}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not _OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl is required.  pip install openpyxl")
        return

    print("=" * 76)
    print("  LOADING DATA")
    print("=" * 76)
    cell_sites, active, passive, infra = load_data()
    print(f"  Cell sites:     {len(cell_sites)}")
    print(f"  Active rows:    {len(active)}")
    print(f"  Passive rows:   {len(passive)}")
    print(f"  Infra rows:     {len(infra)}")

    ef_table = _enrich._load_ef_table()
    cs_lookup_main = {str(s.get("cell_site_id", "")).strip(): s for s in cell_sites}
    site_op  = _build_site_op(cs_lookup_main, ef_table)

    _na_op = {f: None for f in model_operational.FIELDS}
    _na_op["power_path"] = "not_applicable"
    _na_op["flags"]      = "operational emissions not applicable for this component type"

    results = []
    for row in active:
        cs_id = str(row.get("cell_site_id", "")).strip()
        site  = cs_lookup_main.get(cs_id, {})
        emb   = model_embodied.compute(row)
        op    = model_operational.compute(row, "active", ef_table, site=site)
        emb, op = _apply_site_modifiers(emb, op, site)
        results.append(("active", row, emb, op))
    for row in passive:
        cs_id = str(row.get("cell_site_id", "")).strip()
        site  = cs_lookup_main.get(cs_id, {})
        emb   = model_embodied.compute(row)
        emb, _ = _apply_site_modifiers(emb, dict(_na_op), site)
        results.append(("passive", row, emb, dict(_na_op)))
    for row in infra:
        cs_id = str(row.get("cell_site_id", "")).strip()
        site  = cs_lookup_main.get(cs_id, {})
        emb   = model_embodied.compute(row)
        emb, _ = _apply_site_modifiers(emb, dict(_na_op), site)
        results.append(("infrastructure", row, emb, dict(_na_op)))

    print_report(results, cell_sites, site_op=site_op)

    out_path = DATA_DIR / "emissions_report.xlsx"
    write_excel(results, cell_sites, out_path, site_op=site_op)

    print("\n" + "=" * 76)
    print("  UNCERTAINTY ANALYSIS  (1σ, sum-in-quadrature propagation)")
    print("=" * 76)
    uncertainty = compute_uncertainty(results, cs_lookup_main)
    g = uncertainty["grand"]
    grand_t = sum(g.get(k, 0) for k in ("emb_sd", "eol_sd", "op_sd"))  # approx display only
    print(f"  Grand total 1σ: ±{g['t_sd']:>14,.1f} kgCO2e/yr")
    print(f"    Embodied:     ±{g['emb_sd']:>14,.1f}")
    print(f"    Operational:  ±{g['op_sd']:>14,.1f}")
    print(f"    End-of-life:  ±{g['eol_sd']:>14,.1f}")
    uncertainty["high_variability"] = compute_high_variability_groups(active, passive, infra)
    unc_path = DATA_DIR / "uncertainty_results.json"
    with open(unc_path, "w", encoding="utf-8") as fh:
        json.dump(uncertainty, fh, indent=2)
    print(f"\n  Saved: {unc_path.name}")

    print("\n" + "=" * 76)
    print("  SENSITIVITY ANALYSIS  (±20% one-at-a-time, impact on grand total)")
    print("=" * 76)
    sensitivity = run_sensitivity(cell_sites, active, passive, infra, ef_table)
    if sensitivity:
        print(f"  {'#':<3} {'Parameter':<35} {'−20%':>8}  {'+20%':>8}  {'Swing':>7}")
        print(f"  {'-'*3}  {'-'*35}  {'-'*8}  {'-'*8}  {'-'*7}")
        for i, r in enumerate(sensitivity, 1):
            print(f"  {i:<3} {r['label']:<35} {r['low']:>7.1f}%  {r['high']:>7.1f}%  {r['swing']:>6.1f}pp")
        sens_path = DATA_DIR / "sensitivity_results.json"
        with open(sens_path, "w", encoding="utf-8") as fh:
            json.dump(sensitivity, fh, indent=2)
        print(f"\n  {len(sensitivity)} parameters tested. Saved: {sens_path.name}")
    else:
        print("  No results — baseline is zero or no data.")
    print()


if __name__ == "__main__":
    main()
